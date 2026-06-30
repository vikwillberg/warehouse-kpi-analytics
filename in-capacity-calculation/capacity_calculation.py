"""
Site 1 Warehouse Capacity Calculation
Reads Capacity Calculation.xlsx (location specs & capacity data),
502 (current stock), 701 (fallback part data: PCS/box & weight/pc),
and Site 1 Parts List (descriptions)
to generate an Excel workbook with capacity analysis.
"""

import shutil
import re
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "Data"
OUTPUT_FILE = BASE_DIR / "Output" / f"IN_Capacity_Calculation_{datetime.now().strftime('%Y%m%d')}.xlsx"

# Designated overflow buckets — stock here has no rack-capacity row but still
# counts toward per-part inventory. Matched on the first 5 chars of LOCATION
# so suffixed codes like OVFLO01000 / ZZZZZ00000 / H1HLD01000 are caught.
OVERFLOW_LOCATIONS = {'ECQPC', 'OVFLO', 'ZZZZZ', 'H1HLD', 'XXPIL'}

# ---------------------------------------------------------------------------
# 1. Load & clean data
# ---------------------------------------------------------------------------
print("Loading data files...")

# --- Capacity Calculation: Primary reference for locations & capacity specs ---
try:
    cap_ref = pd.read_excel(DATA_DIR / "Capacity Calculation.xlsx",
                             sheet_name='Capacity Calculation')
except PermissionError:
    _temp = Path.home() / "AppData" / "Local" / "Temp" / "CapCalc_temp.xlsx"
    shutil.copy2(DATA_DIR / "Capacity Calculation.xlsx", _temp)
    cap_ref = pd.read_excel(_temp, sheet_name='Capacity Calculation')

cap_ref.columns = [str(c).strip() for c in cap_ref.columns]
cap_ref = cap_ref.rename(columns={
    'PRODUCT CODE': 'PRODUCT_CODE',
    'Weight': 'WEIGHT_STR',
    'Packaging Description': 'PKG_DESCRIPTION',
    'QTY per box': 'PCS_MODULE',
    'BOX TYPE': 'MODULE_TYPE',
    'BOX STACK HEIGHT': 'STACK_HEIGHT',
    'BOX PER LAYER': 'MODULES_PER_LAYER',
    'PALLETS PER WINDOW': 'PALLETS_PER_WIN',
    'WINDOWS': 'LOC_WINDOWS',
})

# Clean strings
cap_ref['PRODUCT_CODE'] = cap_ref['PRODUCT_CODE'].astype(str).str.strip()
cap_ref['LOCATION'] = cap_ref['LOCATION'].astype(str).str.strip()
cap_ref['UNIT'] = cap_ref['UNIT'].astype(str).str.strip()
cap_ref['RACK'] = cap_ref['RACK'].astype(str).str.strip()
cap_ref['MODULE_TYPE'] = cap_ref['MODULE_TYPE'].astype(str).str.strip()
cap_ref['PKG_DESCRIPTION'] = cap_ref['PKG_DESCRIPTION'].fillna('').astype(str).str.strip()
cap_ref['WEIGHT_STR'] = cap_ref['WEIGHT_STR'].fillna('').astype(str).str.strip()

# Parse weight string to numeric lbs (e.g., "29 lbs" → 29.0)
def parse_weight_lbs(w):
    if not w or w in ('nan', 'None'):
        return 0.0
    m = re.search(r'[\d]+(?:\.\d+)?', w.replace(',', ''))
    return float(m.group()) if m else 0.0

cap_ref['WEIGHT_LBS'] = cap_ref['WEIGHT_STR'].apply(parse_weight_lbs)

# Numeric fields
cap_ref['PCS_MODULE'] = pd.to_numeric(cap_ref['PCS_MODULE'], errors='coerce').fillna(0).astype(int)
cap_ref['STACK_HEIGHT'] = pd.to_numeric(cap_ref['STACK_HEIGHT'], errors='coerce').fillna(1).astype(int)
cap_ref['MODULES_PER_LAYER'] = pd.to_numeric(cap_ref['MODULES_PER_LAYER'], errors='coerce').fillna(1).astype(int)
cap_ref['PALLETS_PER_WIN'] = pd.to_numeric(cap_ref['PALLETS_PER_WIN'], errors='coerce').fillna(1).astype(int)
cap_ref['LOC_WINDOWS'] = pd.to_numeric(cap_ref['LOC_WINDOWS'], errors='coerce').fillna(1).astype(int)

# Derive LEVEL from LOCATION (last 2 characters, always "00")
cap_ref['LEVEL'] = cap_ref['LOCATION'].str[-2:]

print(f"  Capacity Calculation: {len(cap_ref)} location-part entries")

# --- 701: Fallback part master data (PCS/Module where Excel is missing, weight per piece) ---
# WMS report 701 = item/part master (MODEL/COLOR CODE, PCS/BOX, WEIGHT PER PCS).
# 702 is the location/slot master and does NOT carry these columns.
df702_raw = pd.read_csv(DATA_DIR / "701.CSV", dtype=str)
df702_raw.columns = [c.strip().strip('"') for c in df702_raw.columns]
for c in df702_raw.columns:
    df702_raw[c] = df702_raw[c].astype(str).str.strip().str.strip('"')

df702 = df702_raw[['MODEL/COLOR CODE', 'PCS/BOX', 'WEIGHT PER PCS']].copy()
df702.columns = ['PRODUCT_CODE', 'PCS_MODULE_702', 'WEIGHT_PCS']
df702['PRODUCT_CODE'] = df702['PRODUCT_CODE'].str.strip()
df702['PCS_MODULE_702'] = pd.to_numeric(df702['PCS_MODULE_702'], errors='coerce').fillna(0).astype(int)
df702['WEIGHT_PCS'] = pd.to_numeric(df702['WEIGHT_PCS'], errors='coerce').fillna(0)
print(f"  701: {len(df702)} part master records (fallback)")

# --- 502: Current stock ---
df502 = pd.read_csv(DATA_DIR / "502.csv", dtype=str)
df502.columns = [c.strip().strip('"') for c in df502.columns]
for c in df502.columns:
    df502[c] = df502[c].astype(str).str.strip().str.strip('"')
df502['QUANTITY'] = pd.to_numeric(df502['QUANTITY'], errors='coerce').fillna(0).astype(int)
df502['COMM PRODUCT'] = df502['COMM PRODUCT'].str.strip()
df502['LOCATION'] = df502['LOCATION'].str.strip()
# Convert 502 LOCATION from dashed format to match Excel (e.g., "M1-G06-100-00" → "M1G0610000")
df502['LOCATION'] = df502['LOCATION'].str.replace('-', '', regex=False)
print(f"  502: {len(df502)} current stock rows")

# --- Site 1 Parts List ---
try:
    df_parts = pd.read_excel(DATA_DIR / "Site 1 Parts List.xlsx",
                              sheet_name='Parts List_IND', header=3)
except PermissionError:
    _temp = Path.home() / "AppData" / "Local" / "Temp" / "Site 1Parts_temp.xlsx"
    shutil.copy2(DATA_DIR / "Site 1 Parts List.xlsx", _temp)
    df_parts = pd.read_excel(_temp, sheet_name='Parts List_IND', header=3)

df_parts.columns = [str(c).strip() for c in df_parts.columns]
print(f"  Site 1 Parts List: {len(df_parts)} rows")

# ---------------------------------------------------------------------------
# 2. Build capacity table
# ---------------------------------------------------------------------------
print("\nBuilding capacity table...")

capacity = cap_ref.copy()

# Create clean part code for matching (strip all whitespace and hyphens)
capacity['PART_CLEAN'] = capacity['PRODUCT_CODE'].str.replace(r'[\s\-]+', '', regex=True)

# Merge 702 fallback data (for PCS_MODULE where Excel is 0, and WEIGHT_PCS)
df702['PART_CLEAN'] = df702['PRODUCT_CODE'].str.replace(r'[\s\-]+', '', regex=True)
capacity = capacity.merge(
    df702[['PART_CLEAN', 'PCS_MODULE_702', 'WEIGHT_PCS']].drop_duplicates('PART_CLEAN'),
    on='PART_CLEAN', how='left'
)
capacity['PCS_MODULE_702'] = capacity['PCS_MODULE_702'].fillna(0).astype(int)
capacity['WEIGHT_PCS'] = capacity['WEIGHT_PCS'].fillna(0)

# Use Excel PCS_MODULE where available, fall back to 702
capacity['PCS_FINAL'] = np.where(
    capacity['PCS_MODULE'] > 0,
    capacity['PCS_MODULE'],
    capacity['PCS_MODULE_702']
).astype(int)

# Get description & area from Site 1 Parts List
def _clean_cell(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ''
    return str(v).strip()

parts_lookup = {}
if 'Supplier Part No.' in df_parts.columns:
    for _, row in df_parts.iterrows():
        part = re.sub(r'[\s\-]+', '', _clean_cell(row.get('Supplier Part No.')))
        if part and part.lower() != 'nan':
            parts_lookup[part] = {
                'description': _clean_cell(row.get('Description')),
                'sourcing': _clean_cell(row.get('Sourcing')),
                'area': _clean_cell(row.get('Area')),
            }

capacity['DESCRIPTION'] = capacity['PART_CLEAN'].map(
    lambda x: parts_lookup.get(x, {}).get('description', ''))
capacity['AREA'] = capacity['PART_CLEAN'].map(
    lambda x: parts_lookup.get(x, {}).get('area', ''))

print(f"  Capacity table: {len(capacity)} rows")

# ---------------------------------------------------------------------------
# 3. Current stock per location+part
# ---------------------------------------------------------------------------
print("Computing current stock...")

current_stock = df502.groupby(['LOCATION', 'COMM PRODUCT']).agg(
    CURRENT_QTY=('QUANTITY', 'sum'),
    MODULE_COUNT=('MODULE#', 'nunique')
).reset_index()
current_stock['COMM PRODUCT'] = current_stock['COMM PRODUCT'].str.replace(r'[\s\-]+', '', regex=True)
current_stock.rename(columns={'COMM PRODUCT': 'PART_CLEAN'}, inplace=True)

print(f"  Current stock: {len(current_stock)} location-part combos")

# Split 502 stock into three buckets:
#   - assigned: (location, part) matches a capacity-spec row → flows into the
#     left-join merge below and counts toward location-level utilization
#   - overflow: location is a designated overflow bucket (ECQPC, OVFLO) —
#     legitimate stock that contributes to per-part inventory but has no
#     rack capacity attached
#   - orphan: neither — stock sitting somewhere it shouldn't; surfaced in the
#     Summary as an alert so it doesn't silently vanish from the merge
_is_overflow = current_stock['LOCATION'].str[:5].isin(OVERFLOW_LOCATIONS).to_numpy()
overflow_stock = current_stock[_is_overflow].copy()

_cap_keys = set(zip(capacity['LOCATION'], capacity['PART_CLEAN']))
_cap_loc_to_part = dict(zip(capacity['LOCATION'], capacity['PART_CLEAN']))
_orphan_mask = np.array([
    (loc, part) not in _cap_keys
    for loc, part in zip(current_stock['LOCATION'], current_stock['PART_CLEAN'])
])
orphan_stock = current_stock[_orphan_mask & ~_is_overflow].copy()
orphan_stock['ASSIGNED_PART'] = orphan_stock['LOCATION'].map(_cap_loc_to_part).fillna('')
orphan_stock['REASON'] = np.where(
    orphan_stock['ASSIGNED_PART'] == '',
    'Location not in capacity spec',
    'Part mismatch at location'
)
if len(overflow_stock) > 0:
    _ov_qty = int(overflow_stock['CURRENT_QTY'].sum())
    print(f"  Overflow: {len(overflow_stock)} rows ({_ov_qty:,} pcs) in "
          f"{sorted(OVERFLOW_LOCATIONS)}")
if len(orphan_stock) > 0:
    _unmatched_qty = int(orphan_stock['CURRENT_QTY'].sum())
    print(f"  WARNING: {len(orphan_stock)} stock rows ({_unmatched_qty:,} pcs) "
          f"don't match the capacity spec")

# ---------------------------------------------------------------------------
# 4. Calculate capacity
# ---------------------------------------------------------------------------
print("Calculating capacity...")

# Merge current stock
capacity = capacity.merge(
    current_stock[['LOCATION', 'PART_CLEAN', 'CURRENT_QTY', 'MODULE_COUNT']],
    on=['LOCATION', 'PART_CLEAN'], how='left'
)
capacity['CURRENT_QTY'] = capacity['CURRENT_QTY'].fillna(0).astype(int)
capacity['MODULE_COUNT'] = capacity['MODULE_COUNT'].fillna(0).astype(int)

# --- Capacity formula ---
# Max Pallets = Windows × Pallets_per_window
# Max Modules = Max Pallets × Modules_per_layer × Stack_height
# Max QTY     = Max Modules × PCS/Module
capacity['MAX_PALLETS'] = (capacity['LOC_WINDOWS'] *
                            capacity['PALLETS_PER_WIN']).astype(int)
capacity['MAX_MODULES'] = (capacity['MAX_PALLETS'] *
                          capacity['MODULES_PER_LAYER'] *
                          capacity['STACK_HEIGHT']).astype(int)
capacity['MAX_QTY'] = (capacity['MAX_MODULES'] * capacity['PCS_FINAL']).astype(int)

# Current utilization
capacity['CURRENT_MODULES'] = np.where(
    capacity['PCS_FINAL'] > 0,
    np.ceil(capacity['CURRENT_QTY'] / capacity['PCS_FINAL']),
    0
).astype(int)
modules_per_pallet = (capacity['MODULES_PER_LAYER'] * capacity['STACK_HEIGHT']).clip(lower=1)
capacity['CURRENT_PALLETS'] = np.where(
    capacity['MODULES_PER_LAYER'] > 0,
    np.ceil(capacity['CURRENT_MODULES'] / modules_per_pallet),
    capacity['CURRENT_MODULES']
).astype(int)
capacity['UTILIZATION'] = np.where(
    capacity['MAX_QTY'] > 0,
    capacity['CURRENT_QTY'] / capacity['MAX_QTY'],
    0
)

# Windows used
modules_per_window = (capacity['PALLETS_PER_WIN'] * capacity['MODULES_PER_LAYER'] *
                    capacity['STACK_HEIGHT']).clip(lower=1)
capacity['WINDOWS_USED'] = np.where(
    capacity['CURRENT_MODULES'] > 0,
    np.minimum(np.ceil(capacity['CURRENT_MODULES'] / modules_per_window),
               capacity['LOC_WINDOWS']),
    0
).astype(int)
capacity['WINDOW_UTIL'] = np.where(
    capacity['LOC_WINDOWS'] > 0,
    capacity['WINDOWS_USED'] / capacity['LOC_WINDOWS'],
    0
)
capacity['MODULE_UTIL'] = np.where(
    capacity['MAX_MODULES'] > 0,
    capacity['CURRENT_MODULES'] / capacity['MAX_MODULES'],
    0
)
capacity['CURRENT_WEIGHT_KG'] = capacity['CURRENT_QTY'] * capacity['WEIGHT_PCS']

# Storage type label (all floor for Site 1)
capacity['STORAGE_LABEL'] = 'Floor'

# Sort by location
capacity = capacity.sort_values(['UNIT', 'RACK', 'POSITION', 'LEVEL']).reset_index(drop=True)

# Validation: flag any location where current stock exceeds capacity
over = capacity[capacity['CURRENT_QTY'] > capacity['MAX_QTY']]
print(f"  Capacity computed for {len(capacity)} rows")
if len(over) > 0:
    print(f"  WARNING: {len(over)} locations have current stock > capacity:")
    for _, ov in over.iterrows():
        print(f"    {ov['LOCATION']} {ov['PART_CLEAN']}: "
              f"current={ov['CURRENT_QTY']:,} > max={ov['MAX_QTY']:,}")

# Utilization band for easy filtering
def util_band(u):
    if u <= 0:
        return 'Empty'
    elif u <= 0.25:
        return 'Low (1-25%)'
    elif u <= 0.50:
        return 'Medium (26-50%)'
    elif u <= 0.75:
        return 'High (51-75%)'
    elif u <= 0.90:
        return 'Very High (76-90%)'
    elif u <= 1.0:
        return 'Critical (91-100%)'
    else:
        return 'Over Capacity'

capacity['UTIL_BAND'] = capacity['UTILIZATION'].apply(util_band)

# Spare capacity (used in Summary sheet)
capacity['SPARE'] = capacity['MAX_QTY'] - capacity['CURRENT_QTY']
capacity['SPARE_MODULES'] = capacity['MAX_MODULES'] - capacity['CURRENT_MODULES']

# ---------------------------------------------------------------------------
# 5. Build Excel output
# ---------------------------------------------------------------------------
print("\nBuilding Excel output...")

wb = Workbook()

# ---- Styles ----
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_fill2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_fill3 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
header_fill4 = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
pct_format = '0.0%'
num_format = '#,##0'
dec_format = '#,##0.00'
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def style_header(ws, row, fill=None):
    if fill is None:
        fill = header_fill
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border


def auto_width(ws, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def color_pct(cell, val):
    """Apply green/yellow/red fill to a percentage cell."""
    if val >= 0.9:
        cell.fill = red_fill
    elif val >= 0.5:
        cell.fill = yellow_fill
    elif val > 0:
        cell.fill = green_fill
    cell.font = Font(bold=True)


# =========================================================================
# Sheet 1: CapacityCalculation
# =========================================================================
ws1 = wb.active
ws1.title = "CapacityCalculation"

# Row 1: Section headers
# Section 1: Part & Location Info (cols 1-9, A-I)
ws1.merge_cells('A1:I1')
ws1['A1'] = 'Part & Location Info'
ws1['A1'].font = Font(bold=True, size=12, color="FFFFFF")
ws1['A1'].fill = header_fill
ws1['A1'].alignment = Alignment(horizontal='center')
for col in range(1, 10):
    ws1.cell(row=1, column=col).fill = header_fill

# Section 2: Packaging & Capacity (cols 10-19, J-S)
ws1.merge_cells('J1:S1')
ws1['J1'] = 'Packaging & Capacity'
ws1['J1'].font = Font(bold=True, size=12, color="FFFFFF")
ws1['J1'].fill = header_fill2
ws1['J1'].alignment = Alignment(horizontal='center')
for col in range(10, 20):
    ws1.cell(row=1, column=col).fill = header_fill2

# Section 3: Current Utilization (cols 20-26, T-Z)
ws1.merge_cells('T1:Z1')
ws1['T1'] = 'Current Utilization'
ws1['T1'].font = Font(bold=True, size=12, color="FFFFFF")
ws1['T1'].fill = header_fill3
ws1['T1'].alignment = Alignment(horizontal='center')
for col in range(20, 27):
    ws1.cell(row=1, column=col).fill = header_fill3

# Section 4: Additional Info (cols 27-31, AA-AE)
ws1.merge_cells('AA1:AE1')
ws1['AA1'] = 'Additional Info'
ws1['AA1'].font = Font(bold=True, size=12, color="FFFFFF")
ws1['AA1'].fill = header_fill4
ws1['AA1'].alignment = Alignment(horizontal='center')
for col in range(27, 32):
    ws1.cell(row=1, column=col).fill = header_fill4

# Row 2: Column headers
headers = [
    # Section 1: Part & Location Info (1-9)
    'Parts', 'Location', 'UNIT', 'RACK', 'POSITION', 'LEVEL',
    'Storage Type', 'Description', 'Packaging Description',
    # Section 2: Packaging & Capacity (cols 10-19, J-S)
    'Module Type', 'PCS/Module', 'Weight (lbs)', 'Modules/Layer', 'Windows',
    'Pallets/Window', 'Stack Height', 'Max Modules', 'Max Pallets',
    'Max QTY (Capacity)',
    # Section 3: Current Utilization (cols 20-26, T-Z)
    'Current QTY', 'Current Modules', 'Windows Used',
    '% Windows Used', '% Modules Used', '% QTY Utilization',
    'Current Pallets',
    # Section 4: Additional Info (cols 27-31, AA-AE)
    'Weight/PCS (kg)', 'Current Weight (kg)',
    'Modules Stored', 'Area', 'Utilization Band'
]
for col_idx, h in enumerate(headers, 1):
    ws1.cell(row=2, column=col_idx, value=h)
style_header(ws1, 2)

# Data rows
for i, row in capacity.iterrows():
    r = i + 3
    # Section 1: Part & Location Info
    ws1.cell(row=r, column=1, value=row['PART_CLEAN'])
    ws1.cell(row=r, column=2, value=row['LOCATION'])
    ws1.cell(row=r, column=3, value=row['UNIT'])
    ws1.cell(row=r, column=4, value=row['RACK'])
    ws1.cell(row=r, column=5, value=row['POSITION'])
    ws1.cell(row=r, column=6, value=row['LEVEL'])
    ws1.cell(row=r, column=7, value=row['STORAGE_LABEL'])
    ws1.cell(row=r, column=8, value=row.get('DESCRIPTION', ''))
    ws1.cell(row=r, column=9, value=row.get('PKG_DESCRIPTION', ''))
    # Section 2: Packaging & Capacity
    ws1.cell(row=r, column=10, value=row.get('MODULE_TYPE', ''))
    ws1.cell(row=r, column=11, value=int(row['PCS_FINAL']))
    c = ws1.cell(row=r, column=12, value=row['WEIGHT_LBS'] if row['WEIGHT_LBS'] > 0 else '')
    if row['WEIGHT_LBS'] > 0:
        c.number_format = '#,##0'
    ws1.cell(row=r, column=13, value=int(row['MODULES_PER_LAYER']))
    ws1.cell(row=r, column=14, value=int(row['LOC_WINDOWS']))
    ws1.cell(row=r, column=15, value=int(row['PALLETS_PER_WIN']))
    ws1.cell(row=r, column=16, value=int(row['STACK_HEIGHT']))
    ws1.cell(row=r, column=17, value=int(row['MAX_MODULES']))
    ws1.cell(row=r, column=18, value=int(row['MAX_PALLETS']))
    c = ws1.cell(row=r, column=19, value=int(row['MAX_QTY']))
    c.number_format = num_format
    # Section 3: Current Utilization
    c = ws1.cell(row=r, column=20, value=int(row['CURRENT_QTY']))
    c.number_format = num_format
    ws1.cell(row=r, column=21, value=int(row['CURRENT_MODULES']))
    ws1.cell(row=r, column=22, value=int(row['WINDOWS_USED']))
    # % Windows Used
    c = ws1.cell(row=r, column=23, value=round(row['WINDOW_UTIL'], 3))
    c.number_format = pct_format
    if row['WINDOW_UTIL'] >= 0.9:
        c.fill = red_fill
    elif row['WINDOW_UTIL'] >= 0.5:
        c.fill = yellow_fill
    elif row['WINDOW_UTIL'] > 0:
        c.fill = green_fill
    # % Modules Used
    c = ws1.cell(row=r, column=24, value=round(row['MODULE_UTIL'], 3))
    c.number_format = pct_format
    if row['MODULE_UTIL'] >= 0.9:
        c.fill = red_fill
    elif row['MODULE_UTIL'] >= 0.5:
        c.fill = yellow_fill
    elif row['MODULE_UTIL'] > 0:
        c.fill = green_fill
    # % QTY Utilization
    c = ws1.cell(row=r, column=25, value=round(row['UTILIZATION'], 3))
    c.number_format = pct_format
    if row['UTILIZATION'] >= 0.9:
        c.fill = red_fill
    elif row['UTILIZATION'] >= 0.5:
        c.fill = yellow_fill
    elif row['UTILIZATION'] > 0:
        c.fill = green_fill
    ws1.cell(row=r, column=26, value=int(row['CURRENT_PALLETS']))
    # Section 4: Additional Info
    c = ws1.cell(row=r, column=27, value=round(row['WEIGHT_PCS'], 5))
    c.number_format = '0.00000'
    c = ws1.cell(row=r, column=28, value=round(row['CURRENT_WEIGHT_KG'], 2))
    c.number_format = dec_format
    ws1.cell(row=r, column=29, value=int(row['MODULE_COUNT']))
    ws1.cell(row=r, column=30, value=row.get('AREA', ''))
    c = ws1.cell(row=r, column=31, value=row.get('UTIL_BAND', ''))
    band = row.get('UTIL_BAND', '')
    if band in ('Critical (91-100%)', 'Over Capacity'):
        c.fill = red_fill
    elif band in ('High (51-75%)', 'Very High (76-90%)'):
        c.fill = yellow_fill
    elif band in ('Low (1-25%)', 'Medium (26-50%)'):
        c.fill = green_fill
    # Add thin borders
    for col in range(1, 32):
        ws1.cell(row=r, column=col).border = thin_border

# Freeze top 2 rows
ws1.freeze_panes = 'A3'
ws1.auto_filter.ref = f"A2:AE{len(capacity) + 2}"
auto_width(ws1)

# =========================================================================
# Sheet 2: Summary
# =========================================================================
ws2 = wb.create_sheet("Summary")

# Totals
# Module/pallet/window counts are rack-only (overflow has no rack capacity).
# QTY and weight totals include overflow so they reflect actual inventory.
total_max_qty = capacity['MAX_QTY'].sum()
total_assigned_qty = int(capacity['CURRENT_QTY'].sum())
total_overflow_qty = int(overflow_stock['CURRENT_QTY'].sum()) if len(overflow_stock) else 0
total_current_qty = total_assigned_qty + total_overflow_qty
total_max_pallets = capacity['MAX_PALLETS'].sum()
total_current_pallets = capacity['CURRENT_PALLETS'].sum()
total_max_modules = capacity['MAX_MODULES'].sum()
total_current_modules = capacity['CURRENT_MODULES'].sum()
total_windows = capacity['LOC_WINDOWS'].sum()
total_windows_used = capacity['WINDOWS_USED'].sum()
total_locations = len(capacity)
occupied_locations = len(capacity[capacity['CURRENT_QTY'] > 0])
overall_util = total_current_qty / total_max_qty if total_max_qty > 0 else 0
module_util = total_current_modules / total_max_modules if total_max_modules > 0 else 0
pal_util = total_current_pallets / total_max_pallets if total_max_pallets > 0 else 0
win_util = total_windows_used / total_windows if total_windows > 0 else 0
loc_util = occupied_locations / total_locations if total_locations > 0 else 0

# Weight: rack-location weight plus overflow weight (using the same
# WEIGHT_PCS resolved earlier; fall back to 702 for overflow-only parts).
_part_weight = capacity.drop_duplicates('PART_CLEAN').set_index('PART_CLEAN')['WEIGHT_PCS'].to_dict()
for _p, _w in zip(df702['PART_CLEAN'], df702['WEIGHT_PCS']):
    _part_weight.setdefault(_p, _w)
_overflow_weight = sum(
    int(q) * _part_weight.get(p, 0)
    for q, p in zip(overflow_stock['CURRENT_QTY'], overflow_stock['PART_CLEAN'])
) if len(overflow_stock) else 0
total_weight = capacity['CURRENT_WEIGHT_KG'].sum() + _overflow_weight

# --- Title & date ---
ws2['A1'] = 'Site 1 Warehouse — Capacity Summary'
ws2['A1'].font = Font(bold=True, size=14, color="2F5496")
ws2.merge_cells('A1:D1')
ws2['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
ws2['A2'].font = Font(italic=True, size=10, color="808080")

# ===== CAPACITY OVERVIEW TABLE (Metric / Max / Current / Utilization) =====
r = 4
ws2.cell(row=r, column=1, value='Capacity Overview').font = Font(bold=True, size=12, color="2F5496")
r += 1
for col_idx, h in enumerate(['Metric', 'Max', 'Current', 'Utilization'], 1):
    ws2.cell(row=r, column=col_idx, value=h)
style_header(ws2, r, fill=header_fill)

overview_rows = [
    ('Modules',   int(total_max_modules),   int(total_current_modules),   module_util),
    ('Pallets',   int(total_max_pallets), int(total_current_pallets), pal_util),
    ('Windows',   int(total_windows),     int(total_windows_used),    win_util),
    ('Locations', int(total_locations),   int(occupied_locations),    loc_util),
    ('QTY (pcs)', int(total_max_qty),     int(total_current_qty),     overall_util),
]
for metric, mx, cur, util in overview_rows:
    r += 1
    ws2.cell(row=r, column=1, value=metric).font = Font(bold=True)
    c = ws2.cell(row=r, column=2, value=mx); c.number_format = num_format
    c = ws2.cell(row=r, column=3, value=cur); c.number_format = num_format
    c = ws2.cell(row=r, column=4, value=util); c.number_format = pct_format
    color_pct(c, util)
    for col in range(1, 5):
        ws2.cell(row=r, column=col).border = thin_border

# Additional reference row: weight & unique parts
r += 1
ws2.cell(row=r, column=1, value='Weight (kg)').font = Font(bold=True)
ws2.cell(row=r, column=2, value='')
c = ws2.cell(row=r, column=3, value=round(total_weight, 1)); c.number_format = num_format
ws2.cell(row=r, column=4, value='')
for col in range(1, 5):
    ws2.cell(row=r, column=col).border = thin_border
r += 1
ws2.cell(row=r, column=1, value='Unique Parts').font = Font(bold=True)
ws2.cell(row=r, column=2, value='')
c = ws2.cell(row=r, column=3, value=capacity['PART_CLEAN'].nunique()); c.number_format = num_format
ws2.cell(row=r, column=4, value='')
for col in range(1, 5):
    ws2.cell(row=r, column=col).border = thin_border

# ===== OVER-CAPACITY ALERTS =====
over_cap = capacity[capacity['CURRENT_QTY'] > capacity['MAX_QTY']]
if len(over_cap) > 0:
    r += 2
    ws2.cell(row=r, column=1, value=f'Over-Capacity Alerts ({len(over_cap)})').font = Font(
        bold=True, size=12, color="C00000")
    r += 1
    for col_idx, h in enumerate(['Location', 'Part', 'Description',
                                  'Current Modules', 'Max Modules', 'Over By (modules)'], 1):
        ws2.cell(row=r, column=col_idx, value=h)
    style_header(ws2, r, fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"))
    for _, ov in over_cap.iterrows():
        r += 1
        ws2.cell(row=r, column=1, value=ov['LOCATION'])
        ws2.cell(row=r, column=2, value=ov['PART_CLEAN'])
        ws2.cell(row=r, column=3, value=ov.get('DESCRIPTION', ''))
        c = ws2.cell(row=r, column=4, value=int(ov['CURRENT_MODULES'])); c.number_format = num_format
        c = ws2.cell(row=r, column=5, value=int(ov['MAX_MODULES'])); c.number_format = num_format
        over_by = int(ov['CURRENT_MODULES'] - ov['MAX_MODULES'])
        c = ws2.cell(row=r, column=6, value=over_by); c.number_format = num_format
        for col in range(1, 7):
            cell = ws2.cell(row=r, column=col)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

# ===== UNASSIGNED / MISMATCHED STOCK =====
if len(orphan_stock) > 0:
    r += 2
    ws2.cell(row=r, column=1,
             value=f'Unassigned / Mismatched Stock ({len(orphan_stock)})').font = Font(
        bold=True, size=12, color="C00000")
    r += 1
    for col_idx, h in enumerate(['Location', 'Stock Part', 'Assigned Part',
                                  'Current QTY', 'Reason'], 1):
        ws2.cell(row=r, column=col_idx, value=h)
    style_header(ws2, r, fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"))
    for _, orow in orphan_stock.sort_values('CURRENT_QTY', ascending=False).iterrows():
        r += 1
        ws2.cell(row=r, column=1, value=orow['LOCATION'])
        ws2.cell(row=r, column=2, value=orow['PART_CLEAN'])
        ws2.cell(row=r, column=3, value=orow['ASSIGNED_PART'])
        c = ws2.cell(row=r, column=4, value=int(orow['CURRENT_QTY']))
        c.number_format = num_format
        ws2.cell(row=r, column=5, value=orow['REASON'])
        for col in range(1, 6):
            cell = ws2.cell(row=r, column=col)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

# Overflow stock is rolled into each part's totals on the PerPartSummary
# sheet (Overflow QTY → Total Current QTY), so no standalone block here.

# ===== BREAKDOWN BY UNIT =====
r += 2
ws2.cell(row=r, column=1, value='Breakdown by UNIT').font = Font(bold=True, size=12, color="2F5496")
r += 1
unit_headers = ['UNIT', 'Locations', 'Max Modules', 'Current Modules', '% Module Util',
                'Spare Modules', 'Max Pallets', 'Current Pallets', '% Pallet Util',
                'Windows', 'Windows Used', '% Win Util']
for col_idx, h in enumerate(unit_headers, 1):
    ws2.cell(row=r, column=col_idx, value=h)
style_header(ws2, r, fill=header_fill2)

for unit in sorted(capacity['UNIT'].unique()):
    r += 1
    subset = capacity[capacity['UNIT'] == unit]
    max_b = int(subset['MAX_MODULES'].sum()); cur_b = int(subset['CURRENT_MODULES'].sum())
    max_p = int(subset['MAX_PALLETS'].sum()); cur_p = int(subset['CURRENT_PALLETS'].sum())
    tot_w = int(subset['LOC_WINDOWS'].sum()); used_w = int(subset['WINDOWS_USED'].sum())
    ws2.cell(row=r, column=1, value=unit).font = Font(bold=True)
    ws2.cell(row=r, column=2, value=len(subset))
    c = ws2.cell(row=r, column=3, value=max_b); c.number_format = num_format
    c = ws2.cell(row=r, column=4, value=cur_b); c.number_format = num_format
    butil = cur_b / max_b if max_b > 0 else 0
    c = ws2.cell(row=r, column=5, value=butil); c.number_format = pct_format; color_pct(c, butil)
    c = ws2.cell(row=r, column=6, value=max_b - cur_b); c.number_format = num_format
    c = ws2.cell(row=r, column=7, value=max_p); c.number_format = num_format
    c = ws2.cell(row=r, column=8, value=cur_p); c.number_format = num_format
    putil = cur_p / max_p if max_p > 0 else 0
    c = ws2.cell(row=r, column=9, value=putil); c.number_format = pct_format; color_pct(c, putil)
    c = ws2.cell(row=r, column=10, value=tot_w); c.number_format = num_format
    c = ws2.cell(row=r, column=11, value=used_w); c.number_format = num_format
    wutil = used_w / tot_w if tot_w > 0 else 0
    c = ws2.cell(row=r, column=12, value=wutil); c.number_format = pct_format; color_pct(c, wutil)
    for col in range(1, 13):
        ws2.cell(row=r, column=col).border = thin_border

# Total row
r += 1
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ws2.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
ws2.cell(row=r, column=2, value=total_locations)
c = ws2.cell(row=r, column=3, value=int(total_max_modules)); c.number_format = num_format
c = ws2.cell(row=r, column=4, value=int(total_current_modules)); c.number_format = num_format
c = ws2.cell(row=r, column=5, value=module_util); c.number_format = pct_format
c = ws2.cell(row=r, column=6, value=int(total_max_modules - total_current_modules)); c.number_format = num_format
c = ws2.cell(row=r, column=7, value=int(total_max_pallets)); c.number_format = num_format
c = ws2.cell(row=r, column=8, value=int(total_current_pallets)); c.number_format = num_format
c = ws2.cell(row=r, column=9, value=pal_util); c.number_format = pct_format
c = ws2.cell(row=r, column=10, value=int(total_windows)); c.number_format = num_format
c = ws2.cell(row=r, column=11, value=int(total_windows_used)); c.number_format = num_format
c = ws2.cell(row=r, column=12, value=win_util); c.number_format = pct_format
for col in range(1, 13):
    cell = ws2.cell(row=r, column=col)
    cell.border = thin_border
    cell.fill = total_fill
    cell.font = Font(bold=True)

# ===== BREAKDOWN BY RACK =====
r += 2
ws2.cell(row=r, column=1, value='Breakdown by RACK').font = Font(bold=True, size=12, color="2F5496")
r += 1
rack_headers = ['UNIT-RACK', 'Locations', 'Max Modules', 'Current Modules', '% Module Util',
                'Spare Modules', 'Max Pallets', 'Current Pallets', '% Pallet Util']
for col_idx, h in enumerate(rack_headers, 1):
    ws2.cell(row=r, column=col_idx, value=h)
style_header(ws2, r, fill=header_fill2)

rack_groups = capacity.groupby(['UNIT', 'RACK']).agg(
    LOC_COUNT=('LOCATION', 'count'),
    MAX_MODULES=('MAX_MODULES', 'sum'),
    CUR_MODULES=('CURRENT_MODULES', 'sum'),
    MAX_PALLETS=('MAX_PALLETS', 'sum'),
    CUR_PALLETS=('CURRENT_PALLETS', 'sum'),
).reset_index().sort_values(['UNIT', 'RACK'])

for _, rg in rack_groups.iterrows():
    r += 1
    max_b = int(rg['MAX_MODULES']); cur_b = int(rg['CUR_MODULES'])
    max_p = int(rg['MAX_PALLETS']); cur_p = int(rg['CUR_PALLETS'])
    ws2.cell(row=r, column=1, value=f"{rg['UNIT']}-{rg['RACK']}").font = Font(bold=True)
    ws2.cell(row=r, column=2, value=int(rg['LOC_COUNT']))
    c = ws2.cell(row=r, column=3, value=max_b); c.number_format = num_format
    c = ws2.cell(row=r, column=4, value=cur_b); c.number_format = num_format
    butil = cur_b / max_b if max_b > 0 else 0
    c = ws2.cell(row=r, column=5, value=butil); c.number_format = pct_format; color_pct(c, butil)
    c = ws2.cell(row=r, column=6, value=max_b - cur_b); c.number_format = num_format
    c = ws2.cell(row=r, column=7, value=max_p); c.number_format = num_format
    c = ws2.cell(row=r, column=8, value=cur_p); c.number_format = num_format
    putil = cur_p / max_p if max_p > 0 else 0
    c = ws2.cell(row=r, column=9, value=putil); c.number_format = pct_format; color_pct(c, putil)
    for col in range(1, 10):
        ws2.cell(row=r, column=col).border = thin_border

# ===== TOP 10 FULLEST LOCATIONS =====
r += 2
ws2.cell(row=r, column=1, value='Top 10 Fullest Locations (by Module %)').font = Font(
    bold=True, size=12, color="548235")
r += 1
for col_idx, h in enumerate(['Location', 'Part', 'Description',
                              'Current Modules', 'Max Modules', '% Module Util'], 1):
    ws2.cell(row=r, column=col_idx, value=h)
style_header(ws2, r, fill=header_fill3)

top_mod = capacity[capacity['CURRENT_MODULES'] > 0].nlargest(10, 'MODULE_UTIL')
for _, tu in top_mod.iterrows():
    r += 1
    ws2.cell(row=r, column=1, value=tu['LOCATION'])
    ws2.cell(row=r, column=2, value=tu['PART_CLEAN'])
    ws2.cell(row=r, column=3, value=tu.get('DESCRIPTION', ''))
    c = ws2.cell(row=r, column=4, value=int(tu['CURRENT_MODULES'])); c.number_format = num_format
    c = ws2.cell(row=r, column=5, value=int(tu['MAX_MODULES'])); c.number_format = num_format
    c = ws2.cell(row=r, column=6, value=round(tu['MODULE_UTIL'], 3)); c.number_format = pct_format
    color_pct(c, tu['MODULE_UTIL'])
    for col in range(1, 7):
        ws2.cell(row=r, column=col).border = thin_border

# ===== TOP 10 MOST SPARE =====
r += 2
ws2.cell(row=r, column=1, value='Top 10 Most Spare Capacity (by Modules)').font = Font(
    bold=True, size=12, color="BF8F00")
r += 1
for col_idx, h in enumerate(['Location', 'Part', 'Description',
                              'Current Modules', 'Max Modules', 'Spare Modules', '% Module Util'], 1):
    ws2.cell(row=r, column=col_idx, value=h)
style_header(ws2, r, fill=header_fill4)

least_mod = capacity.nlargest(10, 'SPARE_MODULES')
for _, lu in least_mod.iterrows():
    r += 1
    ws2.cell(row=r, column=1, value=lu['LOCATION'])
    ws2.cell(row=r, column=2, value=lu['PART_CLEAN'])
    ws2.cell(row=r, column=3, value=lu.get('DESCRIPTION', ''))
    c = ws2.cell(row=r, column=4, value=int(lu['CURRENT_MODULES'])); c.number_format = num_format
    c = ws2.cell(row=r, column=5, value=int(lu['MAX_MODULES'])); c.number_format = num_format
    c = ws2.cell(row=r, column=6, value=int(lu['SPARE_MODULES'])); c.number_format = num_format
    c = ws2.cell(row=r, column=7, value=round(lu['MODULE_UTIL'], 3)); c.number_format = pct_format
    for col in range(1, 8):
        ws2.cell(row=r, column=col).border = thin_border

# ===== UTILIZATION BAND DISTRIBUTION =====
r += 2
ws2.cell(row=r, column=1, value='Location Count by Utilization Band').font = Font(
    bold=True, size=12, color="2F5496")
r += 1
for col_idx, h in enumerate(['Utilization Band', 'Locations', '% of Total'], 1):
    ws2.cell(row=r, column=col_idx, value=h)
style_header(ws2, r, fill=header_fill)

band_order = ['Empty', 'Low (1-25%)', 'Medium (26-50%)', 'High (51-75%)',
              'Very High (76-90%)', 'Critical (91-100%)', 'Over Capacity']
band_counts = capacity['UTIL_BAND'].value_counts()
band_fills = {
    'Empty': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    'Low (1-25%)': green_fill, 'Medium (26-50%)': green_fill,
    'High (51-75%)': yellow_fill, 'Very High (76-90%)': yellow_fill,
    'Critical (91-100%)': red_fill, 'Over Capacity': red_fill,
}
for band in band_order:
    r += 1
    cnt = int(band_counts.get(band, 0))
    c = ws2.cell(row=r, column=1, value=band)
    c.fill = band_fills.get(band, PatternFill())
    c = ws2.cell(row=r, column=2, value=cnt); c.number_format = num_format
    c = ws2.cell(row=r, column=3, value=cnt / total_locations if total_locations > 0 else 0)
    c.number_format = pct_format
    for col in range(1, 4):
        ws2.cell(row=r, column=col).border = thin_border

ws2.freeze_panes = 'A3'
auto_width(ws2)

# =========================================================================
# Sheet 3: PerPartSummary
# =========================================================================
ws3 = wb.create_sheet("PerPartSummary")

ws3['A1'] = 'Per-Part Capacity Summary'
ws3['A1'].font = Font(bold=True, size=14, color="2F5496")
ws3.merge_cells('A1:M1')

part_headers = ['Part', 'Description', 'Module Type', 'PCS/Module', 'Modules/Layer',
                'Locations Assigned', 'Total Max QTY', 'Assigned QTY', 'Overflow QTY',
                'Total Current QTY', '% Utilization', 'Spare Capacity', 'Weight/PCS (kg)']
for col_idx, h in enumerate(part_headers, 1):
    ws3.cell(row=2, column=col_idx, value=h)
style_header(ws3, 2)

part_summary = capacity.groupby('PART_CLEAN').agg(
    DESCRIPTION=('DESCRIPTION', 'first'),
    MODULE_TYPE=('MODULE_TYPE', 'first'),
    PCS_MODULE=('PCS_FINAL', 'first'),
    MODULE_PLT=('MODULES_PER_LAYER', 'first'),
    LOC_COUNT=('LOCATION', 'count'),
    MAX_QTY=('MAX_QTY', 'sum'),
    ASSIGNED_QTY=('CURRENT_QTY', 'sum'),
    WEIGHT_PCS=('WEIGHT_PCS', 'first')
).reset_index()

# Attach overflow-location stock per part
_overflow_by_part = overflow_stock.groupby('PART_CLEAN')['CURRENT_QTY'].sum()
part_summary['OVERFLOW_QTY'] = part_summary['PART_CLEAN'].map(_overflow_by_part).fillna(0).astype(int)

# Parts that exist only in overflow (no rack location assigned at all)
_overflow_only = set(overflow_stock['PART_CLEAN']) - set(part_summary['PART_CLEAN'])
if _overflow_only:
    _extra = overflow_stock[overflow_stock['PART_CLEAN'].isin(_overflow_only)].groupby(
        'PART_CLEAN')['CURRENT_QTY'].sum().reset_index()
    _extra.rename(columns={'CURRENT_QTY': 'OVERFLOW_QTY'}, inplace=True)
    _extra['DESCRIPTION'] = _extra['PART_CLEAN'].map(
        lambda x: parts_lookup.get(x, {}).get('description', ''))
    _extra['MODULE_TYPE'] = ''
    _extra['PCS_MODULE'] = 0
    _extra['MODULE_PLT'] = 0
    _extra['LOC_COUNT'] = 0
    _extra['MAX_QTY'] = 0
    _extra['ASSIGNED_QTY'] = 0
    _extra['WEIGHT_PCS'] = _extra['PART_CLEAN'].map(_part_weight).fillna(0)
    part_summary = pd.concat([part_summary, _extra], ignore_index=True)

part_summary['CURRENT_QTY'] = part_summary['ASSIGNED_QTY'] + part_summary['OVERFLOW_QTY']
part_summary['UTIL'] = np.where(part_summary['MAX_QTY'] > 0,
                                 part_summary['CURRENT_QTY'] / part_summary['MAX_QTY'], 0)
part_summary['SPARE'] = part_summary['MAX_QTY'] - part_summary['CURRENT_QTY']
part_summary = part_summary.sort_values('CURRENT_QTY', ascending=False).reset_index(drop=True)

for i, row in part_summary.iterrows():
    r = 3 + i
    ws3.cell(row=r, column=1, value=row['PART_CLEAN'])
    ws3.cell(row=r, column=2, value=row['DESCRIPTION'])
    ws3.cell(row=r, column=3, value=row['MODULE_TYPE'])
    ws3.cell(row=r, column=4, value=int(row['PCS_MODULE']))
    ws3.cell(row=r, column=5, value=int(row['MODULE_PLT']))
    ws3.cell(row=r, column=6, value=int(row['LOC_COUNT']))
    c = ws3.cell(row=r, column=7, value=int(row['MAX_QTY'])); c.number_format = num_format
    c = ws3.cell(row=r, column=8, value=int(row['ASSIGNED_QTY'])); c.number_format = num_format
    c = ws3.cell(row=r, column=9, value=int(row['OVERFLOW_QTY'])); c.number_format = num_format
    c = ws3.cell(row=r, column=10, value=int(row['CURRENT_QTY'])); c.number_format = num_format
    c = ws3.cell(row=r, column=11, value=round(row['UTIL'], 3))
    c.number_format = pct_format
    if row['UTIL'] >= 0.9:
        c.fill = red_fill
    elif row['UTIL'] >= 0.5:
        c.fill = yellow_fill
    elif row['UTIL'] > 0:
        c.fill = green_fill
    c = ws3.cell(row=r, column=12, value=int(row['SPARE']))
    c.number_format = num_format
    c = ws3.cell(row=r, column=13, value=round(float(row['WEIGHT_PCS']), 5))
    c.number_format = '0.00000'

ws3.freeze_panes = 'A3'
ws3.auto_filter.ref = f"A2:M{len(part_summary) + 2}"
auto_width(ws3)

# =========================================================================
# Sheet 4: Raw 502 (current stock)
# =========================================================================
ws_502 = wb.create_sheet("502 - Current Stock")
for col_idx, h in enumerate(df502.columns, 1):
    ws_502.cell(row=1, column=col_idx, value=h)
style_header(ws_502, 1, fill=PatternFill(start_color="808080", end_color="808080", fill_type="solid"))
for i, (_, row) in enumerate(df502.iterrows()):
    for col_idx, col_name in enumerate(df502.columns, 1):
        val = row[col_name]
        ws_502.cell(row=2 + i, column=col_idx, value=val)
ws_502.freeze_panes = 'A2'
ws_502.auto_filter.ref = f"A1:{get_column_letter(len(df502.columns))}{len(df502) + 1}"
auto_width(ws_502)

# =========================================================================
# Sheet 5: Capacity Reference (raw data from Capacity Calculation.xlsx)
# =========================================================================
ws_ref = wb.create_sheet("Capacity Reference")
ref_cols = ['LOCATION', 'UNIT', 'RACK', 'POSITION', 'PRODUCT_CODE',
            'WEIGHT_STR', 'PKG_DESCRIPTION', 'PCS_MODULE', 'MODULE_TYPE',
            'STACK_HEIGHT', 'MODULES_PER_LAYER', 'PALLETS_PER_WIN', 'LOC_WINDOWS']
ref_headers = ['LOCATION', 'UNIT', 'RACK', 'POSITION', 'PRODUCT CODE',
               'Weight', 'Packaging Description', 'QTY per module', 'MODULE TYPE',
               'MODULE STACK HEIGHT', 'MODULES PER LAYER', 'PALLETS PER WINDOW', 'WINDOWS']
for col_idx, h in enumerate(ref_headers, 1):
    ws_ref.cell(row=1, column=col_idx, value=h)
style_header(ws_ref, 1, fill=PatternFill(start_color="808080", end_color="808080", fill_type="solid"))
for i, (_, row) in enumerate(cap_ref.iterrows()):
    for col_idx, col_name in enumerate(ref_cols, 1):
        val = row[col_name]
        ws_ref.cell(row=2 + i, column=col_idx, value=val)
ws_ref.freeze_panes = 'A2'
ws_ref.auto_filter.ref = f"A1:{get_column_letter(len(ref_cols))}{len(cap_ref) + 1}"
auto_width(ws_ref)

# =========================================================================
# Reorder sheets: Summary -> PerPartSummary -> CapacityCalculation -> Raw
# =========================================================================
desired_order = ['Summary', 'PerPartSummary', 'CapacityCalculation',
                 '502 - Current Stock', 'Capacity Reference']
for idx, name in enumerate(desired_order):
    current_idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=idx - current_idx)

# ---------------------------------------------------------------------------
# 6. Save
# ---------------------------------------------------------------------------
print(f"\nSaving to {OUTPUT_FILE}...")
OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT_FILE)
print(f"Done! Output: {OUTPUT_FILE}")
print(f"\nSummary:")
print(f"  Locations:  {occupied_locations} occupied / {total_locations} total ({loc_util:.0%})")
print(f"  Windows:    {total_windows_used:,} used / {total_windows:,} total ({win_util:.1%})")
print(f"  Modules:    {total_current_modules:,} used / {total_max_modules:,} total ({module_util:.1%})")
print(f"  QTY:        {total_current_qty:,} stored / {total_max_qty:,} capacity ({overall_util:.1%})")
if total_overflow_qty > 0:
    print(f"              (incl. {total_overflow_qty:,} pcs in overflow: {sorted(OVERFLOW_LOCATIONS)})")
print(f"  Pallets:    {int(total_current_pallets):,} used / {int(total_max_pallets):,} total ({pal_util:.1%})")
print(f"  Weight:     {total_weight:,.1f} kg")
