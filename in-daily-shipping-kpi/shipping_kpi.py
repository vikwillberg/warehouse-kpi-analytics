"""IN Daily Shipping KPI — V3.4.

V3.3's data logic and single-page layout, plus a delivery pass:
the email PNG is now a full-page screenshot of the finished HTML
report, captured with headless Edge/Chrome at 2x device scale
(write_report_screenshot; the old matplotlib KPI-card snapshot,
write_snapshot_png, remains as an automatic fallback when no browser
is found or the capture fails); the Copy Email Summary button also
opens the legacy email body in a new tab so it can be checked before
pasting; and a passive phone layout aimed at Outlook attachment
preview, where no JS runs: the dead day-tab strip, tap affordances,
and filter controls hide themselves (html:not(.js) at phone widths),
every day panel carries a static day title, and today's Unshipped
Orders table ships pre-rendered in the HTML so it reads without JS.
"""
from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
import html
import json
import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

import pandas as pd
from pandas.tseries.offsets import BDay
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

STATUS_BUCKET_MAP = {
    "Load.Conf*": "shipped",
    "Load.Conf": "shipped",
    "Load.Entry": "load_ent",
    "Load.Set": "load_set",
    "Pick & ski": "picked",
    "Tmp.Traile": "allocated",
}

UNSHIPPED_STATUS_MAP = {
    "Tmp.Traile": "2.Allocated",
    "Pick & ski": "3.Picked",
    "Load.Set": "4.Load.Set",
    "Load.Entry": "5.Load.Ent",
}

# Status -> dot color for the Unshipped table. Mirrors the JS `USC` map —
# the static no-JS render and the JS re-render must paint the same dots.
UNSHIPPED_STATUS_COLORS = {
    "1.No process": "#94a3b8",
    "2.Allocated": "#f97316",
    "3.Picked": "#f59e0b",
    "4.Load.Set": "#eab308",
    "5.Load.Ent": "#84cc16",
}

UNSHIPPED_HEADERS = [
    "TargetDate",
    "CUSTOMER ORDER NO.",
    "Plan_Ship",
    "PRODUCT NO.",
    "QUANTITY",
    "Status",
    "Module Count",
]

# On-screen labels for the same columns (the raw source-system names above are
# kept for the Excel export and as JS sort keys).
UNSHIPPED_DISPLAY_HEADERS = [
    "Target Date",
    "Customer Order No.",
    "Plan Ship",
    "Product No.",
    "Qty",
    "Status",
    "Modules",
]

# Table headers of the original emailed report (generate_daily_shipping_kpi.py),
# reused verbatim by the HTML's "Copy Email Summary" button so the copied
# content matches the legacy email exactly. "\n" renders as <br> in the cells.
LEGACY_SUMMARY2_HEADERS = [
    "Ship Date",
    "1.No process\n(order)",
    "2.Allocated\n(module)",
    "3.Picked\n(module)",
    "4.Load.Set\n(module)",
    "5.Load.Ent\n(module)",
    "6.Shipped\n(module)",
    "7.Shortage\n(order)",
    "8.Canceled\n(order)",
    "Plan Total\n(module)",
    "Ship Ratio",
]

LEGACY_SUMMARY1_HEADERS = [
    "Plan Date (min)",
    "Plan Date (max)",
    "Trailer",
    "BOL",
    "Actual Load (max)",
    "Order# (max)",
    "Skid Count",
    "Module Count",
]

STAGE_NAMES = ["Allocated", "Picked", "Load.Set", "Load.Entry", "Shipped"]

STAGE_CSS_CLASS = {
    "Allocated": "seg-alloc",
    "Picked": "seg-pick",
    "Load.Set": "seg-lset",
    "Load.Entry": "seg-lent",
    "Shipped": "seg-ship",
}

STAGE_COLORS = {
    "Allocated": "#f97316",
    "Picked": "#f59e0b",
    "Load.Set": "#eab308",
    "Load.Entry": "#84cc16",
    "Shipped": "#22c55e",
}

# Columns each CSV export must contain (only the ones the code actually
# reads). Checked up front so a wrong/partial export fails with a readable
# message instead of a pandas KeyError traceback.
REQUIRED_COLUMNS = {
    "201S.csv": [
        "CUSTOMER ORDER NO.", "STATUS", "SHIP DATE", "SHIP TIME",
        "UC/CNL", "PRODUCT NO.", "QUANTITY",
    ],
    "201P.csv": [
        "CUSTOMER ORDER NO.", "SHIP DATE", "SHIP TIME",
        "PRODUCT NO.", "QUANTITY",
    ],
    "202.csv": [
        "ORDER NO", "MODULE NO", "PLAN SHIP DATE", "PLAN SHIP TIME",
        "PICKING DATE", "PICKING TIME", "LOADING SET DATE", "LOADING SET TIME",
        "LOADING ENTRY DATE", "LOADING ENTRY TIME",
        "SHIPMENT LOAD DATE", "SHIPMENT LOAD TIME",
    ],
    "210.csv": [
        "Module No", "Trailer#", "Bol", "Skid No", "Weight", "Boxes",
        "Product No", "Qty",
    ],
    "701.CSV": [
        "MODEL/COLOR CODE", "WEIGHT PER PCS",
    ],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate Daily Shipping KPI report from CSV inputs."
    )
    parser.add_argument(
        "--base-dir",
        default=str(Path(__file__).resolve().parent),
        help="Base directory containing Data folder.",
    )
    parser.add_argument(
        "--data-dir",
        default=None,
        help="Optional override for CSV data folder.",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Date2 in YYYY-MM-DD format (defaults to today).",
    )
    parser.add_argument(
        "--dates",
        nargs=3,
        metavar=("DATE1", "DATE2", "DATE3"),
        default=None,
        help="Override all three dates in YYYY-MM-DD format (e.g. --dates 2026-02-12 2026-02-13 2026-02-16).",
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output HTML path (defaults to Output/HTML/Site1_Shipping_KPI_YYYYMMDD.html). "
            "Unshipped_List is saved to Output/Excel with a matching .xlsx file name."
        ),
    )
    parser.add_argument(
        "--no-png",
        action="store_true",
        help=(
            "Skip the full-report snapshot PNG "
            "(Output/PNG/Site1_Shipping_KPI_YYYYMMDD.png)."
        ),
    )
    parser.add_argument(
        "--open",
        action="store_true",
        help="Open the HTML report in the default browser after generating.",
    )
    return parser.parse_args()


def resolve_dates(
    date2_str: str | None, dates_override: list[str] | None = None
) -> tuple[list[date], date]:
    if dates_override:
        dates = [datetime.strptime(d, "%Y-%m-%d").date() for d in dates_override]
        return dates, dates[len(dates) // 2]
    if date2_str:
        date2 = datetime.strptime(date2_str, "%Y-%m-%d").date()
    else:
        date2 = datetime.today().date()
    # Generate 9 business days: BDay(-4) through BDay(+4) from today
    dates = [(pd.Timestamp(date2) + BDay(offset)).date() for offset in range(-4, 5)]
    return dates, date2


def clean_date_column(df: pd.DataFrame, col: str) -> pd.Series:
    values = df[col].mask(df[col].isin(["0001-01-01", ""]), pd.NA)
    return pd.to_datetime(values, errors="coerce").dt.date


def build_date_mask(series: pd.Series, target_dates: set[date]) -> pd.Series:
    target_series = pd.to_datetime(list(target_dates), errors="coerce").normalize()
    series_dt = pd.to_datetime(series, errors="coerce").dt.normalize()
    return series_dt.isin(target_series)


def clean_time_column(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.replace(".", ":", regex=False)


def combine_datetime(date_series: pd.Series, time_series: pd.Series) -> pd.Series:
    date_vals = date_series.fillna("").astype(str)
    time_vals = clean_time_column(time_series)
    return pd.to_datetime(date_vals + " " + time_vals, errors="coerce")


def format_date(value: date | None) -> str:
    if value is None or pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime("%m/%d/%Y")


def format_datetime(value) -> str:
    if value is None or pd.isna(value):
        return "N/A"
    return pd.Timestamp(value).strftime("%m/%d/%Y %H:%M")


def normalize_trailer(value: str | None) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).replace("#", "").strip()


def _pct_color(pct: float) -> str:
    """Green / amber / red threshold color shared by the header, overview
    table, and the PNG snapshot (>=80 green, >=50 amber, else red)."""
    if pct >= 80:
        return "#22c55e"
    if pct >= 50:
        return "#eab308"
    return "#ef4444"


def _day_label(target_date: date, date2: date, short: bool = False) -> str:
    """Relative day label (Yesterday / Today / Tomorrow / weekday) shared by the
    day tabs, the overview table, and the PNG snapshot. ``short`` trims the
    long "Next Ship Day: Friday" form down to just the weekday for tight cells."""
    if target_date < date2:
        if (date2 - target_date).days == 1:
            return "Yesterday"
        return target_date.strftime("%A")
    if target_date == date2:
        return "Today"
    if (target_date - date2).days == 1:
        return "Tomorrow"
    weekday = target_date.strftime("%A")
    return weekday if short else f"Next Ship Day: {weekday}"


def _cell_int(value) -> int:
    """Coerce a Summary2 cell (already an int, or a string like '83.00%') to int,
    falling back to 0. Used by the overview table and the PNG snapshot."""
    try:
        if isinstance(value, (int, float)):
            return int(value)
        return int(float(str(value).replace("%", "").strip() or 0))
    except (ValueError, TypeError):
        return 0


def validate_inputs(data_dir: Path) -> list[str]:
    """Pre-flight check of the four CSV exports: existence + required columns.
    Returns a list of human-readable problems (empty when all good)."""
    problems: list[str] = []
    for name, required in REQUIRED_COLUMNS.items():
        path = data_dir / name
        if not path.is_file():
            problems.append(f"{name} is missing from {data_dir}")
            continue
        try:
            header = pd.read_csv(path, dtype=str, nrows=0)
        except Exception as exc:  # unreadable/empty export
            problems.append(f"{name} could not be read ({exc})")
            continue
        missing = [c for c in required if c not in header.columns]
        if missing:
            problems.append(
                f"{name} is missing column(s): {', '.join(missing)} "
                "(wrong export selected?)"
            )
    return problems


def data_freshness(data_dir: Path) -> tuple[str | None, bool]:
    """Newest daily-export time (display string) and a stale flag set when the
    oldest of the four daily inputs (201S/201P/202/210) is more than 12 hours
    old. The 701 product master is deliberately excluded: it is slow-changing
    reference data, not a daily export, so an old 701 must not trip the stale
    warning (nor skew the as-of stamp). Fail-soft: any OS error just disables
    the stamp."""
    daily = [name for name in REQUIRED_COLUMNS if name != "701.CSV"]
    try:
        mtimes = [
            datetime.fromtimestamp((data_dir / name).stat().st_mtime)
            for name in daily
        ]
    except OSError:
        return None, False
    stale = (datetime.now() - min(mtimes)) > timedelta(hours=12)
    return max(mtimes).strftime("%m/%d/%Y %H:%M"), stale


def fill_zero_weights(df210: pd.DataFrame, df701: pd.DataFrame) -> int:
    """Backfill 210 rows whose Weight is zero/blank from the 701 product master.

    701's "WEIGHT PER PCS" (column U) is the per-piece weight of each product,
    keyed by "MODEL/COLOR CODE" — which matches 210's "Product No". For every
    210 row with no real weight we set Weight = WEIGHT PER PCS * Qty (the row's
    own piece quantity). Rows that already carry a non-zero weight, whose product
    is absent from 701, or whose Qty is blank are left untouched. df210["Weight"]
    is rewritten in place (as a 2-decimal string, matching the export's format)
    and the count of filled rows is returned.
    """
    weight_per_pcs = (
        pd.DataFrame({
            "key": df701["MODEL/COLOR CODE"].astype(str).str.strip(),
            "wpp": pd.to_numeric(df701["WEIGHT PER PCS"], errors="coerce"),
        })
        .dropna(subset=["wpp"])
        .drop_duplicates("key")
        .set_index("key")["wpp"]
    )

    zero_mask = pd.to_numeric(df210["Weight"], errors="coerce").fillna(0) == 0
    product = df210["Product No"].astype(str).str.strip()
    qty = pd.to_numeric(df210["Qty"], errors="coerce")
    computed = product.map(weight_per_pcs) * qty

    fill_mask = zero_mask & computed.notna()
    df210.loc[fill_mask, "Weight"] = computed[fill_mask].map("{:.2f}".format)
    return int(fill_mask.sum())


def compute_summary2(
    df201s: pd.DataFrame,
    df201p: pd.DataFrame,
    df202: pd.DataFrame,
    df210: pd.DataFrame,
    target_date: date,
) -> list:
    modules_date = df202[df202["PLAN SHIP DATE"] == target_date][
        ["ORDER NO", "MODULE NO"]
    ].copy()

    joined = modules_date.merge(
        df210[["Module No", "Trailer#"]].drop_duplicates("Module No"),
        left_on="MODULE NO",
        right_on="Module No",
        how="left",
    )
    has_trailer = joined["Trailer#"].fillna("").str.strip() != ""
    joined["is_shipped"] = has_trailer
    modules_date = joined[["ORDER NO", "MODULE NO", "is_shipped"]].copy()

    order_status = (
        df201s[df201s["SHIP DATE"] == target_date]
        .drop_duplicates("CUSTOMER ORDER NO.")
        .set_index("CUSTOMER ORDER NO.")["STATUS"]
    )

    def get_bucket(row):
        if row["is_shipped"]:
            return "shipped"
        status = order_status.get(row["ORDER NO"], "")
        return STATUS_BUCKET_MAP.get(status, "other")

    modules_date["bucket"] = modules_date.apply(get_bucket, axis=1)

    bucket_counts = modules_date.groupby("bucket").size()
    allocated = int(bucket_counts.get("allocated", 0))
    picked = int(bucket_counts.get("picked", 0))
    load_set = int(bucket_counts.get("load_set", 0))
    load_ent = int(bucket_counts.get("load_ent", 0))
    shipped = int(bucket_counts.get("shipped", 0))

    plan_total = allocated + picked + load_set + load_ent + shipped
    no_process = (
        df201p[df201p["SHIP DATE"] == target_date]["CUSTOMER ORDER NO."]
        .nunique()
    )

    df201s_date = df201s[df201s["SHIP DATE"] == target_date]
    uc_cnl = df201s_date["UC/CNL"].fillna("").astype(str)
    shortage = df201s_date[uc_cnl.str.contains("short", case=False, na=False)][
        "CUSTOMER ORDER NO."
    ].nunique()
    canceled = df201s_date[
        uc_cnl.str.contains("cnl|cancel", case=False, na=False)
    ]["CUSTOMER ORDER NO."].nunique()

    ship_ratio = shipped / plan_total if plan_total else 0
    ship_ratio_display = f"{ship_ratio * 100:.2f}%"

    return [
        format_date(target_date),
        int(no_process),
        allocated,
        picked,
        load_set,
        load_ent,
        shipped,
        int(shortage),
        int(canceled),
        plan_total,
        ship_ratio_display,
    ]


def compute_summary1(
    df202: pd.DataFrame,
    df210: pd.DataFrame,
    target_date: date,
) -> list[dict]:
    sub = df202[df202["PLAN SHIP DATE"] == target_date]
    if sub.empty:
        return []

    joined = sub.merge(df210, left_on="MODULE NO", right_on="Module No", how="left")
    joined["Trailer_norm"] = joined["Trailer#"].map(normalize_trailer)
    joined["Bol_norm"] = joined["Bol"].fillna("").astype(str).str.strip()

    missing = (joined["Trailer_norm"] == "") | (joined["Bol_norm"] == "")
    joined.loc[missing, ["Trailer_norm", "Bol_norm"]] = "N/A"

    def _module_stage(row):
        if pd.notna(row.get("SHIPMENT LOAD DATE")):
            return "Shipped"
        if pd.notna(row.get("LOADING ENTRY DATE")):
            return "Load.Entry"
        if pd.notna(row.get("LOADING SET DATE")):
            return "Load.Set"
        if pd.notna(row.get("PICKING DATE")):
            return "Picked"
        return "Allocated"

    joined["stage"] = joined.apply(_module_stage, axis=1)

    grouped = (
        joined.groupby(["Trailer_norm", "Bol_norm"], dropna=False)
        .agg(
            plan_min=("plan_dt", "min"),
            plan_max=("plan_dt", "max"),
            actual_load_max=("ship_dt", "max"),
            order_max=("ORDER NO", "max"),
            skid_count=("Skid No", "nunique"),
            module_count=("Module No", "size"),
        )
        .reset_index()
    )

    stage_counts = (
        joined.groupby(["Trailer_norm", "Bol_norm"])["stage"]
        .value_counts()
        .unstack(fill_value=0)
        .reset_index()
    )
    for s in STAGE_NAMES:
        if s not in stage_counts.columns:
            stage_counts[s] = 0
    grouped = grouped.merge(stage_counts, on=["Trailer_norm", "Bol_norm"], how="left")

    grouped["plan_sort"] = grouped["plan_min"].fillna(pd.Timestamp.max)
    grouped = grouped.sort_values(["plan_sort", "Trailer_norm", "Bol_norm"]).drop(
        columns=["plan_sort"]
    )

    # Collect per-module detail rows for each trailer group
    def _safe(val):
        if val is None or pd.isna(val):
            return ""
        return str(val).strip()

    def _fmt_dt(d, t):
        ds = _safe(d)
        ts = _safe(t).replace(".", ":")
        if not ds:
            return ""
        return f"{ds} {ts}".strip() if ts else ds

    module_details = {}
    for _, m in joined.iterrows():
        key = (m["Trailer_norm"], m["Bol_norm"])
        if key not in module_details:
            module_details[key] = []
        module_details[key].append({
            "mod": _safe(m.get("Module No", m.get("MODULE NO", ""))),
            "parts": _safe(m.get("PARTS NO", "")),
            "qty": _safe(m.get("QTY", "")),
            "serial": _safe(m.get("SERIAL NO", "")),
            "order": _safe(m.get("ORDER NO", "")),
            "stage": m["stage"],
            "pick": _fmt_dt(m.get("PICKING DATE"), m.get("PICKING TIME")),
            "lset": _fmt_dt(m.get("LOADING SET DATE"), m.get("LOADING SET TIME")),
            "lent": _fmt_dt(m.get("LOADING ENTRY DATE"), m.get("LOADING ENTRY TIME")),
            "ship": _fmt_dt(m.get("SHIPMENT LOAD DATE"), m.get("SHIPMENT LOAD TIME")),
            "weight": _safe(m.get("Weight", "")),
            "boxes": _safe(m.get("Boxes", "")),
            "supplier": _safe(m.get("supplier short name", _safe(m.get("SUPPLIER CODE", "")))),
        })

    results = []
    for _, row in grouped.iterrows():
        order_max = row["order_max"] if pd.notna(row["order_max"]) else "N/A"
        key = (row["Trailer_norm"], row["Bol_norm"])
        details = module_details.get(key, [])
        # "Late" means the trailer shipped on a LATER DATE than planned (a
        # missed ship day) — an intraday delay on the planned day counts as on
        # time. `is_late` is that date-based flag, shared by the "Late: N" day
        # chip, the modal Ship Delay row, and the red timeline marker so they
        # always agree. `late_min` keeps the full actual-vs-first-planned delta
        # as the delay *magnitude* the modal shows when a trailer is late.
        late_min = None
        is_late = False
        if pd.notna(row["actual_load_max"]) and pd.notna(row["plan_min"]):
            late_min = int(
                (row["actual_load_max"] - row["plan_min"]).total_seconds() // 60
            )
            is_late = row["actual_load_max"].date() > row["plan_min"].date()
        total_weight = 0.0
        # 210.csv's "Boxes" column actually holds piece quantity per module,
        # so the sum is exposed (and labeled) as a quantity.
        total_qty = 0
        for d in details:
            try:
                total_weight += float(d["weight"]) if d["weight"] else 0
            except (ValueError, TypeError):
                pass
            try:
                total_qty += int(float(d["boxes"])) if d["boxes"] else 0
            except (ValueError, TypeError):
                pass
        results.append({
            "plan_min": format_datetime(row["plan_min"]),
            "plan_max": format_datetime(row["plan_max"]),
            "trailer": row["Trailer_norm"],
            "bol": row["Bol_norm"],
            "actual_load": format_datetime(row["actual_load_max"]),
            "order_max": order_max,
            "skid_count": int(row["skid_count"] or 0),
            "module_count": int(row["module_count"] or 0),
            "stages": {s: int(row.get(s, 0)) for s in STAGE_NAMES},
            "total_weight": round(total_weight, 1),
            "total_qty": total_qty,
            "late_min": late_min,
            "is_late": is_late,
            "modules": details,
        })
    return results


def compute_unshipped_list(
    df201s: pd.DataFrame,
    df201p: pd.DataFrame,
    df202: pd.DataFrame,
    module_counts: pd.Series,
    target_dates: set[date],
) -> pd.DataFrame:
    plan_min = df202.groupby("ORDER NO")["plan_dt"].min()

    df201s_unshipped = df201s[df201s["STATUS"].isin(UNSHIPPED_STATUS_MAP)].copy()
    df201s_unshipped = df201s_unshipped[
        build_date_mask(df201s_unshipped["SHIP DATE"], target_dates)
    ]
    df201s_unshipped["Status"] = df201s_unshipped["STATUS"].map(UNSHIPPED_STATUS_MAP)
    df201s_unshipped["Module Count"] = (
        df201s_unshipped["CUSTOMER ORDER NO."].map(module_counts).fillna(0).astype(int)
    )
    df201s_unshipped["Plan_Ship"] = df201s_unshipped["CUSTOMER ORDER NO."].map(plan_min)

    fallback_plan_ship = combine_datetime(
        df201s_unshipped["SHIP DATE"].astype(str),
        df201s_unshipped["SHIP TIME"],
    )
    df201s_unshipped["Plan_Ship"] = df201s_unshipped["Plan_Ship"].fillna(
        fallback_plan_ship
    )

    df201s_out = df201s_unshipped[
        [
            "SHIP DATE",
            "CUSTOMER ORDER NO.",
            "Plan_Ship",
            "PRODUCT NO.",
            "QUANTITY",
            "Status",
            "Module Count",
        ]
    ].copy()

    df201p_unshipped = df201p[
        build_date_mask(df201p["SHIP DATE"], target_dates)
    ].copy()
    df201p_unshipped["Status"] = "1.No process"
    df201p_unshipped["Module Count"] = (
        df201p_unshipped["CUSTOMER ORDER NO."].map(module_counts).fillna(0).astype(int)
    )
    df201p_unshipped["Plan_Ship"] = combine_datetime(
        df201p_unshipped["SHIP DATE"].astype(str),
        df201p_unshipped["SHIP TIME"],
    )

    df201p_out = df201p_unshipped[
        [
            "SHIP DATE",
            "CUSTOMER ORDER NO.",
            "Plan_Ship",
            "PRODUCT NO.",
            "QUANTITY",
            "Status",
            "Module Count",
        ]
    ].copy()

    combined = pd.concat([df201s_out, df201p_out], ignore_index=True)
    combined = combined.rename(columns={"SHIP DATE": "TargetDate"})
    combined["TargetDate"] = combined["TargetDate"].apply(format_date)
    combined["Plan_Ship"] = combined["Plan_Ship"].apply(format_datetime)

    combined = combined.sort_values(
        ["TargetDate", "Status", "CUSTOMER ORDER NO."], ignore_index=True
    )
    return combined


def html_escape(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return html.escape(str(value))


def render_header(value: str) -> str:
    return html_escape(value).replace("\n", "<br>")


def _report_css() -> str:
    return """
/* ════════════════════════════════════════════════════
   DESIGN TOKENS — every color/shadow lives here; change
   a variable below to retheme the whole report.
   ════════════════════════════════════════════════════ */
:root{
  --bg:#eef2f6;
  --bg-img:radial-gradient(1100px 500px at 85% -10%,rgba(59,130,246,.07),transparent 60%),radial-gradient(900px 420px at -10% 110%,rgba(13,148,136,.06),transparent 60%);
  --card:#ffffff;--card-2:#f8fafc;--track:#eef2f7;
  --line:#e3e9f1;--line-2:#eef2f7;
  --ink:#0f172a;--txt:#1e293b;--ink-2:#334155;--mut:#64748b;--faint:#94a3b8;
  --brand:#1e40af;--brand-2:#3b82f6;--brand-soft:#eff6ff;
  --hdr-grad:linear-gradient(120deg,#0b1220 0%,#1d3a8f 52%,#0c6e62 100%);
  --tab-act:linear-gradient(135deg,#1e40af,#1c3565);
  --thead:#1e40af;
  --sel:#e8f0fe;--hover:#eff6ff;--even:#fafbfd;
  --ok:#22c55e;--warn:#eab308;--bad:#ef4444;--alert:#dc2626;
  --ok-bg:#dcfce7;--ok-t:#15803d;
  --warn-bg:#fef3c7;--warn-t:#92400e;
  --orange-bg:#ffedd5;--orange-t:#9a3412;
  --bad-bg:#fee2e2;--bad-t:#991b1b;
  --mod-bg:#eef2f7;--mod-t:#1e40af;
  --strip:rgba(255,255,255,.88);
  --shadow-1:0 1px 3px rgba(15,23,42,.05);
  --shadow-2:0 3px 14px rgba(15,23,42,.07);
  --shadow-3:0 14px 38px rgba(15,23,42,.16);
  --ring:rgba(59,130,246,.35);
}
/* ── Base ── */
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{font-family:'Segoe UI Variable Display','Segoe UI',system-ui,-apple-system,sans-serif;background:var(--bg);background-image:var(--bg-img);background-attachment:fixed;color:var(--txt);min-height:100vh;line-height:1.5}
.db{max-width:1300px;margin:0 auto;padding:24px}
::selection{background:rgba(59,130,246,.25)}
button{font-family:inherit}
:focus-visible{outline:2px solid var(--brand-2);outline-offset:2px}
::-webkit-scrollbar{width:9px;height:9px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--track);border-radius:5px}
::-webkit-scrollbar-thumb:hover{background:var(--faint)}

/* ── Header ── */
/* Title left, KPI cards pinned to the right edge; the fixed-width cards keep
   the group from shifting when the numbers or the date text change length */
.hdr{background:var(--hdr-grad);color:#fff;padding:22px 34px;border-radius:18px;margin-bottom:26px;position:relative;overflow:hidden;display:flex;align-items:center;justify-content:space-between;gap:24px;flex-wrap:wrap;box-shadow:var(--shadow-2)}
.hdr::before{content:'';position:absolute;top:-70%;right:-8%;width:520px;height:520px;background:radial-gradient(circle,rgba(255,255,255,.08) 0%,transparent 70%);border-radius:50%}
.hdr::after{content:'';position:absolute;left:0;right:0;bottom:0;height:3px;background:linear-gradient(90deg,#f97316,#f59e0b,#eab308,#84cc16,#22c55e);opacity:.85}
.hdr-left{position:relative}
.hdr h1{font-size:1.55rem;font-weight:800;letter-spacing:-.03em}
.hdr .sub{font-size:1rem;opacity:.85;margin-top:3px;font-weight:600;font-variant-numeric:tabular-nums}
.hero{display:flex;gap:12px;position:relative;flex-wrap:wrap;justify-content:flex-end;margin-left:auto}
/* Fixed equal width + centered text so the cards never shift when the
   numbers change size while flipping through days */
.hero-card{background:rgba(255,255,255,.09);backdrop-filter:blur(10px);border:1px solid rgba(255,255,255,.14);border-radius:12px;padding:11px 12px;width:148px;flex:0 0 auto;text-align:center}
.hero-card .v{font-size:1.45rem;font-weight:800;line-height:1.1;font-variant-numeric:tabular-nums}
.hero-card .l{font-size:.62rem;opacity:.65;text-transform:uppercase;letter-spacing:.07em;margin-top:3px;white-space:nowrap}
.hero-bar{height:4px;border-radius:2px;background:rgba(255,255,255,.18);margin-top:7px;overflow:hidden}
.hero-bar i{display:block;height:100%;border-radius:2px;transition:width .5s ease,background .5s ease}

/* ── Sections ── */
.sec{margin-bottom:30px}
#ship-sec{scroll-margin-top:16px}
.sec-hd{font-size:1.02rem;font-weight:800;color:var(--brand);margin-bottom:14px;display:flex;align-items:center;gap:10px;letter-spacing:-.01em}
.sec-hd::after{content:'';flex:1;height:2px;background:linear-gradient(90deg,var(--brand-2),transparent);border-radius:1px;opacity:.5}
.sec-sub{font-size:.78rem;color:var(--faint);font-weight:600}

/* ── Daily Snapshot table (compact: glanceable, not a wall) ── */
.ov-snap{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:12px 16px;box-shadow:var(--shadow-1)}
.ov-snap-title{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:8px;flex-wrap:wrap}
.ov-snap-title .d{font-size:.72rem;color:var(--mut);font-weight:600}
/* Without JS rows aren't clickable and the title row is empty — hide it */
html:not(.js) .ov-snap-title{display:none}
.ov-tbl{width:100%;border-collapse:collapse;font-size:.82rem}
.ov-tbl th{background:var(--thead);color:#fff;padding:6px 12px;text-align:right;font-weight:600;font-size:.64rem;white-space:nowrap;text-transform:uppercase;letter-spacing:.05em}
.ov-tbl th:first-child{border-radius:7px 0 0 7px;text-align:left}
.ov-tbl th:last-child{border-radius:0 7px 7px 0}
.ov-tbl th:nth-child(2){text-align:left}
.ov-tbl td{padding:5px 12px;border-bottom:1px solid var(--line-2);text-align:right;color:var(--ink-2);white-space:nowrap;font-variant-numeric:tabular-nums}
.ov-tbl td:first-child,.ov-tbl td:nth-child(2){text-align:left}
/* Pack the numeric columns shrink-to-fit on the right so Plan / Shipped /
   On-Time sit next to each other for easy comparison */
.ov-tbl th:nth-child(n+3),.ov-tbl td:nth-child(n+3){width:1%}
.ov-tbl tbody tr{cursor:pointer;transition:background .12s}
.ov-tbl tbody tr:last-child td{border-bottom:none}
.ov-tbl tbody tr:nth-child(even){background:var(--even)}
.ov-tbl tbody tr:hover{background:var(--hover)}
.ov-tbl tr.ov-future td{color:var(--faint)}
.ov-tbl tr.ov-today-row{background:var(--sel)}
.ov-tbl tr.ov-today-row td{color:var(--ink);font-weight:700}
.ov-tbl tr.ov-sel td:first-child{box-shadow:inset 3px 0 0 var(--brand-2)}
.ov-day{font-weight:700;color:var(--brand)}
.ov-ratio{font-weight:700;color:#fff;border-radius:5px;padding:1px 8px;display:inline-block;min-width:46px;text-align:center;font-size:.72rem}
.neg{color:var(--alert);font-weight:700}
.dash{color:var(--faint)}
.ov-btn{display:inline-flex;align-items:center;gap:6px;padding:5px 12px;border:none;border-radius:8px;background:var(--brand);color:#fff;font-size:.74rem;font-weight:700;cursor:pointer;box-shadow:var(--shadow-1);transition:filter .15s,transform .15s}
.ov-btn:hover{filter:brightness(1.12);transform:translateY(-1px)}
html:not(.js) .ov-btn{display:none}
.ov-btn.ghost{background:var(--card);color:var(--brand);border:1px solid var(--line)}
.ov-btn.ghost:hover{filter:none;background:var(--hover)}
/* Right side of the snapshot title row: action buttons */
.ov-actions{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-left:auto}
.ov-go-h{width:26px}
.ov-go{text-align:center;color:var(--faint);font-weight:800;font-size:.92rem;transition:color .12s,transform .12s;display:inline-block}
.ov-tbl tbody tr:hover .ov-go{color:var(--brand);transform:translateX(3px)}
.ov-tbl tbody tr.ov-today-row .ov-go{color:var(--brand)}

/* ── Sticky day-switcher bar (tabs + prev/next) ── */
.tabs-wrap{position:sticky;top:10px;z-index:350;margin-bottom:14px}
html:not(.js) .tabs-wrap{position:static}
.tabs{display:flex;flex-wrap:wrap;align-items:stretch;gap:6px;padding:6px;background:var(--strip);backdrop-filter:blur(14px);-webkit-backdrop-filter:blur(14px);border:1px solid var(--line);border-radius:16px;box-shadow:var(--shadow-2)}
.tab{flex:1 1 26%;min-width:0;padding:12px 14px 16px;border-radius:11px;cursor:pointer;text-align:center;transition:background .15s,box-shadow .15s,color .15s;position:relative;overflow:hidden}
.tab:hover{background:var(--hover)}
.tab.act{background:var(--tab-act);color:#fff;box-shadow:0 4px 14px rgba(30,64,175,.3)}
.tab.act:hover{background:var(--tab-act)}
.tab-day{font-size:.72rem;text-transform:uppercase;letter-spacing:.06em;opacity:.55;font-weight:700}
.tab-date{font-size:1.06rem;font-weight:800;margin:3px 0 2px;font-variant-numeric:tabular-nums}
.tab-mini{font-size:.72rem;opacity:.55;font-weight:600}
.tab.act .tab-day,.tab.act .tab-mini{opacity:.85}
.tab-today::after{content:'';position:absolute;top:9px;right:9px;width:8px;height:8px;border-radius:50%;background:var(--ok);box-shadow:0 0 0 2px var(--card)}
.tab.act.tab-today::after{box-shadow:0 0 0 2px rgba(255,255,255,.6)}
.tab-prog{position:absolute;left:10px;right:10px;bottom:6px;height:3px;border-radius:2px;background:rgba(148,163,184,.25);overflow:hidden}
.tab-prog i{display:block;height:100%;border-radius:2px}
.tab.act .tab-prog{background:rgba(255,255,255,.25)}
.tab-nav-btn{align-self:stretch;display:flex;align-items:center;justify-content:center;width:42px;border:none;border-radius:11px;background:transparent;color:var(--brand);font-size:1.25rem;font-weight:700;cursor:pointer;transition:background .15s;line-height:1}
.tab-nav-btn:hover:not(:disabled){background:var(--hover)}
.tab-nav-btn:disabled{opacity:.25;cursor:default}
html:not(.js) .tab-nav-btn{display:none}

/* ── Day panel ── */
.dpanel{display:none;background:var(--card);border:1px solid var(--line);border-radius:16px;padding:26px;box-shadow:var(--shadow-1)}
.dpanel.act{display:block;animation:panelIn .28s ease}
@keyframes panelIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:none}}
.dpanel.flash{animation:panelFlash 1s ease}
@keyframes panelFlash{0%{box-shadow:0 0 0 3px var(--ring)}100%{box-shadow:var(--shadow-1)}}
/* Static day title inside each panel — only shown when JS (and thus the
   tab strip) is unavailable, e.g. the Outlook attachment preview */
.dp-title{display:none;font-size:1.02rem;font-weight:800;color:var(--ink);margin-bottom:12px}
.dp-title .dt-sub{color:var(--mut);font-weight:600;font-size:.8rem;margin-left:8px}
html:not(.js) .dp-title{display:block}
.seg-alloc{background:#f97316}.seg-pick{background:#f59e0b}.seg-lset{background:#eab308;color:#1e293b}.seg-lent{background:#84cc16;color:#1e293b}.seg-ship{background:#22c55e}
.dp-head{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:16px;flex-wrap:wrap}
.dp-meta{font-size:.76rem;color:var(--mut);font-weight:600;white-space:nowrap;margin-left:auto}
.dp-chips{display:flex;gap:6px;flex-wrap:wrap}
.chip{display:inline-flex;align-items:center;gap:3px;padding:3px 9px;border-radius:6px;font-size:.7rem;font-weight:700}
.chip-alert{background:var(--bad-bg);color:var(--bad-t)}
.chip-mut{background:var(--mod-bg);color:var(--mut)}

/* ── Trailer cards ── */
.tg{display:grid;grid-template-columns:repeat(auto-fill,minmax(520px,1fr));gap:16px}
.tc{background:var(--card);border-radius:16px;overflow:hidden;border:1px solid var(--line);box-shadow:var(--shadow-1);transition:box-shadow .2s,transform .15s,border-color .2s;cursor:pointer}
.tc:hover{box-shadow:var(--shadow-3);transform:translateY(-2px);border-color:var(--brand-2)}
.tc-top{padding:18px 22px 0;display:flex;justify-content:space-between;align-items:flex-start}
.tc-id{font-weight:800;color:var(--ink);font-size:1.05rem}
.tc-bol{color:var(--mut);font-size:.86rem;margin-top:2px}
.tc-badges{display:flex;gap:6px;align-items:center;flex-shrink:0}
.badge{padding:3px 10px;border-radius:7px;font-size:.72rem;font-weight:700;white-space:nowrap}
.b-ship{background:var(--ok-bg);color:var(--ok-t)}
.b-prog{background:var(--warn-bg);color:var(--warn-t)}
.b-early{background:var(--orange-bg);color:var(--orange-t)}
.b-pend{background:var(--bad-bg);color:var(--bad-t)}
.b-mod{background:var(--mod-bg);color:var(--mod-t)}

/* ── Chevron stage steps ── */
.chevs{display:flex;margin:16px 22px 0;gap:3px}
.chev{flex:1;text-align:center;padding:6px 4px 6px 14px;font-size:.67rem;font-weight:600;background:var(--track);color:var(--faint);clip-path:polygon(0 0,calc(100% - 10px) 0,100% 50%,calc(100% - 10px) 100%,0 100%,10px 50%)}
.chev:first-child{clip-path:polygon(0 0,calc(100% - 10px) 0,100% 50%,calc(100% - 10px) 100%,0 100%);padding-left:10px;border-radius:6px 0 0 6px}
.chev:last-child{clip-path:polygon(0 0,100% 0,100% 100%,0 100%,10px 50%);border-radius:0 6px 6px 0}
.chev.on{color:#fff}.chev.s0.on{background:#f97316}.chev.s1.on{background:#f59e0b}
.chev.s2.on{background:#eab308;color:#1e293b}.chev.s3.on{background:#84cc16;color:#1e293b}.chev.s4.on{background:#22c55e}
.chev.cur{font-weight:800;text-decoration:underline;text-underline-offset:2px}

/* ── Card progress bar + legend + meta ── */
.tc-bar-wrap{padding:12px 22px 0}
.tc-bar{display:flex;height:26px;border-radius:7px;overflow:hidden;background:var(--track)}
.tc-seg{display:flex;align-items:center;justify-content:center;color:#fff;font-size:.7rem;font-weight:700;min-width:0;white-space:nowrap;overflow:hidden;padding:0 5px}
.tc-seg.seg-lset,.tc-seg.seg-lent{color:#1e293b}
.tc-legend{display:flex;gap:10px;padding:6px 22px 0;flex-wrap:wrap}
.lg-i{display:flex;align-items:center;gap:4px;font-size:.68rem;color:var(--mut)}
.lg-d{width:9px;height:9px;border-radius:2px;flex-shrink:0}
.tc-meta{display:grid;grid-template-columns:1fr 1fr;gap:5px 16px;padding:16px 22px;margin-top:12px;background:var(--card-2);border-top:1px solid var(--line-2);font-size:.86rem;color:var(--mut)}
.tc-meta b{font-weight:600;color:var(--ink-2)}

/* ── Mini timeline (plan vs actual) ── */
.tc-tl{margin:10px 22px 0;height:18px;background:var(--track);border-radius:4px;position:relative;overflow:visible}
.tc-tl-range{position:absolute;top:0;height:100%;background:rgba(59,130,246,.15);border-radius:4px}
.tc-tl-marker{position:absolute;top:0;width:3px;height:100%;border-radius:2px}
.tc-tl-plan{background:#3b82f6}
.tc-tl-actual{background:#22c55e}
.tc-tl-actual.late{background:#ef4444}
.tc-tl-lbl{position:absolute;top:-14px;font-size:.58rem;font-weight:600;white-space:nowrap;transform:translateX(-50%)}
.tc-tl-lbl.plan-lbl{color:#3b82f6}
.tc-tl-lbl.act-lbl{color:#22c55e}
.tc-tl-lbl.act-lbl.late{color:#ef4444}
.tc-tl-legend{display:flex;gap:10px;justify-content:flex-end;padding:2px 22px 0;font-size:.6rem;color:var(--faint)}
.tc-tl-legend span::before{content:'';display:inline-block;width:8px;height:3px;border-radius:1px;margin-right:3px;vertical-align:middle}
.tc-tl-legend .tl-plan::before{background:#3b82f6}
.tc-tl-legend .tl-act::before{background:#22c55e}

/* ── Misc ── */
.no-data{padding:32px;color:var(--faint);font-style:italic;text-align:center;background:var(--card);border:1px dashed var(--line);border-radius:16px}
.tc-focus{outline:3px solid var(--brand-2);outline-offset:2px;box-shadow:0 0 0 6px var(--ring)!important}
.tc-tip{position:fixed;z-index:900;background:#0f172a;color:#fff;padding:10px 14px;border-radius:9px;font-size:.78rem;pointer-events:none;opacity:0;transition:opacity .15s;max-width:280px;box-shadow:0 8px 24px rgba(0,0,0,.3);line-height:1.5;border:1px solid rgba(148,163,184,.25)}
.tc-tip.show{opacity:1}
.tc-tip b{color:#93c5fd}

/* ── Modal ── */
.modal-bg{position:fixed;inset:0;background:rgba(8,12,22,.55);backdrop-filter:blur(4px);z-index:1000;display:none;align-items:center;justify-content:center;padding:20px}
.modal-bg.show{display:flex}
.modal{background:var(--card);border:1px solid var(--line);border-radius:18px;max-width:950px;width:95%;max-height:90vh;overflow-y:auto;position:relative;box-shadow:var(--shadow-3);animation:mIn .2s ease}
@keyframes mIn{from{opacity:0;transform:scale(.95) translateY(10px)}to{opacity:1;transform:scale(1) translateY(0)}}
.modal-hd{padding:24px 28px 18px;border-bottom:1px solid var(--line-2);display:flex;justify-content:space-between;align-items:flex-start}
.m-id{font-size:1.3rem;font-weight:800;color:var(--ink)}
.m-bol{font-size:.92rem;color:var(--mut);margin-top:2px}
.m-badges{display:flex;gap:8px;margin-top:8px}
.modal-x{width:34px;height:34px;border-radius:9px;border:none;background:var(--track);color:var(--mut);font-size:1.2rem;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:background .15s,color .15s;flex-shrink:0}
.modal-x:hover{background:var(--line);color:var(--ink)}
.modal-body{padding:24px 28px}
.m-section{font-size:.76rem;font-weight:800;color:var(--brand);text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px}
.m-sg{display:grid;gap:6px;margin-bottom:22px}
.m-sr{display:grid;grid-template-columns:100px 45px 52px 1fr;gap:10px;align-items:center;font-size:.88rem;padding:6px 0}
.m-sr:not(:last-child){border-bottom:1px solid var(--line-2)}
.m-sn{font-weight:600;color:var(--ink-2)}
.m-sc{font-weight:700;color:var(--brand);text-align:right}
.m-sp{color:var(--mut);text-align:right;font-size:.82rem}
.m-sb{height:14px;border-radius:4px;background:var(--track);overflow:hidden}
.m-sf{height:100%;border-radius:4px}
.m-bar{display:flex;height:34px;border-radius:9px;overflow:hidden;background:var(--track);margin-bottom:22px}
.m-bar .seg{display:flex;align-items:center;justify-content:center;color:#fff;font-size:.78rem;font-weight:700;min-width:0;white-space:nowrap;overflow:hidden;padding:0 6px}
.m-meta{display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:.92rem;color:var(--ink-2)}
.m-meta b{color:var(--ink);font-weight:600}
.m-tbl-wrap{max-height:400px;overflow-y:auto;border:1px solid var(--line);border-radius:10px;margin-bottom:22px}
.m-tbl{width:100%;border-collapse:collapse;font-size:.78rem}
.m-tbl th{position:sticky;top:0;background:var(--thead);color:#fff;padding:8px 10px;text-align:left;font-weight:600;font-size:.72rem;white-space:nowrap;cursor:pointer;user-select:none}
.m-tbl th:hover{filter:brightness(1.18)}
.m-tbl td{padding:6px 10px;border-bottom:1px solid var(--line-2)}
.m-tbl tr:nth-child(even){background:var(--even)}
.m-tbl tbody tr:hover{background:var(--hover)}
.m-dot{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:4px;vertical-align:middle}
.m-summary{display:flex;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.m-row{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:22px}
@media(max-width:700px){.m-row{grid-template-columns:1fr}}
.m-sum-card{padding:10px 16px;border-radius:10px;background:var(--card-2);border:1px solid var(--line-2);text-align:center;flex:1 1 90px}
.m-sum-card .m-sv{font-size:1.1rem;font-weight:800;color:var(--brand)}
.m-sum-card .m-sl{font-size:.66rem;color:var(--mut);text-transform:uppercase;letter-spacing:.05em}
.sort-arrow{font-size:.65rem;margin-left:3px;opacity:.8}
.m-filter{padding:7px 11px;border:1px solid var(--line);border-radius:8px;font-size:.78rem;width:100%;margin-bottom:10px;outline:none;background:var(--card-2);color:var(--txt)}
.m-filter:focus{border-color:var(--brand-2);box-shadow:0 0 0 3px var(--ring)}
.m-filter::placeholder{color:var(--faint)}

/* ── Unshipped section ── */
.us-controls{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap}
.us-input{flex:1;min-width:200px;max-width:360px;padding:8px 12px;border:1px solid var(--line);border-radius:9px;font-size:.85rem;background:var(--card);color:var(--txt);outline:none}
.us-input:focus{border-color:var(--brand-2);box-shadow:0 0 0 3px var(--ring)}
.us-input::placeholder{color:var(--faint)}
.us-select{padding:8px 12px;border:1px solid var(--line);border-radius:9px;font-size:.85rem;background:var(--card);color:var(--txt);outline:none}
.us-select:focus{border-color:var(--brand-2)}
.us-count{font-size:.7rem;color:var(--faint);margin-bottom:8px}
.us-overdue{color:var(--alert);font-weight:700}
.us-btn{padding:8px 14px;border:1px solid var(--line);border-radius:9px;background:var(--card);color:var(--brand);font-size:.78rem;font-weight:700;cursor:pointer;transition:background .15s}
.us-btn:hover{background:var(--hover)}
/* The section pre-renders today's rows so it reads without JS — but the
   filter/export controls and sort arrows only work with JS. Hide them. */
html:not(.js) .us-controls{display:none}
html:not(.js) .sort-arrow{display:none}

/* ── Footer: keyboard hints + timestamp ── */
.foot{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-top:14px;flex-wrap:wrap}
.kbd-hints{display:flex;gap:14px;font-size:.7rem;color:var(--faint);flex-wrap:wrap}
.kbd{display:inline-block;min-width:18px;text-align:center;padding:1px 5px;border:1px solid var(--line);border-bottom-width:2px;border-radius:5px;background:var(--card);font-size:.66rem;font-weight:700;color:var(--mut);margin-right:2px}
html:not(.js) .kbd-hints{display:none}
.ts{font-size:.72rem;color:var(--faint);text-align:right}
.stale{color:var(--warn-t);background:var(--warn-bg);padding:2px 8px;border-radius:6px;font-weight:700;margin-left:8px}

/* ── Print ── */
@media print{
  .modal-bg,.tabs-wrap,.tc-tip,.us-controls,.ov-btn,.us-btn,.kbd-hints{display:none!important}
  .dpanel{display:block!important;page-break-inside:avoid;box-shadow:none;border:1px solid #ccc;margin-bottom:14px}
  .tc{page-break-inside:avoid}
  body{background:#fff!important;color:#1e293b!important}
  .hdr{-webkit-print-color-adjust:exact;print-color-adjust:exact}
  .sec{page-break-before:auto}
}

/* ════════════════════════════════════════════════════
   PHONE LAYOUT — activates only on small touch screens.
   Desktop/tablet are untouched: every rule lives inside
   the query.
   ════════════════════════════════════════════════════ */
@keyframes sheetIn{from{transform:translateY(100%)}to{transform:translateY(0)}}

@media (max-width:600px){
  /* Guard: a single overflowing element must not widen the whole page
     (it clips the right edge of everything else on a phone) */
  body{overflow-x:hidden}
  .db{padding:10px}

  /* Header → stacked title with a compact 3-up KPI strip */
  .hdr{flex-direction:column;align-items:stretch;gap:12px;padding:16px 18px;border-radius:14px;margin-bottom:16px}
  .hdr h1{font-size:1.25rem}
  .hdr .sub{font-size:.88rem}
  .hero{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;width:100%}
  .hero-card{min-width:0;width:auto;padding:9px 6px;text-align:center}
  .hero-card .v{font-size:1.1rem}
  .hero-card .l{font-size:.52rem}

  .sec{margin-bottom:20px}
  .sec-hd{font-size:.95rem;margin-bottom:12px}

  /* Snapshot table: side-scrolls when the 7 columns don't fit */
  .ov-snap{padding:14px 12px;overflow-x:auto;-webkit-overflow-scrolling:touch}
  .ov-tbl{font-size:.78rem}
  .ov-tbl th,.ov-tbl td{padding:6px 7px}

  /* Sticky day bar: snug tabs, compact arrows */
  .tabs-wrap{top:6px}
  .tabs{gap:4px;padding:4px;border-radius:13px}
  .tab{flex:1 1 24%;padding:8px 3px 13px;border-radius:9px}
  .tab:hover{background:none}
  .tab.act:hover{background:var(--tab-act)}
  .tab-day{font-size:.54rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .tab-date{font-size:.8rem;margin:2px 0 1px}
  .tab-mini{font-size:.54rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .tab-today::after{width:6px;height:6px;top:4px;right:4px}
  .tab-prog{left:6px;right:6px;bottom:4px;height:2px}
  .tab-nav-btn{width:30px;font-size:1.05rem}

  /* Day panel */
  .dpanel{padding:14px 12px;border-radius:14px}
  .dp-chips{gap:5px}
  .chip{font-size:.66rem}
  .dp-meta{white-space:normal;margin-left:0}

  /* Trailer cards: full-width, tighter, with tap feedback */
  .tg{grid-template-columns:1fr;gap:12px}
  .tc{border-radius:12px}
  .tc:hover{transform:none;box-shadow:var(--shadow-1);border-color:var(--line)}
  .tc:active{transform:scale(.992)}
  .tc-top{padding:14px 16px 0}
  .tc-id{font-size:1rem}
  .tc-bol{font-size:.82rem}
  .tc-badges{flex-wrap:wrap;justify-content:flex-end;gap:5px}

  /* Chevron stepper → slim, label-free colour strip
     (per-stage counts remain in the legend just below) */
  .chevs{margin:12px 16px 0;gap:2px}
  .chev{font-size:0;padding:0;height:9px;clip-path:polygon(0 0,calc(100% - 6px) 0,100% 50%,calc(100% - 6px) 100%,0 100%,6px 50%)}
  .chev:first-child{clip-path:polygon(0 0,calc(100% - 6px) 0,100% 50%,calc(100% - 6px) 100%,0 100%);padding:0;border-radius:5px 0 0 5px}
  .chev:last-child{clip-path:polygon(0 0,100% 0,100% 100%,0 100%,6px 50%);border-radius:0 5px 5px 0}

  .tc-bar-wrap{padding:10px 16px 0}
  .tc-bar{height:24px}
  .tc-legend{padding:6px 16px 0;gap:8px}
  /* Single column: Order #/BOL strings have no break points and two
     columns at phone width force the page wider than the viewport */
  .tc-meta{grid-template-columns:1fr;padding:14px 16px;font-size:.82rem;gap:5px 14px}
  .tc-tl{margin:12px 16px 0}
  .tc-tl-legend{padding:2px 16px 0}

  /* Modal → bottom sheet that slides up, near full-height */
  .modal-bg{padding:0;align-items:flex-end}
  .modal{width:100%;max-width:100%;max-height:92vh;border-radius:18px 18px 0 0;animation:sheetIn .25s ease}
  .modal-hd{padding:18px 18px 14px;position:sticky;top:0;background:inherit;z-index:5;border-radius:18px 18px 0 0}
  .modal-body{padding:18px 16px}
  .m-id{font-size:1.15rem}
  .m-summary{gap:7px}
  .m-sum-card{flex:1 1 calc(33.333% - 7px);min-width:0;padding:8px 5px}
  .m-sum-card .m-sv{font-size:.95rem}
  .m-row{grid-template-columns:1fr;gap:16px;margin-bottom:16px}
  .m-sr{grid-template-columns:84px 34px 44px 1fr;gap:8px;font-size:.8rem}
  .m-meta{grid-template-columns:1fr}
  .m-tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch}
  .m-tbl{font-size:.72rem}
  .m-filter{font-size:.85rem;padding:9px 12px}

  /* Unshipped list */
  .us-controls{flex-direction:column}
  .us-select{width:100%}
  .us-input{width:100%;max-width:none}

  /* No hover tooltip on touch — hide it */
  .tc-tip{display:none!important}

  .foot{justify-content:center}
  .kbd-hints{display:none}
  .ts{text-align:center}
}

/* Very small phones (iPhone SE etc.) */
@media (max-width:380px){
  .hdr h1{font-size:1.15rem}
  .hero-card .v{font-size:1rem}
  .tab-date{font-size:.74rem}
  .tc-meta{grid-template-columns:1fr}
  .m-sum-card{flex-basis:calc(50% - 7px)}
}

/* ════════════════════════════════════════════════════
   PASSIVE PHONE VIEW — the Outlook attachment preview on
   a phone runs no JS, so interactive chrome is dead
   weight: hide it and let the page read top-to-bottom as
   a static brief. The interactive phone layout above
   still applies whenever the file opens in a real
   browser (html.js present).
   ════════════════════════════════════════════════════ */
@media (max-width:600px){
  /* Day tabs can't switch panels without JS — today's panel is already open
     and carries its own .dp-title */
  html:not(.js) .tabs-wrap{display:none}
  /* Snapshot rows aren't tappable: drop the chevron column + pointer */
  html:not(.js) .ov-go,html:not(.js) .ov-go-h{display:none}
  html:not(.js) .ov-tbl tbody tr{cursor:default}
  /* Trailer cards are read-only: no tap/hover feedback */
  html:not(.js) .tc{cursor:default}
  html:not(.js) .tc:active{transform:none}
  html:not(.js) .tc:hover{box-shadow:var(--shadow-1);border-color:var(--line);transform:none}
  /* Unshipped table scrolls natively; sort headers are inert */
  html:not(.js) .m-tbl-wrap{-webkit-overflow-scrolling:touch}
  html:not(.js) .m-tbl th{cursor:default}
}

/* ── Reduced motion ── */
@media (prefers-reduced-motion:reduce){
  *,*::before,*::after{animation:none!important;transition:none!important}
  html{scroll-behavior:auto}
}
"""


def _trailer_status(t: dict) -> tuple[str, str]:
    stages = t["stages"]
    total = t["module_count"]
    shipped = stages.get("Shipped", 0)
    if shipped == total and total > 0:
        return "Shipped", "b-ship"
    if shipped > 0:
        return "In Progress", "b-prog"
    if t["trailer"] == "N/A":
        return "Pending", "b-pend"
    for s in reversed(STAGE_NAMES[1:]):
        if stages.get(s, 0) > 0:
            return "Processing", "b-prog"
    return "Allocated", "b-early"


def _render_timeline(plan_min: str, plan_max: str, actual: str, is_late: bool = False) -> str:
    """Return HTML for a mini plan-vs-actual timeline bar on a trailer card.

    `is_late` (the date-based flag from compute_summary1) decides whether the
    actual marker is colored red, so the timeline agrees with the day chip and
    the modal — the bar still plots time-of-day, but a same-day late-in-the-day
    load is on time."""
    def _to_minutes(dt_str: str) -> int | None:
        if not dt_str or dt_str == "N/A":
            return None
        parts = dt_str.split()
        if len(parts) < 2:
            return None
        try:
            h, m = parts[1].split(":")
            return int(h) * 60 + int(m)
        except (ValueError, IndexError):
            return None

    DAY_START, DAY_END = 360, 1320  # 06:00 to 22:00
    span = DAY_END - DAY_START

    pm = _to_minutes(plan_min)
    px = _to_minutes(plan_max)
    pa = _to_minutes(actual)

    if pm is None:
        return ""

    def clamp_pct(minutes: int) -> float:
        return max(0.0, min(100.0, (minutes - DAY_START) / span * 100))

    plan_pct = clamp_pct(pm)
    plan_max_pct = clamp_pct(px) if px is not None else plan_pct

    # Build the HTML
    parts = ['<div class="tc-tl">']

    # Range shading between plan_min and plan_max
    if plan_max_pct > plan_pct:
        parts.append(
            f'<div class="tc-tl-range" style="left:{plan_pct:.1f}%;width:{plan_max_pct - plan_pct:.1f}%"></div>'
        )

    # Plan marker
    pm_h, pm_m = pm // 60, pm % 60
    parts.append(
        f'<div class="tc-tl-marker tc-tl-plan" style="left:{plan_pct:.1f}%">'
        f'<span class="tc-tl-lbl plan-lbl">{pm_h}:{pm_m:02d}</span></div>'
    )

    # Actual marker
    if pa is not None:
        act_pct = clamp_pct(pa)
        late_cls = " late" if is_late else ""
        pa_h, pa_m = pa // 60, pa % 60
        parts.append(
            f'<div class="tc-tl-marker tc-tl-actual{late_cls}" style="left:{act_pct:.1f}%">'
            f'<span class="tc-tl-lbl act-lbl{late_cls}">{pa_h}:{pa_m:02d}</span></div>'
        )

    parts.append('</div>')

    # Legend
    legend_items = '<span class="tl-plan">Plan</span>'
    if pa is not None:
        legend_items += '<span class="tl-act">Actual</span>'
    parts.append(f'<div class="tc-tl-legend">{legend_items}</div>')

    return "\n".join(parts)


def render_overview_table(
    date2: date,
    summary2_rows: list[list],
    summary1_blocks: list[tuple[date, list[dict]]],
) -> str:
    """Compact 9-business-day table (the Daily Snapshot). Each row jumps to that
    day's tab via ``jumpToDay``. Renders without JS too (Outlook preview / print);
    JS only adds the click-to-jump behavior. Mirrors the PNG snapshot table."""
    L = ['<div class="ov-snap">']
    L.append('<div class="ov-snap-title">')
    L.append('<span class="d">Click a row to open that day below</span>')
    L.append('<div class="ov-actions">')
    L.append('<button class="ov-btn ghost" id="copy-btn" onclick="copySummary(this)" title="Copy the KPI tables (selected day &#177;1 business day) in the original email format and open a preview in a new tab">Copy Email Summary</button>')
    L.append('<button class="ov-btn" onclick="jumpToDay(TODAY_IDX)">Jump to Today &rarr;</button>')
    L.append("</div>")
    L.append("</div>")
    L.append(
        '<table class="ov-tbl"><thead><tr>'
        "<th>Day</th><th>Date</th><th>Plan</th><th>Shipped</th>"
        '<th>On-Time %</th><th>Short</th><th>Cncl</th><th class="ov-go-h"></th></tr></thead><tbody>'
    )
    for i, (row, (td, _trailers)) in enumerate(zip(summary2_rows, summary1_blocks)):
        plan = _cell_int(row[9])
        ship = _cell_int(row[6])
        pct = (ship / plan * 100) if plan else 0
        short = _cell_int(row[7])
        cncl = _cell_int(row[8])
        is_future = td > date2
        if td == date2:
            # ov-sel tracks the selected tab; today is the selection on load
            cls = "ov-today-row ov-sel"
        elif is_future:
            cls = "ov-future"
        else:
            cls = ""
        # Ratio is only meaningful once the ship day has arrived.
        if plan and not is_future:
            ratio_disp = f'<span class="ov-ratio" style="background:{_pct_color(pct)}">{pct:.0f}%</span>'
        else:
            ratio_disp = '<span class="dash">&mdash;</span>'
        short_disp = f'<span class="neg">{short}</span>' if short else "0"
        cncl_disp = f'<span class="neg">{cncl}</span>' if cncl else "0"
        L.append(
            f'<tr class="{cls}" id="ov-row-{i}" onclick="jumpToDay({i})" title="Open {html_escape(row[0])} below">'
            f'<td class="ov-day">{html_escape(_day_label(td, date2, short=True))}</td>'
            f'<td>{html_escape(row[0])}</td>'
            f"<td>{plan}</td><td>{ship}</td>"
            f"<td>{ratio_disp}</td><td>{short_disp}</td><td>{cncl_disp}</td>"
            f'<td class="ov-go">&rsaquo;</td></tr>'
        )
    L.append("</tbody></table>")
    L.append("</div>")  # ov-snap
    return "\n".join(L)


def _plan_ship_overdue(plan_ship, gen_dt: datetime) -> bool:
    """Mirror of the JS planTs/overdue check: 'MM/DD/YYYY[ HH:MM]' earlier
    than the report-generation time. Unparseable values are not overdue."""
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})(?: (\d{2}):(\d{2}))?", str(plan_ship or ""))
    if not m:
        return False
    mo, dd, yy, hh, mi = m.groups()
    try:
        return datetime(int(yy), int(mo), int(dd), int(hh or 0), int(mi or 0)) < gen_dt
    except ValueError:
        return False


def _render_unshipped_rows(records: list[dict], gen_dt: datetime) -> str:
    """Static <tr> markup for the Unshipped table — the same cells the JS
    renderUnshipped builds, so the no-JS view (Outlook preview) matches what
    the browser shows after JS re-renders."""
    rows = []
    for r in records:
        plan = r["Plan_Ship"]
        if _plan_ship_overdue(plan, gen_dt):
            plan_td = (
                '<td class="us-overdue" title="Past planned ship time">'
                f"{html_escape(plan)} &#9888;</td>"
            )
        else:
            plan_td = f"<td>{html_escape(plan)}</td>"
        dot = UNSHIPPED_STATUS_COLORS.get(r["Status"], "#94a3b8")
        rows.append(
            f'<tr><td>{html_escape(r["TargetDate"])}</td>'
            f'<td>{html_escape(r["CUSTOMER ORDER NO."])}</td>'
            f"{plan_td}"
            f'<td>{html_escape(r["PRODUCT NO."])}</td>'
            f'<td>{html_escape(r["QUANTITY"])}</td>'
            f'<td><span class="m-dot" style="background:{dot}"></span>{html_escape(r["Status"])}</td>'
            f'<td>{html_escape(r["Module Count"])}</td></tr>'
        )
    return "".join(rows)


def build_html(
    date2: date,
    summary1_blocks: list[tuple[date, list[dict]]],
    summary2_rows: list[list],
    unshipped: pd.DataFrame | None = None,
    data_asof: str | None = None,
    data_stale: bool = False,
) -> str:
    css = _report_css()
    now = datetime.now()
    now_str = now.strftime("%m/%d/%Y %H:%M")
    seg_colors = list(STAGE_COLORS.values())
    seg_cls = ["seg-alloc", "seg-pick", "seg-lset", "seg-lent", "seg-ship"]

    # ── Today's hero stats ──
    today_row = next((r for r, (d, _) in zip(summary2_rows, summary1_blocks) if d == date2), summary2_rows[0] if summary2_rows else None)
    today_trailers = next((tr for d, tr in summary1_blocks if d == date2), [])
    total_plan = _cell_int(today_row[9]) if today_row else 0
    total_ship = _cell_int(today_row[6]) if today_row else 0
    total_trailers = len(today_trailers)
    overall_pct = (total_ship / total_plan * 100) if total_plan else 0
    pct_color = _pct_color(overall_pct)

    lines = [
        "<!doctype html>", '<html lang="en">', "<head>",
        '<meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        f"<title>SITE1 Shipping KPI - {format_date(date2)}</title>",
        # Mark that scripts run, so JS-only controls (pagination nav) show only
        # when JS is available. In no-JS viewers (Outlook preview) they stay hidden.
        '<script>document.documentElement.className+=" js"</script>',
        f"<style>{css}</style>", "</head>", "<body>", '<div class="db">',
    ]

    # ── Header with hero stats ──
    lines.append('<div class="hdr">')
    lines.append('<div class="hdr-left">')
    lines.append("<h1>SITE1 Shipping KPI</h1>")
    trailer_word = "trailer" if total_trailers == 1 else "trailers"
    lines.append(
        f'<div class="sub"><span id="hdr-date">{date2.strftime("%A")}, {format_date(date2)}</span>'
        f' &middot; <span id="hdr-trail">{total_trailers} {trailer_word}</span></div>'
    )
    lines.append("</div>")
    lines.append('<div class="hero">')
    lines.append(f'<div class="hero-card"><div class="v" id="hdr-plan">{total_plan}</div><div class="l">Plan Ship</div></div>')
    lines.append(f'<div class="hero-card"><div class="v" id="hdr-ship">{total_ship}</div><div class="l">Actual Ship</div></div>')
    lines.append(
        f'<div class="hero-card"><div class="v" id="hdr-ratio" style="color:{pct_color}">{overall_pct:.1f}%</div>'
        f'<div class="l">On-Time Shipping</div>'
        f'<div class="hero-bar"><i id="hdr-ratio-bar" style="width:{min(overall_pct, 100):.0f}%;background:{pct_color}"></i></div></div>'
    )
    lines.append("</div>")
    lines.append("</div>")

    # Tooltip container
    lines.append('<div class="tc-tip" id="tc-tip"></div>')

    # Build per-day stats array for JS header updates
    day_stats_js = []
    for row, (td, trl) in zip(summary2_rows, summary1_blocks):
        ds_plan = sum(_cell_int(row[si]) for si in (2, 3, 4, 5, 6))
        ds_ship = _cell_int(row[6])
        ds_trail = len(trl)
        ds_pct = (ds_ship / ds_plan * 100) if ds_plan > 0 else 0
        is_future = 1 if td > date2 else 0
        day_stats_js.append('{' + f'"date":"{format_date(td)}","dow":"{td.strftime("%A")}","plan":{ds_plan},"ship":{ds_ship},"trail":{ds_trail},"pct":{ds_pct:.1f},"future":{is_future}' + '}')
    lines.append(f'<script>var DS=[{",".join(day_stats_js)}];</script>')

    # Legacy email tables for the Copy Email Summary button: the exact
    # Summary2 / per-Trailer/BOL rows the original generate_daily_shipping_kpi.py
    # report contained, indexed by day. S1 rows follow LEGACY_SUMMARY1_HEADERS.
    s1_js = [
        [
            [t["plan_min"], t["plan_max"], t["trailer"], t["bol"],
             t["actual_load"], t["order_max"], t["skid_count"], t["module_count"]]
            for t in trl
        ]
        for _td, trl in summary1_blocks
    ]
    lines.append(
        "<script>"
        f"var S2={json.dumps(summary2_rows)};"
        f"var S2H={json.dumps(LEGACY_SUMMARY2_HEADERS)};"
        f"var S1H={json.dumps(LEGACY_SUMMARY1_HEADERS)};"
        f"var S1={json.dumps(s1_js)};"
        "</script>"
    )

    # ── Pagination setup ──
    page_size = 3
    today_idx = next((i for i, (d, _) in enumerate(summary1_blocks) if d == date2), 0)
    total_days = len(summary1_blocks)
    total_pages = max(1, (total_days + page_size - 1) // page_size)
    default_page = today_idx // page_size
    lines.append(
        f'<script>var TODAY_IDX={today_idx};var PAGE_SIZE={page_size};'
        f'var TOTAL_PAGES={total_pages};var DEFAULT_PAGE={default_page};'
        # Report-generation timestamp (ms) — overdue checks compare against
        # the data snapshot's time, not the viewer's clock.
        f'var GEN_TS={int(now.timestamp() * 1000)};</script>'
    )

    # ── Daily Snapshot: 9-business-day overview (mirrors the PNG snapshot) ──
    lines.append('<div class="sec">')
    lines.append('<div class="sec-hd">Daily Snapshot<span class="sec-sub">9 business days</span></div>')
    lines.append(render_overview_table(date2, summary2_rows, summary1_blocks))
    lines.append("</div>")

    # ── Tab-based Day Selector ──
    lines.append('<div class="sec" id="ship-sec">')
    lines.append('<div class="sec-hd">Shipping Overview</div>')

    chip_labels = ["Alloc", "Picked", "Load.Set", "Load.Ent", "Shipped"]

    # Sticky day-switcher: the Prev/Next arrows live inside the tab strip so
    # the whole bar travels together when it sticks to the top on scroll.
    lines.append('<div class="tabs-wrap"><div class="tabs">')
    if total_pages > 1:
        lines.append('<button class="tab-nav-btn" id="nav-prev" onclick="changePage(-1)" title="Earlier days">&#8249;</button>')
    for ti, (row, (target_date, _trailers)) in enumerate(zip(summary2_rows, summary1_blocks)):
        day_label = _day_label(target_date, date2)
        tab_extra = " tab-today" if target_date == date2 else ""
        act = " act" if target_date == date2 else ""
        plan_t = _cell_int(row[9])
        try:
            rv = float(str(row[10]).replace("%", ""))
        except (ValueError, TypeError):
            rv = 0
        # Thin ship-ratio strip along the tab's bottom edge; future days keep
        # just the empty track (ratio is not meaningful until the day arrives).
        if target_date <= date2 and plan_t > 0:
            prog = f'<div class="tab-prog"><i style="width:{min(rv, 100):.0f}%;background:{_pct_color(rv)}"></i></div>'
        else:
            prog = '<div class="tab-prog"></div>'
        lines.append(
            f'<div class="tab{act}{tab_extra}" id="tab-{ti}" onclick="showDay({ti})">'
            f'<div class="tab-day">{day_label}</div>'
            f'<div class="tab-date">{html_escape(row[0])}</div>'
            f'<div class="tab-mini">{plan_t} modules &bull; {rv:.0f}%</div>'
            f'{prog}</div>'
        )
    if total_pages > 1:
        lines.append('<button class="tab-nav-btn" id="nav-next" onclick="changePage(1)" title="Later days">&#8250;</button>')
    lines.append("</div></div>")

    # Build day panels
    for pi, (row, (target_date, trailers)) in enumerate(zip(summary2_rows, summary1_blocks)):
        act = " act" if target_date == date2 else ""
        stage_vals = [_cell_int(row[si]) for si in (2, 3, 4, 5, 6)]
        no_proc = _cell_int(row[1])
        shortage = _cell_int(row[7])
        canceled = _cell_int(row[8])

        lines.append(f'<div class="dpanel{act}" id="panel-{pi}">')

        # Static day title — visible only without JS (Outlook preview), where
        # the tab strip is hidden and today's panel shows on its own.
        lines.append(
            f'<div class="dp-title">{_day_label(target_date, date2)}'
            f'<span class="dt-sub">{target_date.strftime("%A")}, {html_escape(row[0])}</span></div>'
        )

        # Panel header: stage/alert chips on the left, a quiet per-day meta
        # line on the right. (The full-width stage bar was removed at the
        # operator's request — each trailer card has its own bar.)
        lines.append('<div class="dp-head">')
        lines.append('<div class="dp-chips">')
        for j, v in enumerate(stage_vals):
            if v > 0:
                lines.append(f'<div class="chip" style="background:{seg_colors[j]}20;color:{seg_colors[j]}">{chip_labels[j]}: {v}</div>')
        if no_proc > 0:
            lines.append(f'<div class="chip chip-mut">No Proc: {no_proc}</div>')
        if shortage > 0:
            lines.append(f'<div class="chip chip-alert">Shortage: {shortage}</div>')
        if canceled > 0:
            lines.append(f'<div class="chip chip-alert">Canceled: {canceled}</div>')
        # Operational attention chips: trailers with no assignment yet, and
        # trailers whose latest load ran past the first planned time. These are
        # trailer-level counts, so they carry the "trailer(s)" unit — the stage
        # chips above (Shipped, etc.) are module counts, and an unlabeled
        # "Late: 1" next to "Shipped: 90" reads as a 91st module.
        n_pending = sum(1 for t in trailers if t["trailer"] == "N/A")
        n_late = sum(1 for t in trailers if t.get("is_late"))
        if n_pending > 0:
            lines.append(f'<div class="chip chip-mut">Unassigned: {n_pending} trailer{"s" if n_pending != 1 else ""}</div>')
        if n_late > 0:
            lines.append(f'<div class="chip chip-alert">Late: {n_late} trailer{"s" if n_late != 1 else ""}</div>')
        lines.append("</div>")
        n_tr = len(trailers)
        plan_mod = _cell_int(row[9])
        lines.append(
            f'<div class="dp-meta">{n_tr} trailer{"s" if n_tr != 1 else ""}'
            f' &middot; {plan_mod} modules planned</div>'
        )
        lines.append("</div>")


        # Trailer cards
        if not trailers:
            lines.append('<div class="no-data">No Containers Scheduled</div>')
        else:
            lines.append('<div class="tg">')
            for t in trailers:
                total = t["module_count"]
                is_na = t["trailer"] == "N/A"
                status_text, status_cls = _trailer_status(t)
                trailer_display = "Pending Assignment" if is_na else html_escape(t["trailer"])
                bol_display_raw = "" if is_na else html_escape(t["bol"])
                sj = json.dumps(t["stages"])
                dj = html_escape(json.dumps(t.get("modules", [])))
                tw = t.get("total_weight", 0)
                tq = t.get("total_qty", 0)
                late_min = t.get("late_min")
                late_attr = "" if late_min is None else str(late_min)
                lines.append(
                    f'<div class="tc" onclick="openModal(this)" '
                    f'data-trailer="{trailer_display}" data-bol="{bol_display_raw}" '
                    f'data-status="{status_text}" data-status-cls="{status_cls}" '
                    f'data-modules="{total}" data-skids="{t["skid_count"]}" '
                    f'data-weight="{tw}" data-qty="{tq}" '
                    f'data-plan-min="{html_escape(t["plan_min"])}" '
                    f'data-plan-max="{html_escape(t["plan_max"])}" '
                    f'data-actual="{html_escape(t["actual_load"])}" '
                    f'data-order="{html_escape(t["order_max"])}" '
                    f'data-late="{late_attr}" '
                    f'data-islate="{1 if t.get("is_late") else 0}" '
                    f"data-stages='{sj}' data-detail='{dj}'>"
                )
                bol_display = "" if is_na else html_escape(t["bol"])
                lines.append('<div class="tc-top">')
                lines.append(f'<div><div class="tc-id">{trailer_display}</div>')
                if bol_display:
                    lines.append(f'<div class="tc-bol">BOL {bol_display}</div>')
                lines.append("</div>")
                lines.append(f'<div class="tc-badges"><span class="badge {status_cls}">{status_text}</span>'
                             f'<span class="badge b-mod">{total} modules</span></div>')
                lines.append("</div>")
                highest = -1
                for si, sn in enumerate(STAGE_NAMES):
                    if t["stages"].get(sn, 0) > 0:
                        highest = si
                lines.append('<div class="chevs">')
                for si, sn in enumerate(STAGE_NAMES):
                    on = " on" if si <= highest else ""
                    cur = " cur" if si == highest else ""
                    lines.append(f'<div class="chev s{si}{on}{cur}">{sn}</div>')
                lines.append("</div>")
                lines.append('<div class="tc-bar-wrap"><div class="tc-bar">')
                for j, stage in enumerate(STAGE_NAMES):
                    count = t["stages"].get(stage, 0)
                    if count > 0 and total > 0:
                        pct = count / total * 100
                        lbl = str(count) if pct >= 8 else ""
                        lines.append(f'<div class="tc-seg {seg_cls[j]}" style="width:{pct:.1f}%" title="{stage}: {count}">{lbl}</div>')
                lines.append("</div></div>")
                active = [(j, s, t["stages"].get(s, 0)) for j, s in enumerate(STAGE_NAMES) if t["stages"].get(s, 0) > 0]
                if active:
                    lines.append('<div class="tc-legend">')
                    for j, stage, count in active:
                        lines.append(f'<div class="lg-i"><div class="lg-d" style="background:{seg_colors[j]}"></div>{stage}: {count}</div>')
                    lines.append("</div>")
                lines.append('<div class="tc-meta">')
                plan_min_v = t["plan_min"]
                plan_max_v = t["plan_max"]
                if plan_min_v == plan_max_v:
                    plan_display = html_escape(plan_min_v)
                elif plan_min_v[:10] == plan_max_v[:10]:
                    plan_display = f'{html_escape(plan_min_v)} &rarr; {html_escape(plan_max_v[11:])}'
                else:
                    plan_display = f'{html_escape(plan_min_v)} &rarr; {html_escape(plan_max_v)}'
                weight_str = f'{tw:,.1f} lbs' if tw else '—'
                lines.append(f'<div style="grid-column:1/-1"><b>Plan Ship Time:</b> {plan_display}</div>')
                lines.append(f'<div><b>Actual Ship Time:</b> {html_escape(t["actual_load"])}</div>')
                lines.append(f'<div><b>Order #:</b> {html_escape(t["order_max"])}</div>')
                lines.append(f'<div><b>Skid Count:</b> {t["skid_count"]}</div>')
                lines.append(f'<div><b>Total Weight:</b> {weight_str}</div>')
                lines.append(f'<div><b>Module Count:</b> {total}</div>')
                lines.append("</div>")
                # Mini timeline (plan vs actual)
                tl_html = _render_timeline(t["plan_min"], t["plan_max"], t["actual_load"], t.get("is_late", False))
                if tl_html:
                    lines.append(tl_html)
                lines.append("</div>")
            lines.append("</div>")

        lines.append("</div>")  # dpanel

    lines.append("</div>")  # sec

    # ── Unshipped List Section ──
    if unshipped is not None and not unshipped.empty:
        unshipped_records = unshipped.to_dict(orient="records")
        unshipped_statuses = sorted(unshipped["Status"].unique())

        # Pre-render today's rows so the section reads without JS (Outlook
        # preview). When JS runs it re-renders the same rows and takes over
        # day switching/filtering. Hidden up front only when today is clean.
        today_str = format_date(date2)
        today_us = [r for r in unshipped_records if r["TargetDate"] == today_str]
        sec_style = "" if today_us else ' style="display:none"'
        lines.append(f'<div class="sec" id="us-sec"{sec_style}>')
        lines.append(f'<div class="sec-hd">Unshipped Orders<span class="sec-sub" id="us-day">{today_str}</span></div>')
        lines.append('<div class="us-controls">')
        lines.append(
            '<input class="us-input" type="text" id="us-q" '
            'placeholder="Filter by order, product, status…" oninput="filterUnshipped()">'
        )
        lines.append('<select class="us-select" id="us-status" onchange="filterUnshipped()">')
        lines.append('<option value="">All Statuses</option>')
        for st in unshipped_statuses:
            lines.append(f'<option value="{html_escape(st)}">{html_escape(st)}</option>')
        lines.append("</select>")
        lines.append(
            '<button class="us-btn" onclick="exportUnshippedCSV()" '
            'title="Download the currently filtered rows as CSV">Export CSV</button>'
        )
        lines.append("</div>")

        n_us = len(today_us)
        lines.append(f'<div class="us-count" id="us-count">{n_us} order{"s" if n_us != 1 else ""} shown</div>')

        # Table
        lines.append('<div class="m-tbl-wrap" style="max-height:500px">')
        lines.append('<table class="m-tbl" id="us-tbl">')
        us_col_keys = ["TargetDate", "CUSTOMER ORDER NO.", "Plan_Ship", "PRODUCT NO.", "QUANTITY", "Status", "Module Count"]
        lines.append("<thead><tr>")
        for ci, h in enumerate(UNSHIPPED_DISPLAY_HEADERS):
            lines.append(f'<th onclick="sortUnshipped(\'{us_col_keys[ci]}\')">{render_header(h)}<span id="us-sort-{ci}"></span></th>')
        lines.append("</tr></thead>")
        lines.append(f'<tbody id="us-tbody">{_render_unshipped_rows(today_us, now)}</tbody>')
        lines.append("</table></div>")
        lines.append("</div>")

        # Embed data
        lines.append(f"<script>var US={json.dumps(unshipped_records)};</script>")

    # ── Modal template ──
    lines.append(
        '<div class="modal-bg" id="modal-bg" onclick="if(event.target===this)closeModal()">'
        '<div class="modal">'
        '<div class="modal-hd"><div>'
        '<div class="m-id" id="m-id"></div>'
        '<div class="m-bol" id="m-bol"></div>'
        '<div class="m-badges" id="m-badges"></div>'
        '</div><button class="modal-x" onclick="closeModal()">&times;</button></div>'
        '<div class="modal-body">'
        '<div class="m-summary" id="m-summary"></div>'
        '<div class="m-row"><div>'
        '<div class="m-section">Stage Breakdown</div>'
        '<div class="m-sg" id="m-stages"></div>'
        '</div><div>'
        '<div class="m-section">Progress</div>'
        '<div class="m-bar" id="m-bar"></div>'
        '<div class="m-section">Shipping Info</div>'
        '<div class="m-meta" id="m-meta"></div>'
        '</div></div>'
        '<div class="m-section">Module Detail</div>'
        '<input class="m-filter" type="text" id="m-filter" placeholder="Filter modules\u2026" oninput="filterModalTable(this.value)">'
        '<div class="m-tbl-wrap"><table class="m-tbl"><thead id="m-thead"></thead>'
        '<tbody id="m-tbody"></tbody></table></div>'
        '</div></div></div>'
    )

    # ── JavaScript ──
    js = r"""
/* ── Constants ── */
var SN=['Allocated','Picked','Load.Set','Load.Entry','Shipped'];
var SC=['#f97316','#f59e0b','#eab308','#84cc16','#22c55e'];
var SCls=['seg-alloc','seg-pick','seg-lset','seg-lent','seg-ship'];
var SM={};SN.forEach(function(s,i){SM[s]=SC[i]});

/* ── Small helpers ── */
function esc(s){return String(s==null?'':s).replace(/[&<>"']/g,function(c){return{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]})}
/* Unsigned duration like '1d 2h 08m' / '40m' for the modal Ship Delay row.
   Late is date-based now, so this is usually a day or more. */
function fmtDur(mins){var m=Math.abs(mins),d=Math.floor(m/1440),h=Math.floor(m%1440/60),mm=m%60,s='';if(d)s+=d+'d ';if(d||h)s+=h+'h ';s+=(s&&mm<10?'0':'')+mm+'m';return s}

/* ── Animated Counters (Feature 7) ── */
function animateValue(el,start,end,dur,fmt){
  fmt=fmt||function(v){return v};
  if(start===end){el.textContent=fmt(end);return}
  var range=end-start,startTime=null;
  function step(ts){
    if(!startTime)startTime=ts;
    var p=Math.min((ts-startTime)/dur,1);
    var eased=p<0.5?2*p*p:1-Math.pow(-2*p+2,2)/2;
    el.textContent=fmt(Math.round(start+range*eased));
    if(p<1)requestAnimationFrame(step);
  }
  requestAnimationFrame(step);
}

/* ── Tab Switching ── */
var curTab=-1,curCard=-1,curPage=0;
document.querySelectorAll('.tab').forEach(function(t,i){if(t.classList.contains('act'))curTab=i});

function showDay(n){
  document.querySelectorAll('.dpanel').forEach(function(p){p.classList.remove('act')});
  document.querySelectorAll('.tab').forEach(function(t){t.classList.remove('act')});
  document.getElementById('panel-'+n).classList.add('act');
  document.getElementById('tab-'+n).classList.add('act');
  curTab=n;curCard=-1;
  document.querySelectorAll('.tc-focus').forEach(function(c){c.classList.remove('tc-focus')});
  var s=DS[n];
  document.getElementById('hdr-date').textContent=s.dow+', '+s.date;
  var ep=document.getElementById('hdr-plan');
  animateValue(ep,parseInt(ep.textContent)||0,s.plan,500);
  var es=document.getElementById('hdr-ship');
  animateValue(es,parseInt(es.textContent)||0,s.ship,500);
  var r=document.getElementById('hdr-ratio');
  var oldPct=Math.round((parseFloat(r.textContent)||0)*10);
  var newPct=Math.round(s.pct*10);
  animateValue(r,oldPct,newPct,500,function(v){return (v/10).toFixed(1)+'%'});
  r.style.transition='color 0.5s';
  r.style.color=s.pct>=80?'#22c55e':s.pct>=50?'#eab308':'#ef4444';
  var tl=document.getElementById('hdr-trail');
  if(tl)tl.textContent=s.trail+' trailer'+(s.trail!==1?'s':'');
  var rb=document.getElementById('hdr-ratio-bar');
  if(rb){rb.style.width=Math.min(100,s.pct)+'%';rb.style.background=r.style.color}
  /* Keep the Daily Snapshot row highlight in sync with the selected tab */
  document.querySelectorAll('.ov-tbl tr.ov-sel').forEach(function(x){x.classList.remove('ov-sel')});
  var orow=document.getElementById('ov-row-'+n);
  if(orow)orow.classList.add('ov-sel');
  /* Update unshipped section: only show for today/yesterday (not future) */
  updateUnshippedForDay(s);
}

/* ── Copy Email Summary: reproduces the original
   generate_daily_shipping_kpi.py email body exactly — <h1> title,
   "KPI per status module/order count" (Summary2) and "KPI per Trailer/BOL"
   tables — for the selected day +/- 1 business day (the legacy 3-date
   window). Copied as rich HTML so Outlook pastes real tables, with a
   tab-separated plain-text alternative. ── */
function buildEmailSummary(n){
  var idxs=[n-1,n,n+1].filter(function(i){return i>=0&&i<S2.length});
  function th(h){return '<th>'+esc(h).replace(/\n/g,'<br>')+'</th>'}
  function td(c){return '<td style="font-weight:normal;">'+esc(c)+'</td>'}
  var H=[];
  H.push('<h1>SITE1 Shipping KPI - '+DS[n].date+'</h1>');
  H.push('<h2>KPI per status module/order count</h2>');
  H.push('<table border="1" cellspacing="0" cellpadding="4">','<thead>');
  H.push('<tr>'+S2H.map(th).join('')+'</tr>','</thead>','<tbody>');
  idxs.forEach(function(i){H.push('<tr>'+S2[i].map(td).join('')+'</tr>')});
  H.push('</tbody>','</table>','<br>');
  H.push('<h2>KPI per Trailer/BOL</h2>');
  idxs.forEach(function(i){
    H.push('<h3>'+DS[i].date+'</h3>');
    var rows=S1[i]||[];
    if(rows.length){
      H.push('<table border="1" cellspacing="0" cellpadding="4">','<thead>');
      H.push('<tr>'+S1H.map(th).join('')+'</tr>','</thead>','<tbody>');
      rows.forEach(function(r){H.push('<tr>'+r.map(td).join('')+'</tr>')});
      H.push('</tbody>','</table>','<br>');
    }else{
      H.push('<p>No Containers Scheduled</p>','<br>');
    }
  });
  var T=[];
  T.push('SITE1 Shipping KPI - '+DS[n].date,'');
  T.push('KPI per status module/order count');
  T.push(S2H.map(function(h){return h.replace(/\n/g,' ')}).join('\t'));
  idxs.forEach(function(i){T.push(S2[i].join('\t'))});
  T.push('','KPI per Trailer/BOL');
  idxs.forEach(function(i){
    T.push('',DS[i].date);
    var rows=S1[i]||[];
    if(rows.length){
      T.push(S1H.join('\t'));
      rows.forEach(function(r){T.push(r.join('\t'))});
    }else{T.push('No Containers Scheduled')}
  });
  return {html:H.join('\n'),text:T.join('\n')};
}

/* Open the freshly built summary in a new tab so it can be checked
   before pasting into the email. Wrapper styling only — the copied
   clipboard content stays the untouched legacy markup. */
function openSummaryTab(out,n){
  try{
    /* Closing tags are split with \/ so this script's own source never
       contains a literal closing-tag sequence (it would confuse HTML
       post-processing such as the screenshot probe injection). */
    var doc='<!doctype html><html><head><meta charset="utf-8">'
      +'<title>Email Summary - '+DS[n].date+'<\/title>'
      +'<style>body{font-family:Calibri,"Segoe UI",Arial,sans-serif;margin:24px;color:#1e293b}'
      +'table{border-collapse:collapse}th,td{padding:3px 8px}<\/style>'
      +'<\/head><body>'+out.html+'<\/body><\/html>';
    var url=URL.createObjectURL(new Blob([doc],{type:'text/html'}));
    window.open(url,'_blank');
    setTimeout(function(){URL.revokeObjectURL(url)},60000);
  }catch(e){}
}

function copySummary(btn){
  var n=curTab>=0?curTab:TODAY_IDX;
  var out=buildEmailSummary(n);
  openSummaryTab(out,n);
  function done(){
    var t=btn.textContent;
    btn.textContent='Copied ✓';
    setTimeout(function(){btn.textContent=t},1600);
  }
  /* Selection-based copy keeps the rich formatting in older browsers */
  function fallback(){
    var div=document.createElement('div');
    div.style.position='fixed';div.style.left='-9999px';
    div.innerHTML=out.html;
    document.body.appendChild(div);
    var range=document.createRange();range.selectNodeContents(div);
    var sel=window.getSelection();sel.removeAllRanges();sel.addRange(range);
    try{if(document.execCommand('copy'))done()}catch(e){}
    sel.removeAllRanges();document.body.removeChild(div);
  }
  if(navigator.clipboard&&window.ClipboardItem){
    navigator.clipboard.write([new ClipboardItem({
      'text/html':new Blob([out.html],{type:'text/html'}),
      'text/plain':new Blob([out.text],{type:'text/plain'})
    })]).then(done,fallback);
  }else{fallback()}
}

/* ── Modal: sortable table (Feature 5) ── */
var _modalDet=[],_modalSort={col:'',asc:true};
var _mCols=['mod','parts','order','qty','stage','pick','lset','lent','ship','weight'];
var _mLabels=['Module','Parts No','Order','Qty','Stage','Picked','Load Set','Load Entry','Shipped','Weight'];

function renderModalBody(){
  var tb='';_modalDet.forEach(function(r){
    var col=SM[r.stage]||'#94a3b8';
    var w=r.weight?parseFloat(r.weight).toLocaleString():'';
    tb+='<tr><td>'+esc(r.mod)+'</td><td>'+esc(r.parts)+'</td><td>'+esc(r.order)+'</td><td>'+esc(r.qty)+'</td>'
    +'<td><span class="m-dot" style="background:'+col+'"></span>'+esc(r.stage)+'</td>'
    +'<td>'+esc(r.pick||'\u2014')+'</td><td>'+esc(r.lset||'\u2014')+'</td>'
    +'<td>'+esc(r.lent||'\u2014')+'</td><td>'+esc(r.ship||'\u2014')+'</td>'
    +'<td>'+w+'</td></tr>';
  });
  document.getElementById('m-tbody').innerHTML=tb;
}

function buildModalThead(){
  var row='<tr>';
  for(var i=0;i<_mLabels.length;i++){
    var arrow='';
    if(_modalSort.col===_mCols[i])arrow='<span class="sort-arrow">'+(_modalSort.asc?'\u25B2':'\u25BC')+'</span>';
    row+='<th onclick="sortModal(\''+_mCols[i]+'\')">'+_mLabels[i]+arrow+'</th>';
  }
  row+='</tr>';
  document.getElementById('m-thead').innerHTML=row;
}

function sortModal(col){
  if(_modalSort.col===col){_modalSort.asc=!_modalSort.asc}else{_modalSort.col=col;_modalSort.asc=true}
  _modalDet.sort(function(a,b){
    var va=a[col]||'',vb=b[col]||'';
    var na=parseFloat(va),nb=parseFloat(vb);
    if(!isNaN(na)&&!isNaN(nb))return _modalSort.asc?na-nb:nb-na;
    return _modalSort.asc?String(va).localeCompare(String(vb)):String(vb).localeCompare(String(va));
  });
  buildModalThead();
  renderModalBody();
}

/* ── Open / Close Modal ── */
function openModal(el){
  var tip=document.getElementById('tc-tip');if(tip)tip.classList.remove('show');
  var d=el.dataset,st=JSON.parse(d.stages),tot=parseInt(d.modules);
  _modalDet=JSON.parse(d.detail||'[]');
  _modalSort={col:'',asc:true};
  document.getElementById('m-id').textContent=d.trailer;
  document.getElementById('m-bol').textContent=d.bol?'BOL '+d.bol:'';
  document.getElementById('m-badges').innerHTML=
    '<span class="badge '+d.statusCls+'">'+d.status+'</span>'
    +'<span class="badge b-mod">'+tot+' modules</span>'
    +'<span class="badge b-mod">'+d.skids+' skids</span>';
  var sm='<div class="m-sum-card"><div class="m-sv">'+tot+'</div><div class="m-sl">Modules</div></div>'
    +'<div class="m-sum-card"><div class="m-sv">'+d.skids+'</div><div class="m-sl">Skids</div></div>'
    +'<div class="m-sum-card"><div class="m-sv">'+(parseInt(d.qty)||0).toLocaleString()+'</div><div class="m-sl">Quantity</div></div>'
    +'<div class="m-sum-card"><div class="m-sv">'+parseFloat(d.weight).toLocaleString()+'</div><div class="m-sl">Weight</div></div>';
  var sp=tot>0?((st['Shipped']||0)/tot*100):0;
  sm+='<div class="m-sum-card"><div class="m-sv" style="color:'+(sp>=80?'#22c55e':sp>=50?'#eab308':'#ef4444')+'">'+sp.toFixed(0)+'%</div><div class="m-sl">Shipped</div></div>';
  document.getElementById('m-summary').innerHTML=sm;
  var h='';for(var i=0;i<SN.length;i++){
    var c=st[SN[i]]||0,p=tot>0?(c/tot*100):0;
    h+='<div class="m-sr"><div class="m-sn">'+SN[i]+'</div><div class="m-sc">'+c+'</div><div class="m-sp">'+p.toFixed(1)+'%</div><div class="m-sb"><div class="m-sf" style="width:'+p+'%;background:'+SC[i]+'"></div></div></div>';
  }
  document.getElementById('m-stages').innerHTML=h;
  var b='';for(var i=0;i<SN.length;i++){
    var c=st[SN[i]]||0;if(c>0&&tot>0){var p=c/tot*100,l=p>=8?c:'';
    b+='<div class="seg '+SCls[i]+'" style="width:'+p.toFixed(1)+'%" title="'+SN[i]+': '+c+'">'+l+'</div>';}}
  document.getElementById('m-bar').innerHTML=b;
  buildModalThead();
  renderModalBody();
  var mf=document.getElementById('m-filter');if(mf)mf.value='';
  var planTime;
  if(d.planMin===d.planMax){planTime=d.planMin}
  else if(d.planMin.substring(0,10)===d.planMax.substring(0,10)){planTime=d.planMin+' \u2192 '+d.planMax.substring(11)}
  else{planTime=d.planMin+' \u2192 '+d.planMax}
  /* Ship delay: late only when shipped on a later DATE than planned (matches
     the day chip and the red timeline marker); an intraday delay is on time. */
  var lateRow='';
  if(d.late!==''&&d.late!==undefined){
    var lm=parseInt(d.late);
    if(!isNaN(lm)){
      var late=d.islate==='1';
      var lbl=late?fmtDur(lm)+' late':'On time';
      var lcol=late?'#dc2626':'#16a34a';
      lateRow='<div><b>Ship Delay:</b> <span style="color:'+lcol+';font-weight:700">'+lbl+'</span></div>';
    }
  }
  document.getElementById('m-meta').innerHTML=
    '<div style="grid-column:1/-1"><b>Plan Ship Time:</b> '+planTime+'</div>'
    +'<div><b>Actual Ship Time:</b> '+d.actual+'</div>'
    +'<div><b>Order #:</b> '+d.order+'</div>'
    +'<div><b>Skid Count:</b> '+(d.skids||'\u2014')+'</div>'
    +'<div><b>Total Weight:</b> '+parseFloat(d.weight).toLocaleString()+' lbs</div>'
    +'<div><b>Total Quantity:</b> '+(parseInt(d.qty)||0).toLocaleString()+'</div>'
    +lateRow;
  document.getElementById('modal-bg').classList.add('show');
}
function closeModal(){document.getElementById('modal-bg').classList.remove('show')}

/* ── Modal Table Filter (Feature 1) ── */
function filterModalTable(q){
  q=q.toLowerCase();
  var rows=document.getElementById('m-tbody').querySelectorAll('tr');
  rows.forEach(function(r){
    r.style.display=(!q||r.textContent.toLowerCase().indexOf(q)>=0)?'':'none';
  });
}


/* ── Tooltip on Hover (Feature 15) ── */
(function(){
  var tip=document.getElementById('tc-tip');
  if(!tip)return;
  document.addEventListener('mouseover',function(e){
    var card=e.target.closest('.tc');
    if(!card){tip.classList.remove('show');return}
    var d=card.dataset;
    var st=JSON.parse(d.stages);
    var tot=parseInt(d.modules);
    var sp=tot>0?((st['Shipped']||0)/tot*100):0;
    tip.innerHTML='<b>'+d.trailer+'</b><br>Status: '+d.status+'<br>Modules: '+tot+'<br>Ship: '+sp.toFixed(0)+'%<br>Plan: '+d.planMin;
    tip.classList.add('show');
  });
  document.addEventListener('mousemove',function(e){
    if(!tip.classList.contains('show'))return;
    tip.style.left=Math.min(e.clientX+12,window.innerWidth-290)+'px';
    tip.style.top=(e.clientY+12)+'px';
  });
  document.addEventListener('mouseout',function(e){
    var from=e.target.closest&&e.target.closest('.tc');
    var to=e.relatedTarget&&e.relatedTarget.closest&&e.relatedTarget.closest('.tc');
    if(from&&!to)tip.classList.remove('show');
  });
})();

/* ── Page Navigation ── */
function showPage(p){
  if(p<0||p>=TOTAL_PAGES)return;
  curPage=p;
  var start=p*PAGE_SIZE;
  document.querySelectorAll('.tab').forEach(function(t,i){
    t.style.display=(i>=start&&i<start+PAGE_SIZE)?'':'none';
  });
  var prev=document.getElementById('nav-prev');
  var next=document.getElementById('nav-next');
  if(prev)prev.disabled=(p===0);
  if(next)next.disabled=(p===TOTAL_PAGES-1);
  if(curTab<start||curTab>=start+PAGE_SIZE){
    var preferred=(TODAY_IDX>=start&&TODAY_IDX<start+PAGE_SIZE)?TODAY_IDX:start;
    showDay(preferred);
  }
}
function changePage(dir){showPage(curPage+dir)}

/* ── Step one day forward/back, crossing page boundaries seamlessly.
   Shared by the arrow keys and the phone swipe gesture. ── */
function stepDay(dir){
  var pgStart=curPage*PAGE_SIZE;
  var pgEnd=Math.min(pgStart+PAGE_SIZE-1,DS.length-1);
  var next=curTab+dir;
  if(next>=pgStart&&next<=pgEnd){showDay(next);return}
  var np=curPage+dir;
  if(np>=0&&np<TOTAL_PAGES){
    showPage(np);
    var ns=np*PAGE_SIZE,ne=Math.min(ns+PAGE_SIZE-1,DS.length-1);
    showDay(dir>0?ns:ne);
  }
}

/* ── Daily Snapshot: jump from an overview row to that day's tab ── */
function jumpToDay(i){
  if(typeof DS==='undefined'||i<0||i>=DS.length)return;
  if(TOTAL_PAGES>1)showPage(Math.floor(i/PAGE_SIZE));
  showDay(i);
  var sec=document.getElementById('ship-sec');
  if(sec)sec.scrollIntoView({behavior:'smooth',block:'start'});
  /* Flash the landing panel so the eye knows where it arrived */
  var p=document.getElementById('panel-'+i);
  if(p){p.classList.remove('flash');void p.offsetWidth;p.classList.add('flash')}
}

/* ── Keyboard Navigation (Feature 10) ── */
document.addEventListener('keydown',function(e){
  if(e.target.tagName==='INPUT'||e.target.tagName==='SELECT')return;
  if(e.key==='Escape'){closeModal();return}
  if(e.key==='t'||e.key==='T'){jumpToDay(TODAY_IDX);return}
  if(e.key==='ArrowLeft'||e.key==='ArrowRight'){
    e.preventDefault();
    stepDay(e.key==='ArrowRight'?1:-1);
    return;
  }
  var panel=document.getElementById('panel-'+curTab);
  if(!panel)return;
  var cards=Array.from(panel.querySelectorAll('.tc')).filter(function(c){return c.style.display!=='none'});
  if(!cards.length)return;
  if(e.key==='ArrowDown'||e.key==='ArrowUp'){
    e.preventDefault();
    document.querySelectorAll('.tc-focus').forEach(function(c){c.classList.remove('tc-focus')});
    var dir=e.key==='ArrowDown'?1:-1;
    curCard=Math.max(0,Math.min(cards.length-1,curCard+dir));
    cards[curCard].classList.add('tc-focus');
    cards[curCard].scrollIntoView({block:'nearest',behavior:'smooth'});
    return;
  }
  if(e.key==='Enter'&&curCard>=0&&curCard<cards.length){
    e.preventDefault();
    openModal(cards[curCard]);
  }
});

/* ── Touch Swipe: change day on phones ──
   Swipe left = next day, right = previous day. Ignored on
   wide screens, while a modal is open, or when the gesture
   starts inside a horizontally scrollable region (tables,
   tab bar) so it never fights with normal scrolling. */
(function(){
  var x0=null,y0=null;
  function inScroller(el){
    while(el&&el.nodeType===1&&el!==document.body){
      if(el.classList&&(el.classList.contains('m-tbl-wrap')
        ||el.classList.contains('modal')
        ||el.classList.contains('tabs')))return true;
      el=el.parentElement;
    }
    return false;
  }
  document.addEventListener('touchstart',function(e){
    x0=null;
    if(window.innerWidth>600)return;
    var mb=document.getElementById('modal-bg');
    if(mb&&mb.classList.contains('show'))return;
    if(inScroller(e.target))return;
    var t=e.touches[0];x0=t.clientX;y0=t.clientY;
  },{passive:true});
  document.addEventListener('touchend',function(e){
    if(x0===null)return;
    var t=e.changedTouches[0],dx=t.clientX-x0,dy=t.clientY-y0;
    x0=null;
    if(Math.abs(dx)<60||Math.abs(dy)>45)return; /* mostly-horizontal flick only */
    stepDay(dx<0?1:-1); /* swipe left advances to the next day */
  },{passive:true});
})();

/* ── Unshipped List (Feature 4) ── */
var _usDate='',_usSort={col:'',asc:true};
var _usColKeys=["TargetDate","CUSTOMER ORDER NO.","Plan_Ship","PRODUCT NO.","QUANTITY","Status","Module Count"];
var _usDebounce=null;

function updateUnshippedForDay(s){
  var sec=document.getElementById('us-sec');
  if(!sec||typeof US==='undefined')return;
  /* Hide for future days; show for today/yesterday only */
  if(s.future){sec.style.display='none';return}
  /* Check if any unshipped orders exist for this date */
  var hasOrders=US.some(function(r){return r.TargetDate===s.date});
  if(!hasOrders){sec.style.display='none';return}
  _usDate=s.date;
  sec.style.display='';
  var dl=document.getElementById('us-day');if(dl)dl.textContent=s.date;
  /* Reset filters when switching days */
  var ss=document.getElementById('us-status');if(ss)ss.value='';
  var qi=document.getElementById('us-q');if(qi)qi.value='';
  renderUnshipped();
}

/* Status -> stage color for the dot in the Status column */
var USC={'1.No process':'#94a3b8','2.Allocated':'#f97316','3.Picked':'#f59e0b','4.Load.Set':'#eab308','5.Load.Ent':'#84cc16'};

/* Currently visible rows (selected day + text/status filters) — shared by
   the table renderer and the CSV export so both always agree. */
function unshippedRows(){
  if(typeof US==='undefined')return[];
  var q=((document.getElementById('us-q')||{}).value||'').toLowerCase();
  var st=(document.getElementById('us-status')||{}).value||'';
  return US.filter(function(r){
    if(r.TargetDate!==_usDate)return false;
    if(st&&r.Status!==st)return false;
    var row=r.TargetDate+' '+r['CUSTOMER ORDER NO.']+' '+r['PRODUCT NO.']+' '+r.Status;
    return !q||row.toLowerCase().indexOf(q)>=0;
  });
}

/* Parse 'MM/DD/YYYY HH:MM' into a ms timestamp (null when unparseable) */
function planTs(s){
  var m=/^(\d{2})\/(\d{2})\/(\d{4})(?: (\d{2}):(\d{2}))?/.exec(s||'');
  if(!m)return null;
  return new Date(+m[3],+m[1]-1,+m[2],+(m[4]||0),+(m[5]||0)).getTime();
}

function renderUnshipped(){
  if(typeof US==='undefined')return;
  var rows=unshippedRows(),tb='';
  rows.forEach(function(r){
    var sc=USC[r.Status]||'#94a3b8';
    /* Overdue = planned ship time already passed when the report was made */
    var pt=planTs(r.Plan_Ship);
    var od=pt!==null&&typeof GEN_TS!=='undefined'&&pt<GEN_TS;
    tb+='<tr><td>'+esc(r.TargetDate)+'</td><td>'+esc(r['CUSTOMER ORDER NO.'])+'</td>'
      +'<td'+(od?' class="us-overdue" title="Past planned ship time"':'')+'>'
      +esc(r.Plan_Ship)+(od?' ⚠':'')+'</td><td>'+esc(r['PRODUCT NO.'])+'</td>'
      +'<td>'+esc(r.QUANTITY)+'</td>'
      +'<td><span class="m-dot" style="background:'+sc+'"></span>'+esc(r.Status)+'</td>'
      +'<td>'+esc(r['Module Count'])+'</td></tr>';
  });
  var tbody=document.getElementById('us-tbody');
  if(tbody)tbody.innerHTML=tb||'<tr><td colspan="7" style="text-align:center;color:#94a3b8;padding:24px">No unshipped orders found</td></tr>';
  var cEl=document.getElementById('us-count');
  if(cEl)cEl.textContent=rows.length+' order'+(rows.length!==1?'s':'')+' shown';
}

/* ── Export the filtered unshipped rows as a CSV download ── */
function exportUnshippedCSV(){
  var rows=unshippedRows();
  if(!rows.length)return;
  function q(v){return '"'+String(v==null?'':v).replace(/"/g,'""')+'"'}
  var heads=["Target Date","Customer Order No.","Plan Ship","Product No.","Qty","Status","Modules"];
  var csv=[heads.map(q).join(',')];
  rows.forEach(function(r){
    csv.push(_usColKeys.map(function(k){return q(r[k])}).join(','));
  });
  var blob=new Blob([csv.join('\r\n')],{type:'text/csv;charset=utf-8'});
  var a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download='Unshipped_'+_usDate.replace(/\//g,'')+'.csv';
  document.body.appendChild(a);a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(a.href);
}

function filterUnshipped(){
  clearTimeout(_usDebounce);
  _usDebounce=setTimeout(renderUnshipped,150);
}

function sortUnshipped(col){
  if(typeof US==='undefined')return;
  if(_usSort.col===col){_usSort.asc=!_usSort.asc}else{_usSort.col=col;_usSort.asc=true}
  US.sort(function(a,b){
    var va=a[col]||'',vb=b[col]||'';
    var na=parseFloat(va),nb=parseFloat(vb);
    if(!isNaN(na)&&!isNaN(nb))return _usSort.asc?na-nb:nb-na;
    return _usSort.asc?String(va).localeCompare(String(vb)):String(vb).localeCompare(String(va));
  });
  _usColKeys.forEach(function(k,i){
    var el=document.getElementById('us-sort-'+i);
    if(el)el.innerHTML=(k===col?'<span class="sort-arrow">'+(_usSort.asc?'\u25B2':'\u25BC')+'</span>':'');
  });
  renderUnshipped();
}

/* Initialize page view and unshipped on load */
(function(){
  if(TOTAL_PAGES>1)showPage(DEFAULT_PAGE);
  if(typeof US!=='undefined'&&typeof DS!=='undefined'){
    var activeIdx=curTab>=0?curTab:TODAY_IDX;
    updateUnshippedForDay(DS[activeIdx]);
  }
})();
"""
    lines.append(f"<script>{js}</script>")

    lines.append('<div class="foot">')
    lines.append(
        '<div class="kbd-hints">'
        '<span><span class="kbd">&larr;</span><span class="kbd">&rarr;</span> switch day</span>'
        '<span><span class="kbd">&uarr;</span><span class="kbd">&darr;</span> trailer focus</span>'
        '<span><span class="kbd">T</span> today</span>'
        '<span><span class="kbd">Enter</span> open</span>'
        '<span><span class="kbd">Esc</span> close</span></div>'
    )
    ts_bits = []
    if data_asof:
        ts_bits.append(f"Data exported: {data_asof}")
    ts_bits.append(f"Report generated: {now_str}")
    stale_html = (
        '<span class="stale">&#9888; inputs older than 12h</span>' if data_stale else ""
    )
    lines.append(f'<div class="ts">{" &middot; ".join(ts_bits)}{stale_html}</div>')
    lines.append("</div>")
    lines.extend(["</div>", "</body>", "</html>"])
    return "\n".join(lines)


def _find_browser() -> str | None:
    """Locate a headless-capable Chromium browser. Edge ships with Windows,
    Chrome is the fallback. Returns the exe path or None."""
    program_dirs = [
        os.environ.get("ProgramFiles", r"C:\Program Files"),
        os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"),
    ]
    candidates = []
    for base in program_dirs:
        candidates.append(Path(base) / "Microsoft" / "Edge" / "Application" / "msedge.exe")
    for base in program_dirs:
        candidates.append(Path(base) / "Google" / "Chrome" / "Application" / "chrome.exe")
    local = os.environ.get("LOCALAPPDATA")
    if local:
        candidates.append(Path(local) / "Google" / "Chrome" / "Application" / "chrome.exe")
    for c in candidates:
        if c.is_file():
            return str(c)
    return None


def write_report_screenshot(
    html_path: Path, output_path: Path, width: int = 1400, scale: int = 2
) -> bool:
    """Screenshot the finished HTML report into the email-ready PNG with
    headless Edge/Chrome: pass 1 loads the page once to measure its full
    height, pass 2 captures it at ``scale``x device scale (2x = crisp in
    email bodies). Fail-soft like the matplotlib snapshot — any problem
    prints a warning and returns False so the caller can fall back."""
    browser = _find_browser()
    if browser is None:
        print(
            "  Full-report PNG skipped: no Edge/Chrome found for the headless "
            "screenshot (falling back to the matplotlib snapshot)."
        )
        return False
    tmp_dir = Path(tempfile.mkdtemp(prefix="kpi_shot_"))
    try:
        # Throwaway copy of the report with a probe that records the page
        # height as an attribute pass 1 reads back out of the dumped DOM.
        probe = (
            "<script>window.addEventListener('load',function(){"
            "document.documentElement.setAttribute('data-shot-h',"
            "document.documentElement.scrollHeight)})</script>"
        )
        page = tmp_dir / "report.html"
        # Inject before the LAST </body> — the document's real closing tag
        # (earlier occurrences could sit inside inline script strings).
        head, sep, tail = html_path.read_text(encoding="utf-8").rpartition("</body>")
        page.write_text(head + probe + sep + tail, encoding="utf-8")
        url = page.resolve().as_uri()
        base_cmd = [
            browser, "--headless", "--disable-gpu", "--hide-scrollbars",
            "--no-first-run", "--no-default-browser-check",
            # Throwaway profile so the headless run never trips over the
            # operator's open Edge windows holding the default profile.
            f"--user-data-dir={tmp_dir / 'profile'}",
            # Fast-forward animations/timers so counters land on final values,
            # and hard-cap the wait in case the load never settles.
            "--virtual-time-budget=4000",
            "--timeout=30000",
        ]
        measured = subprocess.run(
            base_cmd + [f"--window-size={width},900", "--dump-dom", url],
            capture_output=True, text=True, timeout=120,
            encoding="utf-8", errors="replace",
        )
        m = re.search(r'data-shot-h="(\d+)"', measured.stdout or "")
        height = min(max(int(m.group(1)) + 20 if m else 2400, 720), 16000)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        subprocess.run(
            base_cmd + [
                f"--window-size={width},{height}",
                f"--force-device-scale-factor={scale}",
                f"--screenshot={output_path}", url,
            ],
            capture_output=True, text=True, timeout=120,
            encoding="utf-8", errors="replace",
        )
        if output_path.is_file() and output_path.stat().st_size > 10_000:
            return True
        print("  Full-report PNG skipped: the headless browser produced no image.")
        return False
    except Exception as exc:  # noqa: BLE001 - keep report generation robust
        print(f"  Full-report PNG skipped: {exc}.")
        return False
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def write_snapshot_png(
    date2: date,
    summary2_rows: list[list],
    summary1_blocks: list[tuple[date, list[dict]]],
    output_path: Path,
) -> bool:
    """Fallback PNG when the full-report screenshot can't run (no
    Edge/Chrome): today's headline KPIs + stage bar on top, the
    9-business-day table below. Fail-soft — if matplotlib is unavailable (or
    the draw fails), print a warning and return False so HTML/Excel still
    succeed."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.patches import Rectangle
    except ImportError:
        print(
            "  PNG snapshot skipped: matplotlib is not installed. "
            "Run 'pip install matplotlib' to enable it (or pass --no-png to silence)."
        )
        return False

    try:
        # ── Pull today's numbers (no recomputation) ──
        today_row = next(
            (r for r, (d, _) in zip(summary2_rows, summary1_blocks) if d == date2),
            summary2_rows[0] if summary2_rows else None,
        )
        today_trailers = next((tr for d, tr in summary1_blocks if d == date2), [])
        t_plan = _cell_int(today_row[9]) if today_row else 0
        t_ship = _cell_int(today_row[6]) if today_row else 0
        t_pct = (t_ship / t_plan * 100) if t_plan else 0
        stage_vals = [_cell_int(today_row[i]) for i in (2, 3, 4, 5, 6)] if today_row else [0] * 5
        no_proc = _cell_int(today_row[1]) if today_row else 0
        shortage = _cell_int(today_row[7]) if today_row else 0
        canceled = _cell_int(today_row[8]) if today_row else 0
        stage_total = sum(stage_vals)

        fig = plt.figure(figsize=(9.6, 6.6), dpi=150)
        fig.patch.set_facecolor("#f0f4f8")

        # ── Top area: header band + KPI cards + stage bar + alerts ──
        top = fig.add_axes([0, 0.50, 1, 0.50])
        top.set_xlim(0, 1)
        top.set_ylim(0, 1)
        top.axis("off")

        def _rect(ax, x, y, w, h, fc, ec="none", lw=0, z=1):
            ax.add_patch(Rectangle((x, y), w, h, facecolor=fc, edgecolor=ec, lw=lw, zorder=z))

        # Header band (full-bleed dark navy)
        _rect(top, 0, 0.80, 1, 0.20, "#0f172a")
        top.text(0.025, 0.90, "SITE1 Shipping KPI", color="#ffffff", fontsize=18,
                 fontweight="bold", va="center", ha="left")
        top.text(0.975, 0.90, date2.strftime("%A, %B %d, %Y"), color="#cbd5e1",
                 fontsize=11.5, va="center", ha="right")

        # KPI cards
        kpis = [
            (str(t_plan), "Plan Ship", "#0f172a"),
            (str(t_ship), "Actual Ship", "#0f172a"),
            (f"{t_pct:.1f}%", "On-Time Shipping", _pct_color(t_pct)),
        ]
        cw, gap, x0 = 0.293, 0.025, 0.03
        for i, (val, lbl, col) in enumerate(kpis):
            x = x0 + i * (cw + gap)
            _rect(top, x, 0.44, cw, 0.30, "#ffffff", ec="#e2e8f0", lw=1)
            top.text(x + cw / 2, 0.605, val, color=col, fontsize=21,
                     fontweight="bold", va="center", ha="center")
            top.text(x + cw / 2, 0.49, lbl.upper(), color="#64748b", fontsize=8.5,
                     va="center", ha="center")

        # Stage bar
        top.text(0.03, 0.345, "TODAY'S PROGRESS", color="#64748b", fontsize=8.5,
                 va="center", ha="left", fontweight="bold")
        bar_x, bar_w, bar_y, bar_h = 0.03, 0.94, 0.18, 0.12
        _rect(top, bar_x, bar_y, bar_w, bar_h, "#f1f5f9")
        if stage_total > 0:
            cursor = bar_x
            for j, v in enumerate(stage_vals):
                if v <= 0:
                    continue
                seg_w = v / stage_total * bar_w
                name = STAGE_NAMES[j]
                _rect(top, cursor, bar_y, seg_w, bar_h, STAGE_COLORS[name], z=2)
                if seg_w > 0.05:
                    txt_col = "#1e293b" if j in (2, 3) else "#ffffff"
                    label = f"{name} {v}" if seg_w > 0.12 else str(v)
                    top.text(cursor + seg_w / 2, bar_y + bar_h / 2, label,
                             color=txt_col, fontsize=8, fontweight="bold",
                             va="center", ha="center", zorder=3)
                cursor += seg_w

        # Alerts line
        alert_bits = []
        if shortage:
            alert_bits.append(f"Shortage {shortage}")
        if canceled:
            alert_bits.append(f"Canceled {canceled}")
        if no_proc:
            alert_bits.append(f"No Process {no_proc}")
        if alert_bits:
            has_red = bool(shortage or canceled)
            top.text(0.03, 0.05, "  •  ".join(alert_bits),
                     color="#dc2626" if has_red else "#64748b",
                     fontsize=9.5, va="center", ha="left", fontweight="bold")
        else:
            top.text(0.03, 0.05, "No shortages or cancellations", color="#16a34a",
                     fontsize=9.5, va="center", ha="left", fontweight="bold")

        # ── Bottom area: 9-business-day table ──
        tax = fig.add_axes([0.03, 0.03, 0.94, 0.42])
        tax.axis("off")
        tax.text(0, 1.0, "9 Business Days", transform=tax.transAxes,
                 fontsize=10.5, fontweight="bold", color="#0f172a",
                 va="bottom", ha="left")

        col_labels = ["Day", "Date", "Plan", "Shipped", "On-Time %", "Short", "Cncl"]
        cell_text, day_meta = [], []
        for row, (td, _trailers) in zip(summary2_rows, summary1_blocks):
            plan = _cell_int(row[9])
            ship = _cell_int(row[6])
            pct = (ship / plan * 100) if plan else 0
            short = _cell_int(row[7])
            cncl = _cell_int(row[8])
            is_future = td > date2
            # Ratio is only meaningful once the ship day has arrived.
            ratio_txt = "—" if (is_future or not plan) else f"{pct:.0f}%"
            cell_text.append([
                _day_label(td, date2, short=True), str(row[0]), str(plan),
                str(ship), ratio_txt, str(short), str(cncl),
            ])
            day_meta.append({
                "today": td == date2, "future": is_future, "plan": plan,
                "pct": pct, "short": short, "cncl": cncl,
            })

        tbl = tax.table(cellText=cell_text, colLabels=col_labels,
                        cellLoc="center", loc="upper center", bbox=[0, 0, 1, 0.92])
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9.5)

        for (r, c), cell in tbl.get_celld().items():
            cell.set_edgecolor("#e2e8f0")
            cell.set_linewidth(0.6)
            if r == 0:
                cell.set_facecolor("#1e40af")
                cell.set_height(cell.get_height() * 1.05)
                t = cell.get_text()
                t.set_color("#ffffff")
                t.set_fontweight("bold")
                continue
            meta = day_meta[r - 1]
            if meta["today"]:
                cell.set_facecolor("#dbeafe")
            elif (r - 1) % 2 == 0:
                cell.set_facecolor("#f8fafc")
            else:
                cell.set_facecolor("#ffffff")
            txt = cell.get_text()
            if meta["future"]:
                txt.set_color("#94a3b8")
            else:
                txt.set_color("#334155")
            if meta["today"]:
                txt.set_fontweight("bold")
            if c == 4 and meta["plan"] and not meta["future"]:
                cell.set_facecolor(_pct_color(meta["pct"]))
                txt.set_color("#ffffff")
                txt.set_fontweight("bold")
            if (c == 5 and meta["short"]) or (c == 6 and meta["cncl"]):
                txt.set_color("#dc2626")
                txt.set_fontweight("bold")

        fig.text(0.975, 0.008,
                 f"Generated {datetime.now().strftime('%m/%d/%Y %H:%M')}",
                 ha="right", va="bottom", fontsize=7, color="#94a3b8")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(output_path, dpi=150, facecolor=fig.get_facecolor())
        plt.close(fig)
        return True
    except Exception as exc:  # noqa: BLE001 - keep report generation robust
        print(f"  PNG snapshot skipped: failed to render ({exc}).")
        return False


def write_unshipped_excel(
    unshipped: pd.DataFrame,
    output_path: Path,
    summary2_rows: list[list] | None = None,
    summary1_blocks: list[tuple[date, list[dict]]] | None = None,
    date2: date | None = None,
) -> None:
    """Unshipped_List stays the first/active sheet (raw source-system headers,
    now styled: header fill, freeze pane, autofilter, numeric Qty/Modules).
    When the summary data is passed, a second Daily_Summary sheet mirrors the
    9-day Daily Snapshot table."""
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    today_fill = PatternFill("solid", fgColor="DBEAFE")

    def _num(value):
        try:
            f = float(str(value))
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            return value

    wb = Workbook()
    ws = wb.active
    ws.title = "Unshipped_List"
    for col_idx, header in enumerate(UNSHIPPED_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (_, row) in enumerate(unshipped.iterrows(), start=2):
        ws.cell(row=row_idx, column=1, value=row["TargetDate"])
        ws.cell(row=row_idx, column=2, value=row["CUSTOMER ORDER NO."])
        ws.cell(row=row_idx, column=3, value=row["Plan_Ship"])
        ws.cell(row=row_idx, column=4, value=row["PRODUCT NO."])
        ws.cell(row=row_idx, column=5, value=_num(row["QUANTITY"]))
        ws.cell(row=row_idx, column=6, value=row["Status"])
        ws.cell(row=row_idx, column=7, value=_num(row["Module Count"]))

    ws.freeze_panes = "A2"
    last_row = len(unshipped) + 1
    ws.auto_filter.ref = f"A1:G{last_row}"
    for col_idx, width in enumerate([12, 22, 18, 18, 10, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Daily_Summary: the 9-day Daily Snapshot table, same numbers ──
    if summary2_rows and summary1_blocks and date2 is not None:
        ws2 = wb.create_sheet("Daily_Summary")
        headers = ["Day", "Date", "Plan", "Shipped", "On-Time %", "Short", "Cncl", "No Process"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        for row_idx, (row, (td, _trailers)) in enumerate(
            zip(summary2_rows, summary1_blocks), start=2
        ):
            plan = _cell_int(row[9])
            ship = _cell_int(row[6])
            is_future = td > date2
            ws2.cell(row=row_idx, column=1, value=_day_label(td, date2, short=True))
            ws2.cell(row=row_idx, column=2, value=str(row[0]))
            ws2.cell(row=row_idx, column=3, value=plan)
            ws2.cell(row=row_idx, column=4, value=ship)
            # Ratio is only meaningful once the ship day has arrived.
            if plan and not is_future:
                pct_cell = ws2.cell(row=row_idx, column=5, value=ship / plan)
                pct_cell.number_format = "0%"
            ws2.cell(row=row_idx, column=6, value=_cell_int(row[7]))
            ws2.cell(row=row_idx, column=7, value=_cell_int(row[8]))
            ws2.cell(row=row_idx, column=8, value=_cell_int(row[1]))
            if td == date2:
                for col_idx in range(1, len(headers) + 1):
                    c = ws2.cell(row=row_idx, column=col_idx)
                    c.fill = today_fill
                    c.font = Font(bold=True)
        ws2.freeze_panes = "A2"
        for col_idx, width in enumerate([12, 12, 8, 10, 11, 8, 8, 12], start=1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    base_dir = Path(args.base_dir)
    data_dir = Path(args.data_dir) if args.data_dir else base_dir / "Data"

    problems = validate_inputs(data_dir)
    if problems:
        print("Cannot generate report - input problems found:")
        for p in problems:
            print(f"  - {p}")
        print(f'Drop fresh CSV exports (201S/201P/202/210/701) into "{data_dir}" and rerun.')
        raise SystemExit(1)

    data_asof, data_stale = data_freshness(data_dir)
    if data_stale:
        print(
            "WARNING: at least one CSV export is more than 12 hours old - "
            "the report may not reflect current operations."
        )

    target_dates, date2 = resolve_dates(args.date, args.dates)
    target_date_set = set(target_dates)

    df201s = pd.read_csv(data_dir / "201S.csv", dtype=str)
    df201p = pd.read_csv(data_dir / "201P.csv", dtype=str)
    df202 = pd.read_csv(data_dir / "202.csv", dtype=str)
    df210 = pd.read_csv(data_dir / "210.csv", dtype=str)
    df701 = pd.read_csv(data_dir / "701.CSV", dtype=str)

    # 210 ships many rows with a blank (".00") weight; backfill those from the
    # 701 product master (WEIGHT PER PCS x the row's Qty) before any weight sums.
    filled = fill_zero_weights(df210, df701)
    print(f"Filled {filled} zero-weight 210 row(s) from 701 (WEIGHT PER PCS x Qty).")

    df201s["SHIP DATE"] = clean_date_column(df201s, "SHIP DATE")
    df201p["SHIP DATE"] = clean_date_column(df201p, "SHIP DATE")

    df202["PLAN SHIP DATE"] = clean_date_column(df202, "PLAN SHIP DATE")
    df202["PICKING DATE"] = clean_date_column(df202, "PICKING DATE")
    df202["LOADING SET DATE"] = clean_date_column(df202, "LOADING SET DATE")
    df202["LOADING ENTRY DATE"] = clean_date_column(df202, "LOADING ENTRY DATE")
    df202["SHIPMENT LOAD DATE"] = clean_date_column(df202, "SHIPMENT LOAD DATE")

    df202["plan_dt"] = combine_datetime(df202["PLAN SHIP DATE"], df202["PLAN SHIP TIME"])
    df202["ship_dt"] = combine_datetime(
        df202["SHIPMENT LOAD DATE"], df202["SHIPMENT LOAD TIME"]
    )

    module_counts = df202.groupby("ORDER NO").size()

    summary2_rows = [
        compute_summary2(df201s, df201p, df202, df210, target_date)
        for target_date in target_dates
    ]

    summary1_blocks = [
        (target_date, compute_summary1(df202, df210, target_date))
        for target_date in target_dates
    ]

    unshipped = compute_unshipped_list(
        df201s, df201p, df202, module_counts, target_date_set
    )

    output_root = base_dir / "Output"
    html_dir = output_root / "HTML"
    excel_dir = output_root / "Excel"
    png_dir = output_root / "PNG"
    html_dir.mkdir(parents=True, exist_ok=True)
    excel_dir.mkdir(parents=True, exist_ok=True)
    png_dir.mkdir(parents=True, exist_ok=True)

    if args.output:
        html_output_path = Path(args.output)
        excel_name = f"{html_output_path.stem}.xlsx"
        png_name = f"{html_output_path.stem}.png"
    else:
        html_output_path = html_dir / f"Site1_Shipping_KPI_{date2.strftime('%Y%m%d')}.html"
        excel_name = f"Site1_Shipping_KPI_{date2.strftime('%Y%m%d')}.xlsx"
        png_name = f"Site1_Shipping_KPI_{date2.strftime('%Y%m%d')}.png"

    html_output_path.parent.mkdir(parents=True, exist_ok=True)
    excel_output_path = excel_dir / excel_name
    png_output_path = png_dir / png_name
    html_output = build_html(
        date2, summary1_blocks, summary2_rows, unshipped,
        data_asof=data_asof, data_stale=data_stale,
    )
    html_output_path.write_text(html_output, encoding="utf-8")
    write_unshipped_excel(
        unshipped, excel_output_path, summary2_rows, summary1_blocks, date2
    )
    print(
        "Saved report to "
        f"{html_output_path} and Unshipped_List to {excel_output_path}"
    )

    if not args.no_png:
        if write_report_screenshot(html_output_path, png_output_path):
            print(f"Saved full-report snapshot to {png_output_path}")
        elif write_snapshot_png(date2, summary2_rows, summary1_blocks, png_output_path):
            print(f"Saved email snapshot to {png_output_path} (matplotlib fallback)")

    # ── Console recap: the 9-day table at a glance (ASCII-safe) ──
    print()
    print(f"  {'Day':<11}{'Date':<13}{'Plan':>6}{'Ship':>6}{'On-Time':>9}{'Short':>7}{'Cncl':>6}")
    for row, (td, _trailers) in zip(summary2_rows, summary1_blocks):
        plan = _cell_int(row[9])
        ship = _cell_int(row[6])
        pct = f"{ship / plan * 100:.0f}%" if plan and td <= date2 else "-"
        marker = ">" if td == date2 else " "
        print(
            f"{marker} {_day_label(td, date2, short=True):<11}{str(row[0]):<13}"
            f"{plan:>6}{ship:>6}{pct:>9}{_cell_int(row[7]):>7}{_cell_int(row[8]):>6}"
        )
    if data_asof:
        print(f"\n  Data exported: {data_asof}" + ("  (STALE)" if data_stale else ""))

    if args.open:
        import webbrowser

        webbrowser.open(html_output_path.resolve().as_uri())


if __name__ == "__main__":
    main()
