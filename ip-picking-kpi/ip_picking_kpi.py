import os
import sys
import json
import shutil
import subprocess
import datetime
from datetime import timedelta

import pandas as pd
import openpyxl
from openpyxl import Workbook

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = './IP Picking'
ENGINE_SCRIPT = os.path.join(BASE_DIR, 'Work', 'IP_Picking_KPI_engine.py')
DATA_DIR = os.path.join(BASE_DIR, 'Data')
WORK_PATH = os.path.join(BASE_DIR, 'work9')
HTML_PATH = os.path.join(BASE_DIR, 'Html')
SHAREPOINT_PATH = './bi_data/IP_Picking/BI Data'

INPUT_CSV = 'Picking_MODULE.csv'

# Files to copy to SharePoint
CSV_FILES = [
    'df_KPI.csv', 'df_KPI0.csv', 'df_KPI_1.csv',
    'df_day_per_pcode_all.csv',
    'df_ope0.csv', 'df_ope1.csv', 'df_ope2.csv',
    'df_pcode0.csv', 'df_pcode1.csv', 'df_pcode2.csv',
    'df_pcode_per_shift_main.csv',
    'df_per_shift.csv', 'df_per_shift_speed.csv',
    'df_week1.csv', 'set_day_pcode_total.csv',
]
EXCEL_FILES = ['IP_Picking Group.xlsx', 'IP_Picking_total.xlsx', 'IP_ShipRecord.xlsx']


# ── Config Generation ──────────────────────────────────────────────────────────

def generate_order_groups(dates):
    """Generate OrderGroups for a list of date strings (YYYYMMDD).
    Each date gets 12 groups of 3 sequential orders."""
    groups = []
    for date_str in dates:
        for k in range(1, 13):
            group_name = f"{date_str}-{k:02d}"
            o1 = f"{date_str}{(k-1)*3+1:02d}"
            o2 = f"{date_str}{(k-1)*3+2:02d}"
            o3 = f"{date_str}{(k-1)*3+3:02d}"
            groups.append([group_name, o1, o2, o3])
    return groups


def generate_config(monday, include_saturday):
    """Generate all date-dependent configuration from a Monday date."""
    # TargeDate: Mon-Fri (always 5 days, used for charts)
    targe_date = [(monday + timedelta(days=i)).strftime('%Y%m%d') for i in range(5)]

    # TargeDate1: Mon-Fri or Mon-Sat (used for data filtering)
    num_days = 6 if include_saturday else 5
    targe_date1 = [(monday + timedelta(days=i)).strftime('%Y%m%d') for i in range(num_days)]

    # FIG_TITLE: "Order week" is 7 days before this Monday
    order_week = monday - timedelta(days=7)
    fig_title = f"IP Picking KPI, Pallet code-S3 [Order week {order_week.strftime('%m/%d/%Y')}]"

    # OrderGroups: use all dates in TargeDate1
    order_groups = generate_order_groups(targe_date1)

    return {
        'TargeDate': targe_date,
        'TargeDate1': targe_date1,
        'FIG_TITLE': fig_title,
        'OrderGroups': order_groups,
    }


# ── IP_Picking Group.xlsx Generation ──────────────────────────────────────────

def generate_picking_group_xlsx(targe_date, output_path):
    """Generate IP_Picking Group.xlsx with Mon-Fri order groups for Power BI."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pick_Group'
    ws.append(['Group name', 'Order No1', 'Order No2', 'Order No3'])

    for date_str in targe_date:  # Mon-Fri only (5 days)
        for k in range(1, 13):
            group_name = f"{date_str}-{k:02d}"
            o1 = f"{date_str}{(k-1)*3+1:02d}"
            o2 = f"{date_str}{(k-1)*3+2:02d}"
            o3 = f"{date_str}{(k-1)*3+3:02d}"
            ws.append([group_name, o1, o2, o3])

    wb.save(output_path)


# ── IP_Picking_total.xlsx Update ──────────────────────────────────────────────

def update_picking_total_xlsx(monday, work_path):
    """Append this week's shift totals to IP_Picking_total.xlsx."""
    csv_path = os.path.join(work_path, 'df_pcode_per_shift_main.csv')
    xlsx_path = os.path.join(work_path, 'IP_Picking_total.xlsx')

    df = pd.read_csv(csv_path)
    shift1_row = df[df['Shift1'] == 'shift1'].iloc[0]
    shift2_row = df[df['Shift1'] == 'shift2'].iloc[0]

    speed_1st = shift1_row['ispeed']
    speed_2nd = shift2_row['ispeed']
    count_1st = int(shift1_row['pickCount'])
    count_2nd = int(shift2_row['pickCount'])

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['IP_Picking']

    monday_dt = datetime.datetime(monday.year, monday.month, monday.day)

    # Check for existing row — overwrite if found
    target_row = None
    for row_num in range(ws.max_row, 1, -1):
        if ws.cell(row_num, 1).value == monday_dt:
            target_row = row_num
            break

    if target_row is None:
        target_row = ws.max_row + 1
        action = 'Appended'
    else:
        action = 'Overwrote'

    ws.cell(target_row, 1, monday_dt)
    ws.cell(target_row, 1).number_format = 'MM/DD/YY'
    ws.cell(target_row, 2, speed_1st)
    ws.cell(target_row, 3, speed_2nd)
    ws.cell(target_row, 4, count_1st)
    ws.cell(target_row, 5, count_2nd)

    wb.save(xlsx_path)
    print(f"  {action}: {monday.strftime('%m/%d/%Y')} | speed=[{speed_1st}, {speed_2nd}] | count=[{count_1st}, {count_2nd}]")


# ── SharePoint Copy ───────────────────────────────────────────────────────────

def copy_to_sharepoint(work_path, sharepoint_path):
    """Copy all output files to the SharePoint sync folder."""
    all_files = CSV_FILES + EXCEL_FILES
    copied = 0
    for filename in all_files:
        src = os.path.join(work_path, filename)
        dst = os.path.join(sharepoint_path, filename)
        if os.path.exists(src):
            shutil.copy2(src, dst)
            copied += 1
        else:
            print(f"  WARNING: {filename} not found in work9/")
    print(f"  Copied {copied}/{len(all_files)} files to BI Data/")


# ── Validation ────────────────────────────────────────────────────────────────

def validate_outputs(work_path):
    """Check all expected output files exist and are non-empty."""
    missing = []
    for f in CSV_FILES:
        path = os.path.join(work_path, f)
        if not os.path.exists(path):
            missing.append(f)
        elif os.path.getsize(path) == 0:
            missing.append(f"{f} (EMPTY)")
    if missing:
        print(f"  WARNING: Missing/empty outputs: {', '.join(missing)}")
        return False
    print(f"  All {len(CSV_FILES)} CSV outputs verified.")
    return True


# ── Display Helpers ───────────────────────────────────────────────────────────

def display_order_groups(order_groups):
    """Display order groups for user review."""
    print(f"\nOrder Groups ({len(order_groups)} entries):")
    for i, g in enumerate(order_groups):
        print(f"  {g[0]}: {g[1]}, {g[2]}, {g[3]}")
        # Print separator between days
        if (i + 1) % 12 == 0 and i + 1 < len(order_groups):
            print("  ---")


def ask_yes_no(prompt):
    """Ask a yes/no question. Returns True for yes."""
    while True:
        answer = input(prompt).strip().lower()
        if answer in ('y', 'yes'):
            return True
        if answer in ('n', 'no'):
            return False
        print("  Please enter y or n.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  IP Picking KPI Report Generator")
    print("=" * 60)

    # Auto-detect previous Monday
    today = datetime.date.today()
    this_monday = today - timedelta(days=today.weekday())
    prev_monday = this_monday - timedelta(days=7)

    prev_friday = prev_monday + timedelta(days=4)
    prev_saturday = prev_monday + timedelta(days=5)

    print(f"\nReport week: {prev_monday.strftime('%m/%d/%Y')} (Mon) - {prev_friday.strftime('%m/%d/%Y')} (Fri)")

    # Allow override
    override = input(f"\nUse {prev_monday.strftime('%m/%d/%Y')} as start date? [y/n]: ").strip().lower()
    if override in ('n', 'no'):
        date_str = input("Enter Monday date (YYYY-MM-DD): ").strip()
        prev_monday = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        if prev_monday.weekday() != 0:
            print(f"ERROR: {prev_monday} is a {prev_monday.strftime('%A')}, not a Monday.")
            sys.exit(1)
        prev_friday = prev_monday + timedelta(days=4)
        prev_saturday = prev_monday + timedelta(days=5)
        print(f"Report week: {prev_monday.strftime('%m/%d/%Y')} (Mon) - {prev_friday.strftime('%m/%d/%Y')} (Fri)")

    # Ask about Saturday
    include_saturday = ask_yes_no(f"\nDid pickers work on Saturday {prev_saturday.strftime('%m/%d/%Y')}? [y/n]: ")

    # Generate config
    config = generate_config(prev_monday, include_saturday)

    # Display order groups for confirmation
    display_order_groups(config['OrderGroups'])
    print(f"\nTitle: {config['FIG_TITLE']}")

    if not ask_yes_no("\nProceed with these settings? [y/n]: "):
        print("Aborted.")
        sys.exit(0)

    # Pre-flight checks
    input_csv_path = os.path.join(DATA_DIR, INPUT_CSV)
    if not os.path.exists(input_csv_path):
        print(f"\nERROR: {INPUT_CSV} not found in Data/")
        sys.exit(1)

    if not os.path.exists(ENGINE_SCRIPT):
        print(f"\nERROR: Engine script not found: {ENGINE_SCRIPT}")
        sys.exit(1)

    if not os.path.isdir(SHAREPOINT_PATH):
        print(f"\nWARNING: SharePoint folder not found: {SHAREPOINT_PATH}")
        print("  Files will NOT be copied to SharePoint.")
        skip_copy = True
    else:
        skip_copy = False

    # Step 1: Run the KPI engine
    print(f"\n[1/4] Running KPI engine...")
    config_json = json.dumps(config)
    result = subprocess.run(
        [sys.executable, ENGINE_SCRIPT, config_json],
        cwd=BASE_DIR,
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"ERROR: Engine script failed.")
        print(result.stderr)
        sys.exit(1)
    print("  Done.")

    # Validate outputs
    if not validate_outputs(WORK_PATH):
        print("WARNING: Some outputs are missing. Continuing anyway...")

    # Step 2: Generate IP_Picking Group.xlsx
    print(f"\n[2/4] Generating IP_Picking Group.xlsx...")
    group_xlsx_path = os.path.join(WORK_PATH, 'IP_Picking Group.xlsx')
    generate_picking_group_xlsx(config['TargeDate'], group_xlsx_path)
    print("  Done.")

    # Step 3: Update IP_Picking_total.xlsx
    print(f"\n[3/4] Updating IP_Picking_total.xlsx...")
    update_picking_total_xlsx(prev_monday, WORK_PATH)

    # Step 4: Copy to SharePoint
    if not skip_copy:
        print(f"\n[4/4] Copying files to SharePoint...")
        copy_to_sharepoint(WORK_PATH, SHAREPOINT_PATH)
    else:
        print(f"\n[4/4] Skipping SharePoint copy (folder not found).")

    print(f"\n{'=' * 60}")
    print(f"  Report complete for week of {prev_monday.strftime('%m/%d/%Y')}!")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
