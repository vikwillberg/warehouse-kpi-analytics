"""
DS Aged Inventory Report
Generates aged inventory count and detailed list from warehouse data.
Requested by the OEM client to track and eliminate old parts from the warehouse.
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION — adjust these as needed
# ============================================================

# Modules with vanning date on or before this date are considered "aged"
AGED_CUTOFF_DATE = "2025-03-01"

# the OEM client's target: zero aged inventory by this date (informational)
TARGET_ZERO_DATE = "2026-03-31"

# Input file paths (relative to this script's directory)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_502 = os.path.join(SCRIPT_DIR, "502.csv")
FILE_201P = os.path.join(SCRIPT_DIR, "201P.csv")
FILE_201S = os.path.join(SCRIPT_DIR, "201S.csv")

# Output file name (date-stamped)
TODAY_STR = datetime.now().strftime("%Y%m%d")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, f"DS_Aged_Inventory_{TODAY_STR}.xlsx")

# ============================================================
# ORIGIN MAPPING — prefix → (origin_name, date_table)
# Longest prefix checked first via sorted lookup
# ============================================================

# Date tables: 1 = Japan-style, 2 = China/Canada/Mexico-style, 3 = Thailand-style (ETA-2mo)
ORIGIN_MAP = {
    # 5-char prefixes
    "KJ550": ("Mexico", 2), "KJ552": ("Mexico", 2), "KJ563": ("Mexico", 2),
    "KJ579": ("Mexico", 2), "KJ598": ("Mexico", 2), "KJ617": ("Mexico", 2),
    "KJ621": ("Mexico", 2), "KJ623": ("Mexico", 2), "KJ624": ("Mexico", 2),
    "KJ626": ("Mexico", 2), "KJ646": ("Mexico", 2), "KJ656": ("Mexico", 2),
    "KJ694": ("Mexico", 2), "KJ698": ("Mexico", 2),
    "KJ699": ("MMVO", 1), "KJ900": ("JS", 1),
    "KJ911": ("3RD(THAI)", 3), "KJ912": ("3RD(CHINA)", 2),
    "KA085": ("Thailand", 3), "KA152": ("Thailand", 3), "KA158": ("Thailand", 3),
    "KA199": ("Thailand", 3), "KA224": ("Thailand", 3), "KA261": ("Thailand", 3),
    "KA296": ("Thailand", 3), "KA331": ("Thailand", 3), "KA118": ("Thailand", 3),
    "KA079": ("China", 2), "KA125": ("China", 2), "KA184": ("China", 2),
    "KA255": ("China", 2), "KA277": ("China", 2),
    "KJ540": ("Canada", 2), "KJ999": ("Japan", 1),
    # 4-char prefixes
    "KJ50": ("Canada", 2), "KJ51": ("Canada", 2), "KJ52": ("Canada", 2),
    "KJ53": ("Canada", 2), "KJ54": ("Canada", 2),
    "KJ55": ("Mexico", 2), "KJ56": ("Mexico", 2), "KJ57": ("Mexico", 2),
    "KJ58": ("Mexico", 2), "KJ59": ("Mexico", 2), "KJ60": ("Mexico", 2),
    "KJ61": ("Mexico", 2), "KJ62": ("Mexico", 2), "KJ63": ("Mexico", 2),
    "KJ64": ("Mexico", 2), "KJ65": ("Mexico", 2), "KJ66": ("Mexico", 2),
    "KJ67": ("Mexico", 2), "KJ68": ("Mexico", 2), "KJ69": ("Mexico", 2),
    # 3-char prefixes
    "KJ7": ("IS", 1), "KJ8": ("IS", 1),
    "KJ0": ("US", 2), "KJ1": ("US", 2), "KJ2": ("US", 2),
    "KJ3": ("US", 2), "KJ4": ("US", 2),
    # 1-char prefixes
    "S": ("Japan", 1), "K": ("Japan", 1),
}

# Pre-sort prefixes by length descending for longest-match-first lookup
_SORTED_PREFIXES = sorted(ORIGIN_MAP.keys(), key=len, reverse=True)

# Origin display labels (numbered for sorting)
ORIGIN_LABELS = {
    "Canada": "1.Canada",
    "Mexico": "2.Mexico",
    "MMVO": "2.Mexico",       # grouped with Mexico
    "China": "3.China",
    "3RD(CHINA)": "3.China",  # grouped with China
    "Japan": "4.Japan",
    "JS": "4.Japan",          # grouped with Japan
    "IS": "4.Japan",          # grouped with Japan
    "US": "4.Japan",          # grouped with Japan
    "Thailand": "5.Thailand",
    "3RD(THAI)": "5.Thailand",  # grouped with Thailand
}

# DateTable 1 letter → month mapping
MONTH_LETTER_MAP = {
    "M": 1, "N": 2, "O": 3, "P": 4, "Q": 5, "R": 6,
    "S": 7, "T": 8, "U": 9, "V": 10, "W": 11, "X": 12,
    "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6,
    "G": 7, "H": 8, "I": 9, "J": 10, "K": 11, "L": 12,
}


# ============================================================
# FUNCTIONS
# ============================================================

def get_origin(module_no: str) -> tuple:
    """Return (origin_name, date_table) for a module number using longest-prefix match."""
    for prefix in _SORTED_PREFIXES:
        if module_no.startswith(prefix):
            return ORIGIN_MAP[prefix]
    return ("Unknown", 0)


def get_vanning_date(module_no: str, eta: datetime, origin: str, date_table: int) -> datetime | None:
    """
    Extract or compute the vanning date for a module.
    Returns the 1st of the computed vanning month, or None if extraction fails.
    """
    if not isinstance(eta, datetime) or pd.isna(eta):
        return None

    # AAA rule takes priority
    if "AAA" in module_no.upper():
        if origin in ("Mexico", "MMVO"):
            return eta.replace(day=1)
        else:
            return (eta - relativedelta(months=2)).replace(day=1)

    # DateTable 3: Thailand-style → ETA minus 2 months
    if date_table == 3:
        return (eta - relativedelta(months=2)).replace(day=1)

    # DateTable 1: Japan-style → 6th char is a letter mapping to month
    if date_table == 1:
        if len(module_no) < 6:
            return None
        char6 = module_no[5].upper()
        month = MONTH_LETTER_MAP.get(char6)
        if month is None:
            return None
        return _resolve_year(month, eta)

    # DateTable 2: China/Canada/Mexico-style → 6th-7th chars are numeric month
    if date_table == 2:
        if len(module_no) < 7:
            return None
        try:
            month = int(module_no[5:7])
        except ValueError:
            return None
        if month < 1 or month > 12:
            return None
        return _resolve_year(month, eta)

    return None


def _resolve_year(month: int, eta: datetime) -> datetime:
    """
    Given an extracted month and the ETA, find the most recent 1st-of-month
    that is on or before the ETA.
    """
    # Try same year first
    candidate = datetime(eta.year, month, 1)
    if candidate <= eta:
        return candidate
    # Otherwise previous year
    return datetime(eta.year - 1, month, 1)


def load_502(path: str) -> pd.DataFrame:
    """Load the 502 inventory file."""
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip().str.strip('"')
    df["ETA"] = pd.to_datetime(df["ETA"], errors="coerce")
    df["QUANTITY"] = pd.to_numeric(df["QUANTITY"], errors="coerce").fillna(0).astype(int)
    return df


def load_201(path_p: str, path_s: str) -> pd.DataFrame:
    """Load 201P and 201S, keep only PRODUCT NO. and SHIP DATE, return combined."""
    frames = []
    for path in [path_p, path_s]:
        if os.path.exists(path):
            df = pd.read_csv(path, dtype=str, usecols=["PRODUCT NO.", "SHIP DATE"])
            df["SHIP DATE"] = pd.to_datetime(df["SHIP DATE"], errors="coerce")
            frames.append(df)
    if not frames:
        return pd.DataFrame(columns=["PRODUCT NO.", "SHIP DATE"])
    combined = pd.concat(frames, ignore_index=True)
    return combined


def get_last_ship_dates(df_201: pd.DataFrame) -> dict:
    """Return dict of product_no → max ship date from 201 data."""
    if df_201.empty:
        return {}
    grouped = df_201.groupby("PRODUCT NO.")["SHIP DATE"].max()
    return grouped.to_dict()


# ============================================================
# EXCEL FORMATTING
# ============================================================

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_TOTAL_FONT = Font(bold=True, size=11)
_TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=14, color="2F5496")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _format_summary(writer, summary: pd.DataFrame, cutoff: datetime):
    """Format the Summary sheet with title, styled pivot table, and info block."""
    wb = writer.book
    ws = wb["Summary"]

    n_rows = summary.shape[0]  # includes Total row
    n_cols = summary.shape[1] + 1  # +1 for the index column

    # --- Title block above the pivot (rows 1-3, pivot starts at row 4 in pandas output) ---
    # Shift pivot data down by inserting rows at top
    ws.insert_rows(1, 3)

    ws["A1"] = "DS Aged Inventory Summary"
    ws["A1"].font = _TITLE_FONT

    ws["A2"] = f"Cutoff Vanning Date: {cutoff.strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(size=11, italic=True)

    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(size=11, italic=True)

    # Pivot table starts at row 4 (header) / row 5 (data) after insert
    header_row = 4
    data_start = 5
    data_end = data_start + n_rows - 1  # last data row (Total)

    # --- Style header row ---
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # --- Style data cells ---
    for row_idx in range(data_start, data_end + 1):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            # Number format for data columns (not the index/origin column)
            if col_idx > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # --- Style Total row (last row) ---
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=data_end, column=col_idx)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL

    # --- Column widths ---
    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # --- Info block below the pivot ---
    info_row = data_end + 2
    ws.cell(row=info_row, column=1, value="Target: Zero aged inventory by")
    ws.cell(row=info_row, column=2, value=TARGET_ZERO_DATE)
    ws.cell(row=info_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=info_row, column=2).font = Font(bold=True, size=11, color="C00000")


def _format_detail(writer, result: pd.DataFrame):
    """Format the Aged Inventory detail sheet with styled headers and auto-width columns."""
    wb = writer.book
    ws = wb["Aged Inventory"]

    n_cols = result.shape[1]

    # --- Style header row ---
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # --- Auto-fit column widths (approximate) ---
    col_widths = {
        "LOCATION": 18, "Location2": 12, "MODULE#": 22, "PRODUCT": 22,
        "QUANTITY": 12, "Origin": 14, "Vanning Date": 14, "Last_ShipDate": 14,
    }
    for col_idx, col_name in enumerate(result.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 14)

    # --- Borders on data rows ---
    for row_idx in range(2, len(result) + 2):
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto-filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{len(result) + 1}"


# ============================================================
# MAIN
# ============================================================

def main():
    print(f"DS Aged Inventory Report")
    print(f"Cutoff vanning date: {AGED_CUTOFF_DATE}")
    print(f"Target zero date:    {TARGET_ZERO_DATE}")
    print("-" * 50)

    cutoff = datetime.strptime(AGED_CUTOFF_DATE, "%Y-%m-%d")

    # --- Load data ---
    print("Loading 502.csv ...")
    df = load_502(FILE_502)
    print(f"  {len(df):,} modules in inventory")

    print("Loading 201P.csv + 201S.csv ...")
    df_201 = load_201(FILE_201P, FILE_201S)
    last_ship = get_last_ship_dates(df_201)
    print(f"  {len(last_ship):,} unique products with ship dates")

    # --- Compute origin & vanning date for every module ---
    print("Computing origin and vanning dates ...")
    origins = []
    origin_labels = []
    date_tables = []
    vanning_dates = []

    for _, row in df.iterrows():
        mod = str(row["MODULE#"]).strip()
        eta = row["ETA"]

        origin, dt = get_origin(mod)
        origins.append(origin)
        date_tables.append(dt)
        origin_labels.append(ORIGIN_LABELS.get(origin, origin))

        van = get_vanning_date(mod, eta, origin, dt)
        vanning_dates.append(van)

    df["Origin_Raw"] = origins
    df["Origin"] = origin_labels
    df["DateTable"] = date_tables
    df["Vanning Date"] = vanning_dates

    # Location2: Overflow vs Original
    df["Location2"] = df["LOCATION"].apply(
        lambda x: "Overflow" if str(x).startswith("OV-FLO") else "Original"
    )

    # --- Filter aged inventory ---
    has_vanning = df["Vanning Date"].notna()
    is_aged = df["Vanning Date"] <= cutoff
    aged = df[has_vanning & is_aged].copy()
    print(f"  {len(aged):,} modules with vanning date <= {AGED_CUTOFF_DATE}")

    # --- Join Last_ShipDate ---
    # Primary match: PRODUCT → PRODUCT NO. in 201
    aged["Last_ShipDate"] = aged["PRODUCT"].map(last_ship)

    # Fallback: strip dashes and spaces from PRODUCT, try matching against
    # stripped versions of 201 product numbers
    missing_mask = aged["Last_ShipDate"].isna()
    if missing_mask.any():
        # Build fallback lookup: stripped 201 product → max ship date
        stripped_ship = {}
        for prod, dt in last_ship.items():
            key = str(prod).replace("-", "").replace(" ", "")
            if key not in stripped_ship or dt > stripped_ship[key]:
                stripped_ship[key] = dt

        fallback = aged.loc[missing_mask, "PRODUCT"].apply(
            lambda x: stripped_ship.get(str(x).replace("-", "").replace(" ", ""))
        )
        aged["Last_ShipDate"] = aged["Last_ShipDate"].astype(object)
        aged.loc[missing_mask, "Last_ShipDate"] = fallback

    # --- Sort ---
    aged = aged.sort_values(["Origin", "Vanning Date", "MODULE#"]).reset_index(drop=True)

    # --- Output columns ---
    output_cols = ["LOCATION", "Location2", "MODULE#", "PRODUCT", "QUANTITY",
                   "Origin", "Vanning Date", "Last_ShipDate"]
    result = aged[output_cols].copy()

    # --- Build summary pivot: module count by Origin × Vanning Month ---
    pivot = result.copy()
    pivot["Vanning Month"] = pd.to_datetime(pivot["Vanning Date"]).dt.to_period("M").dt.to_timestamp()
    summary = pivot.pivot_table(
        index="Origin", columns="Vanning Month", values="MODULE#",
        aggfunc="count", fill_value=0, margins=True, margins_name="Total"
    )
    summary.columns = [c.strftime("%b %Y") if isinstance(c, datetime) else c for c in summary.columns]
    summary.index.name = None

    # --- Write Excel ---
    print(f"Writing {OUTPUT_FILE} ...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", datetime_format="YYYY-MM-DD") as writer:
        result.to_excel(writer, sheet_name="Aged Inventory", index=False)
        summary.to_excel(writer, sheet_name="Summary")
        _format_summary(writer, summary, cutoff)
        _format_detail(writer, result)

    # --- Console summary ---
    print()
    print("=" * 50)
    print(f"AGED INVENTORY TOTAL: {len(result):,} modules")
    print("=" * 50)
    print()
    print("Breakdown by origin:")
    breakdown = result.groupby("Origin").size()
    for origin, count in breakdown.items():
        print(f"  {origin}: {count:,}")

    no_ship = result["Last_ShipDate"].isna().sum()
    if no_ship > 0:
        print(f"\n  ({no_ship} modules have no matching ship date in 201)")

    print(f"\nOutput saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
