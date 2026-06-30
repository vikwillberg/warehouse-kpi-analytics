import pandas as pd
import datetime as dt
import heapq
from collections import defaultdict
from pathlib import Path

class FIFOAnalyzer:
    
    def __init__(self, data_folder, output_folder='./DS FIFO Pick/Power BI Data/'):
        self.data_folder = Path(data_folder)
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
    
    def clean_part_number(self, part_str):
        if pd.isna(part_str):
            return ''
        return str(part_str).replace(' ', '').replace('-', '')

    def extract_vanning_date(self, module_str, eta_val=None):
        """
        Extract vanning date from module number using Sachiko's PQ_VanningDate lookup tables.
        
        TblOrigin: prefix -> DateTable mapping (most specific prefix wins)
          5-char: KJ699,KJ900,KJ999,22200,26700,27100,2Z400,2G400,WN000 -> DT1 (Japan)
                  KJ550..KJ698,KJ540,KA079..KA277,KJ912 -> DT2 (Mexico/Canada/China)
                  KJ911,KA085..KA331,KA118 -> DT3 (Thailand)
          4-char: KJ50-KJ54 -> DT2 (Canada), KJ55-KJ69 -> DT2 (Mexico)
          3-char: KJ0-KJ4 -> US (no DT), KJ7-KJ8 -> IS (no DT)
          1-char: S,K -> DT1 (Japan) [default]
        
        TblDate:
          DT1 (Japan): 6th char = month letter (M=1..X=12 or A=1..L=12)
          DT2 (China/Canada/Mexico): 6th-7th chars = month number (01-12)
          DT3 (Thailand): 6th char = year digit (1=2021, 2=2022, ...)
        
        Year: from ETA comparison (DT1/DT2) or module digit (DT3). Day = always 1.
        AAA modules: DT2 origins -> (ETA_year, ETA_month, 1); others -> ETA minus 2 months.
        """
        try:
            if pd.isna(module_str):
                return None
            module_str = str(module_str).strip().upper()
            
            if len(module_str) < 6:
                return None
            
            # Parse ETA
            eta_date = None
            if eta_val is not None and not pd.isna(eta_val):
                try:
                    if isinstance(eta_val, str):
                        eta_date = pd.to_datetime(eta_val).date()
                    elif isinstance(eta_val, (dt.datetime, pd.Timestamp)):
                        eta_date = eta_val.date()
                    elif isinstance(eta_val, dt.date):
                        eta_date = eta_val
                    else:
                        eta_date = pd.to_datetime(eta_val).date()
                except Exception:
                    eta_date = None
            
            # --- Prefix -> DateTable lookup (Sachiko's TblOrigin) ---
            dt1_5 = {'KJ699', 'KJ900', 'KJ999', '22200', '26700', '27100', '2Z400', '2G400', 'WN000'}
            dt2_5 = {
                'KJ550', 'KJ552', 'KJ563', 'KJ579', 'KJ598', 'KJ617', 'KJ621', 'KJ623', 'KJ624', 'KJ626',
                'KJ646', 'KJ656', 'KJ694', 'KJ698', 'KJ540', 'KA079', 'KA125', 'KA184', 'KA255', 'KA277',
                'KJ912'
            }
            dt3_5 = {'KJ911', 'KA085', 'KA152', 'KA158', 'KA199', 'KA224', 'KA261', 'KA296', 'KA331', 'KA118'}
            dt2_4 = {
                'KJ50', 'KJ51', 'KJ52', 'KJ53', 'KJ54',
                'KJ55', 'KJ56', 'KJ57', 'KJ58', 'KJ59',
                'KJ60', 'KJ61', 'KJ62', 'KJ63', 'KJ64',
                'KJ65', 'KJ66', 'KJ67', 'KJ68', 'KJ69'
            }
            no_dt_3 = {'KJ0', 'KJ1', 'KJ2', 'KJ3', 'KJ4', 'KJ7', 'KJ8'}
            
            prefix5 = module_str[:5]
            prefix4 = module_str[:4]
            prefix3 = module_str[:3]
            
            date_table = None
            if prefix5 in dt1_5:
                date_table = 1
            elif prefix5 in dt2_5:
                date_table = 2
            elif prefix5 in dt3_5:
                date_table = 3
            elif prefix4 in dt2_4:
                date_table = 2
            elif prefix3 in no_dt_3:
                return None  # US/IS origins — no date table
            elif module_str[0] in ('S', 'K'):
                date_table = 1  # Default: Japan
            else:
                return None
            
            # --- AAA modules: special ETA-based logic ---
            if 'AAA' in module_str and eta_date is not None:
                if date_table == 2:
                    # Mexico/Canada/China: vanning = (ETA_year, ETA_month, 1)
                    return dt.date(eta_date.year, eta_date.month, 1)
                else:
                    # Japan/Thailand/other: vanning = ETA minus 2 months
                    month = eta_date.month - 2
                    year = eta_date.year
                    if month <= 0:
                        month += 12
                        year -= 1
                    return dt.date(year, month, 1)
            
            # --- Non-AAA: extract from module string ---
            month_map_dt1 = {
                'M': 1, 'N': 2, 'O': 3, 'P': 4, 'Q': 5, 'R': 6,
                'S': 7, 'T': 8, 'U': 9, 'V': 10, 'W': 11, 'X': 12,
                'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6,
                'G': 7, 'H': 8, 'I': 9, 'J': 10, 'K': 11, 'L': 12
            }
            
            if date_table == 1:
                # DateTable1 (Japan): 6th char -> month
                month = month_map_dt1.get(module_str[5])
                if month is None:
                    return None
                if eta_date is not None:
                    year = eta_date.year - 1 if month > eta_date.month else eta_date.year
                else:
                    today = dt.date.today()
                    year = today.year - 1 if month > today.month else today.year
                return dt.date(year, month, 1)
            
            elif date_table == 2:
                # DateTable2 (China/Canada/Mexico): 6th-7th chars -> month number
                if len(module_str) < 7:
                    return None
                month_str = module_str[5:7]
                if not month_str.isdigit():
                    return None
                month = int(month_str)
                if not (1 <= month <= 12):
                    return None
                if eta_date is not None:
                    year = eta_date.year - 1 if month > eta_date.month else eta_date.year
                else:
                    today = dt.date.today()
                    year = today.year - 1 if month > today.month else today.year
                return dt.date(year, month, 1)
            
            elif date_table == 3:
                # DateTable3 (Thailand): 6th char -> year digit (1=2021, 2=2022, etc.)
                year_char = module_str[5]
                if not year_char.isdigit():
                    return None
                year = 2020 + int(year_char)
                # Month from ETA - 2 offset
                if eta_date is not None:
                    month = eta_date.month - 2
                    if month <= 0:
                        month += 12
                        year -= 1
                    return dt.date(year, month, 1)
                return dt.date(year, 6, 1)  # Default mid-year if no ETA
            
            return None
            
        except Exception:
            return None
        
    def find_picking_file(self):
        """
        Find the most recent picking file.
        The picking file has a dynamic name like: Picking_MODULE_LIST_SITE1_CUST1_YYYYMMDD_HHMMSS.csv
        """
        # Look for files matching the pattern
        pattern = 'Picking_MODULE_LIST_*.csv'
        picking_files = list(self.data_folder.glob(pattern))
        
        if not picking_files:
            # Fallback to the original static name
            fallback = self.data_folder / 'Picking_MODULE_LIST.csv'
            if fallback.exists():
                return fallback
            raise FileNotFoundError(
                f"No picking file found matching '{pattern}' or 'Picking_MODULE_LIST.csv' "
                f"in {self.data_folder}"
            )
        
        # Return the most recently modified file
        most_recent = max(picking_files, key=lambda p: p.stat().st_mtime)
        print(f"Using picking file: {most_recent.name}")
        return most_recent

    def load_module_loc_files(self):
        """
        Combine all CSV files in the 'MODULE_LOC Files' subfolder into one
        deduplicated DataFrame. Duplicates are identified by MODULE# and the
        record from the most-recently-modified file is kept.
        """
        module_loc_dir = self.data_folder / 'MODULE_LOC Files'
        csv_files = sorted(module_loc_dir.glob('*.csv'), key=lambda p: p.stat().st_mtime)

        if not csv_files:
            raise FileNotFoundError(
                f"No CSV files found in {module_loc_dir}"
            )

        print(f"Found {len(csv_files)} MODULE_LOC file(s) to combine:")
        frames = []
        for f in csv_files:
            print(f"  Reading {f.name}")
            df = pd.read_csv(f, low_memory=False)
            frames.append(df)

        df_combined = pd.concat(frames, ignore_index=True)
        before = len(df_combined)
        # Keep the last occurrence per MODULE# (comes from the most recent file)
        df_combined = df_combined.drop_duplicates(subset=['MODULE#'], keep='last')
        after = len(df_combined)
        print(f"Combined {before:,} rows -> {after:,} unique modules ({before - after:,} duplicates removed)")
        return df_combined

    def load_and_merge_data(self):
        """
        Load picking and inventory
        Extract VANNING DATES from module numbers (not arrival dates!)
        """
        print("\n=== Loading Data ===")
        
        # Load picking data
        pick_file = self.find_picking_file()
        df_pick = pd.read_csv(pick_file)
        df_pick['PROCESS TIME'] = df_pick['PROCESS TIME'].apply(lambda x: int(float(x)))
        
        # Clean part numbers
        df_pick['PRODUCT CODE'] = df_pick['PRODUCT CODE'].apply(self.clean_part_number)
        
        # Normalize operator codes
        df_pick['OPCD'] = df_pick['OPCD'].fillna('').astype(str).str.replace(r'\s+', '', regex=True).str.upper().str.replace(r'\.0$', '', regex=True)
        
        df_pick = df_pick.sort_values('PROCESS TIME').reset_index(drop=True)
        print(f"Loaded {len(df_pick):,} picks (all available data)")
        
        # Load current inventory (combined from all MODULE_LOC Files)
        df_inv = self.load_module_loc_files()
        
        # Filter and clean
        df_inv = df_inv[df_inv['DAMAGE'] != 'Y'].copy()
        df_inv = df_inv[~df_inv['LOCATION'].str.contains('-HLD-', na=False)].copy()
        df_inv = df_inv[~df_inv['LOCATION'].str.contains('EC-QPC', na=False)].copy()
        
        df_inv['PRODUCT'] = df_inv['PRODUCT'].apply(self.clean_part_number)
        
        print(f"Loaded {len(df_inv):,} current inventory modules")
        
        # Extract VANNING DATES for current inventory
        print("Extracting vanning dates from current inventory module numbers...")
        df_inv['VANNING DATE'] = df_inv.apply(
            lambda row: self.extract_vanning_date(row['MODULE#'], row.get('ETA')), 
            axis=1
        )
        
        vanning_dated = df_inv['VANNING DATE'].notna().sum()
        print(f"  Vanning dates extracted: {vanning_dated:,}/{len(df_inv):,} ({vanning_dated/len(df_inv)*100:.1f}%)")
        
        # Create lookup for vanning dates
        inv_vanning_lookup = dict(zip(df_inv['MODULE#'], df_inv['VANNING DATE']))
        
        # Add picked modules to inventory with VANNING DATES
        picked_modules_data = []
        date_sources = {'from_inventory': 0, 'extracted': 0, 'no_date': 0}
        
        for _, row in df_pick.iterrows():
            module = row['MODULE NO']
            
            # Check if module is in current inventory first
            vanning_date = inv_vanning_lookup.get(module)
            
            if vanning_date is not None and not pd.isna(vanning_date):
                date_sources['from_inventory'] += 1
            else:
                # Don't fall back to string extraction — if the module isn't in
                # the MODULE_LOC inventory, it has no "Picked" date (Sachiko's method)
                date_sources['no_date'] += 1
            
            picked_modules_data.append({
                'MODULE#': module,
                'PRODUCT': row['PRODUCT CODE'],
                'VANNING DATE': vanning_date,
                'LOCATION': row.get('PICK LOCATION', ''),
                'DAMAGE': 'N'
            })
        
        df_picked_modules = pd.DataFrame(picked_modules_data)
        print(f"\nCreated {len(df_picked_modules):,} picked module records")
        print(f"  Vanning dates - from inventory: {date_sources['from_inventory']:,}, extracted: {date_sources['extracted']:,}, none: {date_sources['no_date']:,}")
        
        # Merge: current inventory + picked modules
        df_inv_combined = pd.concat([df_inv[['MODULE#', 'PRODUCT', 'VANNING DATE', 'LOCATION', 'DAMAGE']], 
                                    df_picked_modules], ignore_index=True)
        
        # Remove duplicates (keep first = current inventory version)
        df_inv_combined = df_inv_combined.drop_duplicates(subset=['MODULE#'], keep='first')
        
        print(f"Combined inventory: {len(df_inv_combined):,} total modules")
        
        return df_pick, df_inv, df_inv_combined
    
    def check_fifo(self, df_pick, df_inv):
        """
        FIFO logic using VANNING DATE
        """
        print("\n=== Calculating FIFO Compliance (VANNING DATE) ===")
        
        total = len(df_pick)
        oldest_dates = [None] * total
        picked_dates = [None] * total
        fifo_flags = [0] * total

        module_to_part = dict(zip(df_inv['MODULE#'], df_inv['PRODUCT']))
        module_to_date = dict(zip(df_inv['MODULE#'], df_inv['VANNING DATE']))
        module_to_pos = {module: pos for pos, module in enumerate(df_inv['MODULE#'])}

        available_count = df_inv['PRODUCT'].value_counts().to_dict()

        part_heaps = defaultdict(list)
        for module, part, vanning_date in df_inv[['MODULE#', 'PRODUCT', 'VANNING DATE']].itertuples(index=False, name=None):
            if pd.notna(vanning_date):
                part_heaps[part].append((vanning_date, module))
        for heap in part_heaps.values():
            heapq.heapify(heap)

        picked_modules = set()

        columns = df_pick.columns
        idx_part = columns.get_loc('PRODUCT CODE')
        idx_module = columns.get_loc('MODULE NO')
        idx_time = columns.get_loc('PROCESS TIME')

        for idx, row in enumerate(df_pick.itertuples(index=False, name=None)):
            part = row[idx_part]
            module = row[idx_module]
            pick_time = row[idx_time]

            has_available = available_count.get(part, 0) > 0
            if not has_available:
                fifo_flags[idx] = 1
            else:
                heap = part_heaps.get(part)
                if heap:
                    while heap and heap[0][1] in picked_modules:
                        heapq.heappop(heap)
                oldest_date = heap[0][0] if heap else None

                if oldest_date is None:
                    fifo_flags[idx] = 1
                else:
                    oldest_dates[idx] = oldest_date
                    picked_date = None
                    if module not in picked_modules:
                        module_date = module_to_date.get(module)
                        if pd.notna(module_date):
                            picked_date = module_date
                    if picked_date is None:
                        fifo_flags[idx] = 0
                    else:
                        picked_dates[idx] = picked_date
                        fifo_flags[idx] = 1 if picked_date == oldest_date else 0

            if module in module_to_pos and module not in picked_modules:
                picked_modules.add(module)
                part_for_module = module_to_part[module]
                available_count[part_for_module] = available_count.get(part_for_module, 0) - 1

            if (idx + 1) % 500 == 0:
                print(f"Processed {idx + 1:,} / {total:,} ({(idx+1)/total*100:.1f}%)")

        df_pick['oldest_vanning_date'] = oldest_dates
        df_pick['picked_vanning_date'] = picked_dates
        df_pick['fifo_compliant'] = fifo_flags
        
        print("FIFO calculation complete!")
        return df_pick
    
    def export_oldest_location_file(self, df_inv):
        """
        Export the oldest location by part for Power BI lookup table.
        This helps warehouse operators find where the oldest module for each part is located.
        """
        print("\n=== Exporting Oldest Location by Part ===")
        
        # Filter to modules with valid vanning dates only
        df_with_dates = df_inv[df_inv['VANNING DATE'].notna()].copy()
        
        if len(df_with_dates) == 0:
            print("Warning: No modules with valid vanning dates found")
            # Create empty file with headers
            oldest_export = pd.DataFrame(columns=['PartNumber', 'OldestLocation', 'OldestVanningDate', 'OldestModuleNumber'])
        else:
            # Group by part and find THE SINGLE oldest module (minimum vanning date)
            oldest_by_part = df_with_dates.loc[
                df_with_dates.groupby('PRODUCT')['VANNING DATE'].idxmin()
            ].copy()
            
            # Prepare export dataframe
            oldest_export = pd.DataFrame({
                'PartNumber': oldest_by_part['PRODUCT'],
                'OldestLocation': oldest_by_part['LOCATION'],
                'OldestVanningDate': oldest_by_part['VANNING DATE'],
                'OldestModuleNumber': oldest_by_part['MODULE#']
            })
            
            # Sort by part number for easier lookup
            oldest_export = oldest_export.sort_values('PartNumber').reset_index(drop=True)
            
            print(f"✓ Exported oldest location for {len(oldest_export)} parts")
            
            # Show sample
            if len(oldest_export) > 0:
                print(f"\nSample entries:")
                print(oldest_export.head(3).to_string(index=False))
        
        # Export to CSV
        oldest_file = self.output_folder / 'PowerBI_OldestLocationByPart.csv'
        oldest_export.to_csv(oldest_file, index=False)
        print(f"✓ Oldest location file: {oldest_file}")
        
        return oldest_file

    def get_oldest_location_lookup(self, df_inv):
        """Create lookup dictionary for oldest location by part"""
        print("\n=== Creating Oldest Location Lookup ===")
        
        # Filter to modules with valid vanning dates only
        df_with_dates = df_inv[df_inv['VANNING DATE'].notna()].copy()
        
        # Group by part and find THE SINGLE oldest module
        oldest_by_part = df_with_dates.loc[
            df_with_dates.groupby('PRODUCT')['VANNING DATE'].idxmin()
        ].copy()
        
        # Create lookup dictionaries
        location_lookup = dict(zip(oldest_by_part['PRODUCT'], oldest_by_part['LOCATION']))
        module_lookup = dict(zip(oldest_by_part['PRODUCT'], oldest_by_part['MODULE#']))
        
        print(f"✓ Created lookup for {len(location_lookup)} parts")
        
        return location_lookup, module_lookup

    def _prepare_pick_data(self, df_pick):
        """Parse dates and shifts from PROCESS TIME. Shared by all report/export methods."""
        df_pick['Operation_Date'] = pd.to_datetime(
            df_pick['PROCESS TIME'].astype(str), format='%Y%m%d%H%M', errors='coerce'
        )
        df_pick['Operation_Date'] = df_pick['Operation_Date'].fillna(pd.Timestamp('1900-01-01'))
        df_pick['Operation_Date_Only'] = df_pick['Operation_Date'].apply(
            lambda dt: (dt - pd.Timedelta(days=1)).date() if dt.hour < 5 else dt.date()
        )

        def get_shift(dt_val):
            hour = dt_val.hour
            if hour < 5:
                return 'shift2'
            elif 6 <= hour < 16 or (hour == 16 and dt_val.minute < 30):
                return 'shift1'
            else:
                return 'shift2'

        df_pick['Shift'] = df_pick['Operation_Date'].apply(get_shift)
        return df_pick

    def generate_report(self, df_pick):
        print("\n" + "="*70)
        print("VANNING DATE FIFO COMPLIANCE SUMMARY")
        print("="*70)
        
        total = len(df_pick)
        fifo = df_pick['fifo_compliant'].sum()
        rate = (fifo / total * 100) if total > 0 else 0
        
        print(f"\nOverall: {fifo:,}/{total:,} ({rate:.1f}% FIFO compliant)")
        
        if 'Operation_Date_Only' not in df_pick.columns:
            df_pick = self._prepare_pick_data(df_pick)
        df_pick['date'] = df_pick['Operation_Date_Only']
        
        # 1. Last 10 Work Days Summary
        by_date = df_pick.groupby('date')['fifo_compliant'].agg(['sum', 'count']).reset_index()
        by_date.columns = ['Date', 'FIFO_Picks', 'Total_Picks']
        by_date['FIFO_Percentage'] = (by_date['FIFO_Picks'] / by_date['Total_Picks'] * 100).round(1)
        by_date = by_date.sort_values('Date').tail(10)
        
        print("\n--- Last 10 Work Days ---")
        print(by_date.to_string(index=False))
        
        # 2. Weekly Average (Monday to Saturday weeks)
        df_pick['Week_Start'] = pd.to_datetime(df_pick['date']).dt.to_period('W-SUN').dt.start_time.dt.date
        weekly_summary = df_pick.groupby('Week_Start')['fifo_compliant'].agg(['sum', 'count']).reset_index()
        weekly_summary.columns = ['Week_Start', 'FIFO_Picks', 'Total_Picks']
        weekly_summary['FIFO_Percentage'] = (weekly_summary['FIFO_Picks'] / weekly_summary['Total_Picks'] * 100).round(1)
        
        print("\n--- Weekly Average ---")
        print(weekly_summary.to_string(index=False))
        
        # 3. Part Summary (Last 10 days) — uses single-pass FIFO for part metrics
        part_flag = 'fifo_compliant_part' if 'fifo_compliant_part' in df_pick.columns else 'fifo_compliant'
        daily_part_details = df_pick.groupby(['date', 'PRODUCT CODE'])[part_flag].agg(['count', 'sum']).reset_index()
        daily_part_details['is_fifo_part'] = daily_part_details['sum'] > 0
        part_summary = daily_part_details.groupby('date').agg(
            Picked_Part_count=('PRODUCT CODE', 'count'),
            FIFO_Picked_Part_count=('is_fifo_part', 'sum')
        ).reset_index()
        part_summary['FIFO_Percentage'] = (part_summary['FIFO_Picked_Part_count'] / part_summary['Picked_Part_count'] * 100).round(1)
        
        part_summary = part_summary.sort_values('date').tail(10)
        part_summary = part_summary.rename(columns={'date': 'Date'})
        
        print("\n--- Daily Part Summary (Last 10 Work Days) ---")
        print(part_summary.to_string(index=False))
        
        # 4. Weekly Part Summary — uses single-pass FIFO for part metrics
        daily_part_details = df_pick.groupby(['date', 'PRODUCT CODE'])[part_flag].agg(['count', 'sum']).reset_index()
        daily_part_details['is_fifo_part'] = daily_part_details['sum'] > 0
        
        # First calculate the daily part summary
        daily_part_summary = daily_part_details.groupby('date').agg(
            Picked_Part_count=('PRODUCT CODE', 'count'),
            FIFO_Picked_Part_count=('is_fifo_part', 'sum')
        ).reset_index()
        daily_part_summary['Daily_FIFO_Percentage'] = (daily_part_summary['FIFO_Picked_Part_count'] / daily_part_summary['Picked_Part_count'] * 100)
        
        # Then calculate the weekly sum and average of the daily percentages
        daily_part_summary['Week_Start'] = pd.to_datetime(daily_part_summary['date']).dt.to_period('W-SUN').dt.start_time.dt.date
        
        weekly_part_summary = daily_part_summary.groupby('Week_Start').agg(
            Total_Weekly_Parts=('Picked_Part_count', 'sum'),
            Total_Weekly_FIFO_Parts=('FIFO_Picked_Part_count', 'sum'),
            Avg_FIFO_Percentage=('Daily_FIFO_Percentage', 'mean')
        ).reset_index()
        
        # Round the values for display
        weekly_part_summary['Avg_FIFO_Percentage'] = weekly_part_summary['Avg_FIFO_Percentage'].round(1)
        
        print("\n--- Weekly Part Summary ---")
        print(weekly_part_summary.to_string(index=False))
        
        return rate
    

    def export_kpi_workbook(self, df_pick, df_inv):
        """Generate styled KPI Excel workbook with formatted tables, colors, and charts."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter

        print("\n=== Generating KPI Workbook ===")

        # ---- Color palette ----
        TEAL = PatternFill(start_color='00796B', end_color='00796B', fill_type='solid')
        DARK_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        LIGHT_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        LIGHT_RED = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        WHITE_FONT = Font(color='FFFFFF', bold=True, size=11)
        BOLD_FONT = Font(bold=True, size=11)
        CENTER = Alignment(horizontal='center', vertical='center')
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def style_header(ws, row, max_col, fill=TEAL, font=WHITE_FONT):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = fill
                cell.font = font
                cell.alignment = CENTER
                cell.border = THIN_BORDER

        def auto_width(ws, max_col, min_width=10, max_width=22):
            for c in range(1, max_col + 1):
                letter = get_column_letter(c)
                max_len = min_width
                for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)) + 2)
                ws.column_dimensions[letter].width = min(max_len, max_width)

        location_lookup, module_lookup = self.get_oldest_location_lookup(df_inv)

        if 'Operation_Date_Only' not in df_pick.columns:
            df_pick = self._prepare_pick_data(df_pick)

        df_pick['Oldest_Vanning_Display'] = df_pick['oldest_vanning_date'].apply(
            lambda x: x.strftime('%m/%d/%Y') if pd.notna(x) and x is not None else ''
        )
        df_pick['Picked_Vanning_Display'] = df_pick['picked_vanning_date'].apply(
            lambda x: x.strftime('%m/%d/%Y') if pd.notna(x) and x is not None else ''
        )

        # Always use previous week (Monday-Saturday) regardless of when report is created
        from datetime import timedelta
        today = dt.date.today()
        days_since_monday = today.weekday()  # 0=Mon, 1=Tue, ..., 6=Sun
        previous_monday = today - timedelta(days=days_since_monday + 7)
        previous_saturday = previous_monday + timedelta(days=5)  # Monday + 5 = Saturday
        
        dates = sorted(df_pick['Operation_Date_Only'].unique())
        week_dates = [d for d in dates if d >= previous_monday and d <= previous_saturday]
        wb_file = self.data_folder.parent / 'Output' / f'DS_FIFO_KPI_Data_{previous_monday.strftime("%Y%m%d")}.xlsx'

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ==================================================================
        # 0. SUMMARY DASHBOARD SHEET
        # ==================================================================
        ws_dash = wb.create_sheet(title='Summary')
        week_picks = df_pick[df_pick['Operation_Date_Only'].isin(week_dates)].copy()

        daily_summary = week_picks.groupby('Operation_Date_Only').agg(
            Total_Picks=('fifo_compliant', 'count'),
            FIFO_Picks=('fifo_compliant', 'sum')
        ).reset_index().sort_values('Operation_Date_Only')
        daily_summary['FIFO_%'] = (daily_summary['FIFO_Picks'] / daily_summary['Total_Picks'] * 100).round(1)
        daily_summary['Non_FIFO'] = daily_summary['Total_Picks'] - daily_summary['FIFO_Picks']

        # Title
        ws_dash.merge_cells('A1:H1')
        title_cell = ws_dash['A1']
        title_cell.value = f'DS FIFO Picking Compliance — Week of {week_dates[0].strftime("%m/%d")} to {week_dates[-1].strftime("%m/%d/%Y")}'
        title_cell.font = Font(bold=True, size=16, color='00796B')
        title_cell.alignment = Alignment(horizontal='center')

        # KPI cards
        total_picks = int(daily_summary['Total_Picks'].sum())
        total_fifo = int(daily_summary['FIFO_Picks'].sum())
        fifo_ratio = round(total_fifo / total_picks * 100, 1) if total_picks > 0 else 0

        for col_offset, (label, value) in enumerate([
            ('Total Picks', f'{total_picks:,}'),
            ('FIFO Picks', f'{total_fifo:,}'),
            ('FIFO Ratio', f'{fifo_ratio}%')
        ]):
            c = 2 + col_offset * 3
            ws_dash.merge_cells(start_row=3, start_column=c, end_row=3, end_column=c+1)
            ws_dash.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c+1)
            val_cell = ws_dash.cell(row=3, column=c, value=value)
            val_cell.font = Font(bold=True, size=22, color='FFFFFF')
            val_cell.fill = TEAL
            val_cell.alignment = CENTER
            ws_dash.cell(row=3, column=c+1).fill = TEAL
            lbl_cell = ws_dash.cell(row=4, column=c, value=label)
            lbl_cell.font = Font(bold=True, size=11, color='FFFFFF')
            lbl_cell.fill = DARK_BLUE
            lbl_cell.alignment = CENTER
            ws_dash.cell(row=4, column=c+1).fill = DARK_BLUE

        # Daily table
        table_start = 6
        headers = ['Date', 'Total Picks', 'FIFO Picks', 'Non-FIFO', 'FIFO %']
        for ci, h in enumerate(headers, 1):
            ws_dash.cell(row=table_start, column=ci, value=h)
        style_header(ws_dash, table_start, len(headers), fill=DARK_BLUE)

        for ri, (_, row) in enumerate(daily_summary.iterrows(), table_start + 1):
            ws_dash.cell(row=ri, column=1, value=row['Operation_Date_Only'].strftime('%m/%d/%Y')).border = THIN_BORDER
            ws_dash.cell(row=ri, column=2, value=int(row['Total_Picks'])).border = THIN_BORDER
            ws_dash.cell(row=ri, column=3, value=int(row['FIFO_Picks'])).border = THIN_BORDER
            ws_dash.cell(row=ri, column=4, value=int(row['Non_FIFO'])).border = THIN_BORDER
            pct_cell = ws_dash.cell(row=ri, column=5, value=row['FIFO_%'] / 100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = THIN_BORDER
            pct_cell.alignment = CENTER
            if row['FIFO_%'] >= 40:
                pct_cell.fill = LIGHT_GREEN
            elif row['FIFO_%'] < 25:
                pct_cell.fill = LIGHT_RED

        data_end = table_start + len(daily_summary)
        auto_width(ws_dash, 5)

        # Operator Weekly FIFO % mini-table
        op_weekly = week_picks.groupby(['Shift', 'OPCD']).agg(
            Total_Picks=('fifo_compliant', 'count'),
            FIFO_Picks=('fifo_compliant', 'sum')
        ).reset_index()
        op_weekly['FIFO_%'] = (op_weekly['FIFO_Picks'] / op_weekly['Total_Picks'] * 100).round(1)
        op_weekly = op_weekly.sort_values(['Shift', 'FIFO_%'], ascending=[True, False])

        op_table_start = data_end + 2
        ws_dash.cell(row=op_table_start, column=1, value='Operator Weekly FIFO %').font = Font(bold=True, size=13, color='00796B')
        op_table_start += 1

        op_headers = ['Shift', 'OPCD', 'Total Picks', 'FIFO Picks', 'FIFO %']
        for ci, h in enumerate(op_headers, 1):
            ws_dash.cell(row=op_table_start, column=ci, value=h)
        style_header(ws_dash, op_table_start, len(op_headers), fill=DARK_BLUE)

        for ri, (_, row) in enumerate(op_weekly.iterrows(), op_table_start + 1):
            ws_dash.cell(row=ri, column=1, value=row['Shift']).border = THIN_BORDER
            ws_dash.cell(row=ri, column=2, value=row['OPCD']).border = THIN_BORDER
            ws_dash.cell(row=ri, column=3, value=int(row['Total_Picks'])).border = THIN_BORDER
            ws_dash.cell(row=ri, column=4, value=int(row['FIFO_Picks'])).border = THIN_BORDER
            pct_cell = ws_dash.cell(row=ri, column=5, value=row['FIFO_%'] / 100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = THIN_BORDER
            pct_cell.alignment = CENTER
            if row['FIFO_%'] >= 40:
                pct_cell.fill = LIGHT_GREEN
            elif row['FIFO_%'] < 25:
                pct_cell.fill = LIGHT_RED

        # FIFO % bar chart
        chart = BarChart()
        chart.type = 'col'
        chart.grouping = 'clustered'
        chart.title = 'Daily FIFO %'
        chart.y_axis.title = 'FIFO %'
        chart.y_axis.numFmt = '0%'
        chart.x_axis.title = 'Date'
        chart.style = 10
        chart.width = 22
        chart.height = 14

        cats = Reference(ws_dash, min_col=1, min_row=table_start + 1, max_row=data_end)
        fifo_pct = Reference(ws_dash, min_col=5, min_row=table_start, max_row=data_end)
        chart.add_data(fifo_pct, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = '00796B'
        ws_dash.add_chart(chart, 'G6')

        # ==================================================================
        # 1. PER OPERATOR SHEET
        # ==================================================================
        ws_op = wb.create_sheet(title='Per Operator')

        op_summary = week_picks.groupby(['Shift', 'OPCD', 'Operation_Date_Only']).agg(
            Total_Pick=('fifo_compliant', 'count'),
            FIFO_Pick=('fifo_compliant', 'sum')
        ).reset_index()

        operators = op_summary.groupby(['Shift', 'OPCD']).size().reset_index()[['Shift', 'OPCD']]
        operators = operators.sort_values(['Shift', 'OPCD']).reset_index(drop=True)

        col = 1
        ws_op.cell(row=2, column=col, value='Shift'); col += 1
        ws_op.cell(row=2, column=col, value='OPCD'); col += 1

        date_start_cols = {}
        for date_val in week_dates:
            date_label = pd.Timestamp(date_val).strftime('%m/%d')
            ws_op.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
            ws_op.cell(row=1, column=col, value=date_label)
            date_start_cols[date_val] = col
            ws_op.cell(row=2, column=col, value='Total Pick')
            ws_op.cell(row=2, column=col + 1, value='FIFO Pick')
            ws_op.cell(row=2, column=col + 2, value='Ratio')
            col += 3

        weekly_start_col = col
        ws_op.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        ws_op.cell(row=1, column=col, value='Weekly Total')
        ws_op.cell(row=2, column=col, value='Total Pick')
        ws_op.cell(row=2, column=col + 1, value='FIFO Pick')
        ws_op.cell(row=2, column=col + 2, value='Ratio')
        col += 3
        ws_op.cell(row=2, column=col, value='OPCD')
        last_col = col

        style_header(ws_op, 1, last_col, fill=DARK_BLUE)
        style_header(ws_op, 2, last_col, fill=TEAL)

        row_num = 3
        for _, op_row in operators.iterrows():
            shift = op_row['Shift']
            opcd = op_row['OPCD']
            ws_op.cell(row=row_num, column=1, value=shift).border = THIN_BORDER
            ws_op.cell(row=row_num, column=2, value=opcd).border = THIN_BORDER

            weekly_total = 0
            weekly_fifo = 0

            for date_val in week_dates:
                dcol = date_start_cols[date_val]
                match = op_summary[(op_summary['Shift'] == shift) &
                                   (op_summary['OPCD'] == opcd) &
                                   (op_summary['Operation_Date_Only'] == date_val)]
                if len(match) > 0:
                    tp = int(match['Total_Pick'].values[0])
                    fp = int(match['FIFO_Pick'].values[0])
                    ratio = round(fp / tp * 100, 1) if tp > 0 else 0.0
                    ws_op.cell(row=row_num, column=dcol, value=tp).border = THIN_BORDER
                    ws_op.cell(row=row_num, column=dcol + 1, value=fp).border = THIN_BORDER
                    ratio_cell = ws_op.cell(row=row_num, column=dcol + 2, value=ratio / 100)
                    ratio_cell.number_format = '0.0%'
                    ratio_cell.border = THIN_BORDER
                    ratio_cell.alignment = CENTER
                    if ratio >= 40:
                        ratio_cell.fill = LIGHT_GREEN
                    elif ratio < 25:
                        ratio_cell.fill = LIGHT_RED
                    weekly_total += tp
                    weekly_fifo += fp
                else:
                    for offset in range(3):
                        ws_op.cell(row=row_num, column=dcol + offset).border = THIN_BORDER

            weekly_ratio = round(weekly_fifo / weekly_total * 100, 1) if weekly_total > 0 else 0.0
            ws_op.cell(row=row_num, column=weekly_start_col, value=weekly_total if weekly_total > 0 else '-').border = THIN_BORDER
            ws_op.cell(row=row_num, column=weekly_start_col + 1, value=weekly_fifo if weekly_total > 0 else '-').border = THIN_BORDER
            wk_cell = ws_op.cell(row=row_num, column=weekly_start_col + 2, value=weekly_ratio / 100)
            wk_cell.number_format = '0.0%'
            wk_cell.border = THIN_BORDER
            wk_cell.alignment = CENTER
            wk_cell.font = BOLD_FONT
            if weekly_ratio >= 40:
                wk_cell.fill = LIGHT_GREEN
            elif weekly_ratio < 25:
                wk_cell.fill = LIGHT_RED
            ws_op.cell(row=row_num, column=last_col, value=opcd).border = THIN_BORDER

            if row_num % 2 == 0:
                for c in range(1, last_col + 1):
                    cell = ws_op.cell(row=row_num, column=c)
                    if cell.fill == PatternFill():
                        cell.fill = LIGHT_BLUE

            row_num += 1

        auto_width(ws_op, last_col, min_width=8, max_width=14)

        # ==================================================================
        # 2. APPEND SHEET
        # ==================================================================
        ws_app = wb.create_sheet(title='Append1')
        ws_app.cell(row=2, column=1, value='Operation Date')
        ws_app.cell(row=2, column=2, value='Values')

        col = 3
        append_date_cols = {}
        for date_val in week_dates:
            try:
                date_str = pd.Timestamp(date_val).strftime('%m/%d/%Y')
            except Exception:
                date_str = str(date_val)
            ws_app.cell(row=3, column=col, value=date_str)
            ws_app.cell(row=4, column=col, value='Sum of Total Pick')
            ws_app.cell(row=3, column=col + 1, value=date_str)
            ws_app.cell(row=4, column=col + 1, value='Sum of FIFO Pick')
            append_date_cols[date_val] = col
            col += 2

        style_header(ws_app, 3, col - 1, fill=DARK_BLUE)
        style_header(ws_app, 4, col - 1, fill=TEAL)

        row_num = 5
        for _, op_row in operators.iterrows():
            shift = op_row['Shift']
            opcd = op_row['OPCD']
            for date_val in week_dates:
                acol = append_date_cols[date_val]
                match = op_summary[(op_summary['Shift'] == shift) &
                                   (op_summary['OPCD'] == opcd) &
                                   (op_summary['Operation_Date_Only'] == date_val)]
                if len(match) > 0:
                    ws_app.cell(row=row_num, column=acol, value=int(match['Total_Pick'].values[0])).border = THIN_BORDER
                    ws_app.cell(row=row_num, column=acol + 1, value=int(match['FIFO_Pick'].values[0])).border = THIN_BORDER
            row_num += 1

        # ==================================================================
        # 3. PER-DAY SHEETS
        # ==================================================================
        for date_val in week_dates:
            sheet_name = pd.Timestamp(date_val).strftime('%m%d')
            ws = wb.create_sheet(title=sheet_name)

            day_picks = df_pick[df_pick['Operation_Date_Only'] == date_val].copy()
            day_picks = day_picks.sort_values('PROCESS TIME')

            pick_headers = ['OPCD', 'PRODUCT CODE', 'MODULE NO', 'PICK LOCATION',
                           'Operation Date', 'Operation DateTime', 'Shift',
                           'Oldest Vanning', 'Picked', 'FIFO Pick']
            for col_idx, header in enumerate(pick_headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            style_header(ws, 1, len(pick_headers))

            for row_idx, (_, row) in enumerate(day_picks.iterrows(), 2):
                ws.cell(row=row_idx, column=1, value=row['OPCD']).border = THIN_BORDER
                ws.cell(row=row_idx, column=2, value=row['PRODUCT CODE']).border = THIN_BORDER
                ws.cell(row=row_idx, column=3, value=row['MODULE NO']).border = THIN_BORDER
                ws.cell(row=row_idx, column=4, value=row.get('PICK LOCATION', '')).border = THIN_BORDER
                ws.cell(row=row_idx, column=5, value=str(date_val)).border = THIN_BORDER
                ws.cell(row=row_idx, column=6, value=row['Operation_Date'].strftime('%m/%d/%Y %H:%M')).border = THIN_BORDER
                ws.cell(row=row_idx, column=7, value=row['Shift']).border = THIN_BORDER
                ws.cell(row=row_idx, column=8, value=row['Oldest_Vanning_Display']).border = THIN_BORDER
                ws.cell(row=row_idx, column=9, value=row['Picked_Vanning_Display']).border = THIN_BORDER
                fifo_cell = ws.cell(row=row_idx, column=10, value=int(row['fifo_compliant']))
                fifo_cell.border = THIN_BORDER
                fifo_cell.alignment = CENTER
                fifo_cell.fill = LIGHT_GREEN if row['fifo_compliant'] == 1 else LIGHT_RED

            # Sidebar part summary
            part_day = day_picks.groupby('PRODUCT CODE').agg(
                Total_Picks=('fifo_compliant', 'count'),
                FIFO_Picks=('fifo_compliant', 'sum')
            ).reset_index().sort_values('Total_Picks', ascending=False)

            sidebar_col = 12
            side_headers = ['Product code', 'Total Picks', 'FIFO Picks']
            for ci, h in enumerate(side_headers):
                ws.cell(row=1, column=sidebar_col + ci, value=h)
            style_header(ws, 1, sidebar_col + len(side_headers) - 1)

            for r_idx, (_, prow) in enumerate(part_day.iterrows(), 2):
                ws.cell(row=r_idx, column=sidebar_col, value=prow['PRODUCT CODE']).border = THIN_BORDER
                ws.cell(row=r_idx, column=sidebar_col + 1, value=int(prow['Total_Picks'])).border = THIN_BORDER
                ws.cell(row=r_idx, column=sidebar_col + 2, value=int(prow['FIFO_Picks'])).border = THIN_BORDER

            auto_width(ws, sidebar_col + 2)

        wb.save(wb_file)
        print(f"  KPI Workbook: {wb_file}")
        print(f"  Sheets: {', '.join(wb.sheetnames)}")
        print("\n=== KPI Workbook Ready ===")

    def export_for_powerbi(self, df_pick, df_inv):
        """Export single consolidated FIFO data file for Power BI dashboard"""
        print("\n=== Exporting for Power BI ===")

        location_lookup, module_lookup = self.get_oldest_location_lookup(df_inv)

        if 'Operation_Date_Only' not in df_pick.columns:
            df_pick = self._prepare_pick_data(df_pick)

        # 1. Base grouping: Date x Part x Operator
        df_group = df_pick.groupby(['Operation_Date_Only', 'PRODUCT CODE', 'OPCD']).agg(
            total_picks=('fifo_compliant', 'count'),
            fifo_picks=('fifo_compliant', 'sum')
        ).reset_index()

        df_group.rename(columns={
            'Operation_Date_Only': 'date',
            'PRODUCT CODE': 'part_number',
            'OPCD': 'operator'
        }, inplace=True)

        df_group['non_fifo_picks'] = df_group['total_picks'] - df_group['fifo_picks']

        # 2. Daily-level FIFO % (same value for every row on a given date)
        daily_fifo = df_pick.groupby('Operation_Date_Only').agg(
            daily_total=('fifo_compliant', 'count'),
            daily_fifo=('fifo_compliant', 'sum')
        ).reset_index()
        daily_fifo.rename(columns={'Operation_Date_Only': 'date'}, inplace=True)
        daily_fifo['fifo%'] = (daily_fifo['daily_fifo'] / daily_fifo['daily_total'] * 100).astype(float)
        daily_fifo['fifo%_fraction'] = (daily_fifo['daily_fifo'] / daily_fifo['daily_total']).astype(float)

        # 3. Part-level metrics: daily % of unique parts with at least one FIFO pick
        part_flag = 'fifo_compliant_part' if 'fifo_compliant_part' in df_pick.columns else 'fifo_compliant'
        part_day = df_pick.groupby(['Operation_Date_Only', 'PRODUCT CODE']).agg(
            part_fifo_picks=(part_flag, 'sum')
        ).reset_index()
        part_day['is_fifo_part'] = (part_day['part_fifo_picks'] > 0).astype(int)

        daily_part_summary = part_day.groupby('Operation_Date_Only').agg(
            total_parts_picked=('PRODUCT CODE', 'count'),
            fifo_parts_picked=('is_fifo_part', 'sum')
        ).reset_index()
        daily_part_summary.rename(columns={'Operation_Date_Only': 'date'}, inplace=True)
        daily_part_summary['part_fifo%'] = (daily_part_summary['fifo_parts_picked'] / daily_part_summary['total_parts_picked'] * 100).astype(float)
        daily_part_summary['part_fifo%_fraction'] = (daily_part_summary['fifo_parts_picked'] / daily_part_summary['total_parts_picked']).astype(float)

        # Merge daily metrics into main table
        daily_metrics = pd.merge(
            daily_fifo[['date', 'fifo%', 'fifo%_fraction']],
            daily_part_summary[['date', 'total_parts_picked', 'fifo_parts_picked', 'part_fifo%', 'part_fifo%_fraction']],
            on='date', how='left'
        )
        master = pd.merge(df_group, daily_metrics, on='date', how='left')

        # Add lookup columns
        master['oldest_location'] = master['part_number'].map(location_lookup)
        master['oldest_module'] = master['part_number'].map(module_lookup)
        missing = master['oldest_module'].isna()
        if missing.any():
            master.loc[missing, 'oldest_module'] = 'UNKNOWN_' + master.loc[missing, 'part_number'].astype(str)

        # Reorder columns
        cols = [
            'date', 'fifo_picks', 'non_fifo_picks', 'total_picks',
            'fifo%', 'fifo%_fraction', 'part_number',
            'oldest_location', 'oldest_module', 'operator',
            'total_parts_picked', 'fifo_parts_picked', 'part_fifo%', 'part_fifo%_fraction'
        ]
        master = master[cols]
        master['date'] = pd.to_datetime(master['date']).dt.strftime('%Y-%m-%d')

        output_file = self.output_folder / 'FIFO_PowerBI_Export.xlsx'
        master.to_excel(output_file, index=False)
        print(f"  Master Export: {output_file}")

        # Cleanup old files
        old_files = ['FIFO_by_Operator.xlsx', 'FIFO_Daily_Picks.xlsx', 'FIFO_Daily_Totals.xlsx', 'FIFO_by_Part.xlsx']
        for f in old_files:
            old_path = self.output_folder / f
            if old_path.exists():
                old_path.unlink()

        print("\n=== Power BI Export Ready ===")


def main():
    analyzer = FIFOAnalyzer(
        data_folder='./DS FIFO Pick/Data/',
        output_folder='./DS FIFO Pick/Power BI Data/'
    )

    # Load and merge data
    df_pick, df_inv_loc, df_inv_combined = analyzer.load_and_merge_data()

    # Prepare dates and shifts first so we can split by day
    df_pick = analyzer._prepare_pick_data(df_pick)

    # Method 3: Day-by-day FIFO using per-day MODULE_LOC snapshots
    # Each pick day uses only that day's inventory snapshot (most realistic)
    import re
    module_loc_dir = analyzer.data_folder / 'MODULE_LOC Files'
    date_pattern = re.compile(r'_(\d{8})_\d{6}\.csv$')
    file_by_date = defaultdict(list)
    for f in module_loc_dir.glob('*.csv'):
        m = date_pattern.search(f.name)
        if m:
            file_date = dt.datetime.strptime(m.group(1), '%Y%m%d').date()
            file_by_date[file_date].append(f)
    available_dates = sorted(file_by_date.keys())

    print("\n=== Calculating FIFO Day-by-Day (per-day inventory) ===")
    inv_cols = ['MODULE#', 'PRODUCT', 'VANNING DATE', 'LOCATION', 'DAMAGE']
    day_frames = []
    day_inv_cache = {}

    for pick_date in sorted(df_pick['Operation_Date_Only'].unique()):
        # Find the best matching MODULE_LOC files for this pick date
        day_files = file_by_date.get(pick_date, [])
        if not day_files:
            # Fallback: use nearest past date's files
            past = [d for d in available_dates if d <= pick_date]
            day_files = file_by_date[past[-1]] if past else list(module_loc_dir.glob('*.csv'))

        # Build/cache per-day inventory
        day_key = tuple(sorted(str(f) for f in day_files))
        if day_key not in day_inv_cache:
            frames = [pd.read_csv(f, low_memory=False) for f in sorted(day_files, key=lambda p: p.stat().st_mtime)]
            df_day_base = pd.concat(frames, ignore_index=True)
            df_day_base = df_day_base.drop_duplicates(subset=['MODULE#'], keep='last')
            df_day_base = df_day_base[df_day_base['DAMAGE'] != 'Y'].copy()
            df_day_base = df_day_base[~df_day_base['LOCATION'].str.contains('-HLD-', na=False)].copy()
            df_day_base = df_day_base[~df_day_base['LOCATION'].str.contains('EC-QPC', na=False)].copy()
            df_day_base['PRODUCT'] = df_day_base['PRODUCT'].apply(analyzer.clean_part_number)
            df_day_base['VANNING DATE'] = df_day_base.apply(
                lambda row: analyzer.extract_vanning_date(row['MODULE#'], row.get('ETA')), axis=1
            )
            day_inv_cache[day_key] = df_day_base
            print(f"  {pick_date}: Built inventory from {len(day_files)} file(s) -> {len(df_day_base):,} modules")
        else:
            df_day_base = day_inv_cache[day_key]
            print(f"  {pick_date}: Reusing cached inventory -> {len(df_day_base):,} modules")

        day_picks = df_pick[df_pick['Operation_Date_Only'] == pick_date].copy()
        day_picks = day_picks.reset_index(drop=True)
        # Use only the raw MODULE_LOC inventory — don't add picked modules (Sachiko's method)
        day_picks = analyzer.check_fifo(day_picks, df_day_base)
        day_frames.append(day_picks)
    df_pick = pd.concat(day_frames, ignore_index=True)

    # Single-pass FIFO against combined inventory for part-level metrics
    df_pick['_row_id'] = range(len(df_pick))
    day_by_day_flags = dict(zip(df_pick['_row_id'], df_pick['fifo_compliant']))
    print("\n=== Calculating Part-Level FIFO (single-pass) ===")
    df_pick = df_pick.sort_values('PROCESS TIME').reset_index(drop=True)
    df_pick = analyzer.check_fifo(df_pick, df_inv_combined)
    df_pick['fifo_compliant_part'] = df_pick['fifo_compliant']
    df_pick['fifo_compliant'] = df_pick['_row_id'].map(day_by_day_flags)
    df_pick.drop(columns='_row_id', inplace=True)

    # Generate console report
    fifo_rate = analyzer.generate_report(df_pick)

    # Export KPI workbook
    analyzer.export_kpi_workbook(df_pick, df_inv_combined)

    # Export Excel files for Power BI
    analyzer.export_for_powerbi(df_pick, df_inv_combined)

if __name__ == "__main__":
    main()
